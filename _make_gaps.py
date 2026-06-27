# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = r"C:\Users\ksang\eclipse-workspace\BooksHQ_Web_Gap_Analysis_v2.xlsx"
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "State of the product"

HEAD = PatternFill("solid", fgColor="0E7C66")
SDONE = PatternFill("solid", fgColor="0B7A3B")   # green section
SBLK  = PatternFill("solid", fgColor="C0392B")   # red section
STODO = PatternFill("solid", fgColor="B45309")   # amber section
DONE = PatternFill("solid", fgColor="E6F4EA")
BLK  = PatternFill("solid", fgColor="FBE1E1")
TODO = PatternFill("solid", fgColor="FFF4E5")
white = Font(color="FFFFFF", bold=True, size=11)
swhite = Font(color="FFFFFF", bold=True, size=12)
thin = Side(style="thin", color="D9DEE7"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top"); center = Alignment(horizontal="center", vertical="top")

ws.merge_cells("A1:D1"); ws["A1"] = "Books HQ (Web) — State of the Product"; ws["A1"].font = Font(bold=True, size=16, color="0E7C66")
ws.merge_cells("A2:D2")
ws["A2"] = ("Refreshed 2026-06-23 after Phases 1-3.   ~24 gaps CLOSED.   8 BLOCKED on an external dependency or a structural prerequisite "
            "(a key / a service / a fresh company).   ~14 minor polish items REMAIN.")
ws["A2"].font = Font(italic=True, color="5A6B8B"); ws["A2"].alignment = Alignment(wrap_text=True, vertical="top"); ws.row_dimensions[2].height = 40

hdr = ["#", "Item", "Status", "Notes"]
HR = 4
for c, h in enumerate(hdr, 1):
    cell = ws.cell(row=HR, column=c, value=h); cell.fill = HEAD; cell.font = white; cell.border = border
    cell.alignment = Alignment(vertical="center")

D, B, T = "DONE", "BLOCKED", "TODO"
rows = [
    ("✅  DONE — built & verified across Phases 1-3", None, None),
    ("Ledger field-filtering per voucher type", D, "Each smart field restricted to the right ledgers (fixed the creditor->creditor trap)"),
    ("Smart guided forms — Contra / Debit Note / Credit Note", D, "Type bar + per-type labels + filtered fields"),
    ("Edit a posted voucher", D, "Edit form rewrites lines/date/narration via update_voucher"),
    ("Cancel a voucher (soft-delete)", D, "Cancel button + Day Book Edit link"),
    ("Bank Reconciliation — full matching workflow", D, "Upload -> auto-match -> create-voucher/manual-match/ignore -> finalise"),
    ("Ledger Reconciliation — full matching workflow", D, "Twin of bank reco, party ledgers + MIRROR/SAME sign mode"),
    ("GST Summary (+ HSN)", D, "Output vs input tax by rate, HSN table, net payable"),
    ("Bill-wise Outstanding", D, "Open bills aged by party; receivable/payable toggle"),
    ("Cash-Flow Planning", D, "Fortnightly projected cash position from open items"),
    ("Mileage Log", D, "Trip CRUD + miles x rate deductible (Schedule C)"),
    ("TDS Report", D, "Renders from engine (empty in this demo - no TDS data)"),
    ("Real Excel export (.xlsx) on every report + lists", D, "Server-built openpyxl workbook; Print/PDF via browser"),
    ("Preferences — Dr/Cr label style", D, "Natural / Traditional / Accounting, applied across screens"),
    ("License & Plan", D, "Live: plan, key, transactions used, seats, expiry"),
    ("Backup (download)", D, "One-click .db download (valid SQLite)"),
    ("Period Locks", D, "Lock date ranges; posting inside a lock is blocked"),
    ("Keyboard shortcuts", D, "Ctrl+Q launcher, Ctrl+1-9 jump, Alt+Left back, Ctrl+S post, Alt+C calculator, F2 add-ledger"),
    ("Dark mode + toggle", D, "Desktop dark palette; remembered per browser"),
    ("Deep-dive on every list/report", D, "Ledger rows -> statement; Day Book -> voucher detail"),
    ("Sort + filter on every table", D, "Click-to-sort + live row filter"),
    ("Inline error keeps your data", D, "Voucher posts via AJAX; validation errors don't wipe the form"),
    ("Company Settings / Users — editable", D, "Edit company; add/remove users"),

    ("🔴  BLOCKED — needs something external (not effort)", None, None),
    ("AI Documents Inbox", B, "Needs an Anthropic API key or wallet credits. UI can be wired; extraction won't run without it"),
    ("AI Verbal Entry", B, "Same Anthropic key requirement (and net-new in the codebase)"),
    ("AI Auto-Post", B, "Depends on the AI extraction path above"),
    ("AI 'fill from document' on Sales/Purchase", B, "Same Anthropic key requirement"),
    ("AI Wallet / credits top-up", B, "Needs the live licence server + Razorpay"),
    ("GST Returns filing (GSTR-1 / 3B / 2B)", B, "Needs a GSP integration (external). The GST *summary* is done"),
    ("Migration wizard", B, "Parsers exist, but migration needs a FRESH zero-voucher company; web is single-company (Sunrise has vouchers)"),
    ("Company switcher", B, "Web is single-company; needs multi-company support first"),

    ("🟡  REMAINING — minor / polish, doable anytime", None, None),
    ("TDS toggle on Payments", T, "Engine auto-deducts; no per-payment opt-in/out UI yet"),
    ("Multi-party voucher", T, "Engine has build_payment_multi; no web form yet"),
    ("Bill-wise allocation at posting", T, "Settle a receipt/payment against specific open bills (PRO)"),
    ("Per-line narration in journal grid", T, "No per-line note field on journal rows"),
    ("Date +/-1 steppers + Ctrl+Return add-row", T, "Small voucher-form conveniences"),
    ("Trial Balance balanced footer", T, "A 'balanced / diff' status line"),
    ("Grand totals on Day Book / Ledger Balances", T, "Summary total labels"),
    ("Balance Sheet Grouped/Flat toggle", T, "Only the grouped layout today"),
    ("Bank Book cleared column + voucher drill", T, "Reconciliation tick + click-to-voucher"),
    ("Date-format / sales-tax-rate / country prefs", T, "Extra preference fields"),
    ("User password change / enable-disable", T, "Add/remove works; no edit yet"),
    ("Backup RESTORE", T, "Download is done; restore-from-file not yet"),
    ("User Manual content", T, "Page exists; needs the actual help content/PDF"),
    ("Audit = full audit_log", T, "Today shows vouchers, not every logged action"),
    ("Feedback structured + system info", T, "Plain textarea today"),
]

r = HR + 1; n = 0
for col, w in {"A": 4, "B": 42, "C": 12, "D": 66}.items():
    ws.column_dimensions[col].width = w
for item, status, notes in rows:
    if status is None:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value=item)
        c.fill = SDONE if item.startswith("✅") else SBLK if item.startswith("🔴") else STODO
        c.font = swhite; c.border = border; c.alignment = Alignment(vertical="center")
    else:
        n += 1
        for ci, v in enumerate([n, item, status, notes], 1):
            c = ws.cell(row=r, column=ci, value=v); c.border = border
            c.alignment = center if ci in (1, 3) else wrap
        sc = ws.cell(row=r, column=3)
        sc.fill = DONE if status == D else BLK if status == B else TODO
        sc.font = Font(bold=True)
    r += 1

ws.freeze_panes = "A5"; ws.row_dimensions[HR].height = 20
wb.save(OUT); print("saved", OUT, "-", n, "items")
