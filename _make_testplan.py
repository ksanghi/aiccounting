# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = r"C:\Users\ksang\eclipse-workspace\BooksHQ_Web_Test_Plan_v4.xlsx"
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Test Plan"

HEAD = PatternFill("solid", fgColor="0E7C66")
SECT = PatternFill("solid", fgColor="DDF3EE")
VERI = PatternFill("solid", fgColor="E6F4EA")
RDY  = PatternFill("solid", fgColor="E3F0FA")
PEND = PatternFill("solid", fgColor="FFF4E5")
WIRE = PatternFill("solid", fgColor="EFEAF7")
NEWU = PatternFill("solid", fgColor="E0F4F3")
NEWF = PatternFill("solid", fgColor="FFE9C7")   # "new since v3" marker
white = Font(color="FFFFFF", bold=True, size=11)
thin = Side(style="thin", color="D9DEE7"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top"); center = Alignment(horizontal="center", vertical="top")

ws.merge_cells("A1:H1"); ws["A1"] = "Books HQ (Web) — Test Plan  (v4)"; ws["A1"].font = Font(bold=True, size=16, color="0E7C66")
ws.merge_cells("A2:H2")
ws["A2"] = ("App: http://127.0.0.1:8800/  |  Company: Sunrise Traders  |  v4, 2026-06-23.   "
            "Since v3: a full element-by-element parity pass vs the desktop (driven by the _element_diff.py tool) + the date-format preference. "
            "★ in the 'New' column = added/changed since v3 — test these first.   "
            "KIND: 'Wiring' = same proven core engine, just confirm inputs/outputs; 'New UI' = new web behaviour, test thoroughly.   "
            "For reconciliation/restore tests, keep a small CSV (Date,Narration,Debit,Credit) and a downloaded .db handy.")
ws["A2"].font = Font(italic=True, color="5A6B8B"); ws["A2"].alignment = Alignment(wrap_text=True, vertical="top"); ws.row_dimensions[2].height = 76

hdr = ["#", "Feature", "Kind", "New", "Steps", "Expected result", "Status", "Pass / Fail"]
HR = 4
for c, h in enumerate(hdr, 1):
    cell = ws.cell(row=HR, column=c, value=h); cell.fill = HEAD; cell.font = white; cell.border = border
    cell.alignment = Alignment(vertical="center")

W, U = "Wiring", "New UI"
S = (None, None, None, None, None)
# (feature, kind, new?, steps, expected, status)
rows = [
    ("NAVIGATION & LOOK", *S),
    ("Tile launcher", U, "", "Go to the app URL; type a few letters of a screen name", "Tiles filter live; clicking one opens that screen", "Verified"),
    ("Ctrl+Q launcher", U, "", "Press Ctrl+Q on any screen", "The tile launcher opens", "Verified"),
    ("Ctrl+1..9 jump", U, "", "Press Ctrl+3", "Jumps to the 3rd sidebar screen", "Verified"),
    ("Dark mode", U, "", "Click the ☾ toggle bottom-left; reload", "Flips to dark, readable, remembered", "Verified"),

    ("POST VOUCHER", *S),
    ("Field filtering", U, "", "Payment -> open 'Paid from'; Income -> 'Source of Income'", "'Paid from' = banks/Cash only; income box = income ledgers only", "Verified"),
    ("Post a Payment", W, "", "Payment -> creditor, from Cash, 500 -> Post", "Posted; in Day Book; Cash Book drops 500", "Ready"),
    ("Post Sales + GST", W, "", "Income -> sales, a customer, 1000, GST 18% -> Post", "Party Dr 1180, sales Cr 1000, CGST 90, SGST 90", "Verified"),
    ("F2 inline new ledger", U, "", "Click + beside a ledger -> name + group -> Create", "Created and selected in place, no reload", "Verified"),
    ("Calculator (Alt+C)", U, "", "Alt+C -> 1200*3= -> Use in field; type 5000+250 in an amount", "3600 drops in; 5000+250 -> 5250", "Ready"),
    ("Ctrl+S to post", U, "", "Fill a voucher, press Ctrl+S", "The voucher posts", "Ready"),
    ("Error keeps your data", U, "", "Post a Journal where Debit != Credit", "Red error on the SAME form; nothing lost", "Verified"),
    ("Edit a voucher", W, "", "Day Book -> Edit a row -> change amount -> Save", "Rewritten; shows in Day Book + reports", "Ready"),
    ("Cancel / Delete a voucher", W, "★", "Day Book -> Delete on a row (or a voucher -> Cancel)", "Soft-deleted; gone from Day Book, kept in audit", "Ready"),

    ("DAY BOOK & REPORTS", *S),
    ("Day Book — KPI strip + totals", U, "★", "Open Day Book", "4 KPI tiles (Vouchers / Receipts / Payments / Net) + a Delete link per row + a totals line", "Verified"),
    ("Trial Balance — full columns", W, "★", "Open Trial Balance", "Group · Nature · Op Dr/Cr · Txn Dr/Cr · Cl Dr/Cr columns + a green 'Balanced' footer", "Verified"),
    ("Balance Sheet — Grouped/Flat", U, "★", "Balance Sheet -> flip the View dropdown", "Flat view shows Ledger · Group · Side · Amount; Grouped shows the two-panel layout", "Verified"),
    ("GST — Net column", W, "★", "Menu -> GST", "Tax table now has a Net column (output - input) + Net payable", "Verified"),
    ("Bill-wise — group dropdown", U, "★", "Bill-wise -> change the Group dropdown", "Switches Receivable (Debtors) / Payable (Creditors)", "Verified"),
    ("Cash-Flow — open items", W, "★", "Menu -> Cash-Flow", "Forecast table + an 'Open items driving the forecast' table (Direction/Party/Outstanding)", "Verified"),
    ("TDS — Register", W, "★", "Menu -> TDS Report", "A By-section table AND a party-wise TDS Register table", "Verified"),
    ("Ledger — New / Edit ledger", W, "★", "Open a ledger statement -> ＋ New ledger / ✎ Edit ledger", "Create a ledger / rename-regroup the current one; change persists", "Ready"),
    ("Deep-dive", U, "", "Click a ledger name in a report; a voucher number in Day Book", "Opens the ledger statement / voucher detail", "Verified"),
    ("Sort + filter", U, "", "Click a column header; type in 'Filter rows'", "Sorts asc/desc; filters live", "Verified"),
    ("Excel + Print/PDF", U, "", "Any report -> Excel / Print", "Real .xlsx downloads; clean printable view", "Verified"),
    ("Mileage Log", U, "", "Mileage -> add a trip -> Remove", "Trip + miles x $0.70 deductible; remove works", "Verified"),

    ("RECONCILIATION", *S),
    ("Bank Reconciliation", U, "", "Bank Reco -> account -> upload CSV -> Import & auto-match", "Review page; some lines auto-match; Matched/Unmatched tiles", "Verified"),
    ("Bank Reco - create / match / finalise", W, "", "On Review: create a voucher, manually match, then Finalise", "Lines clear; reconciliation snapshots to history", "Ready"),
    ("Ledger Reconciliation", U, "", "Ledger Reco -> a party -> MIRROR -> upload CSV -> auto-match", "Same review flow for the party ledger", "Verified"),

    ("SETTINGS & ADMIN", *S),
    ("Preferences — Dr/Cr labels", U, "", "Preferences -> Accounting -> Save -> Trial Balance", "Dr/Cr or By/To applied to voucher/ledger surfaces", "Verified"),
    ("Preferences — Date format", U, "★", "Preferences -> Date format -> dd/MM/yyyy -> Save -> open Day Book; then dd-MMM-yyyy", "Every date in reports/tables switches format (23/06/2026 <-> 23-Jun-2026)", "Verified"),
    ("Backup & Restore", U, "★", "Backup -> 💾 Backup now (downloads .db); ↩ Restore -> upload a .db", "Backup downloads; restore replaces the books from the file", "Ready"),
    ("Feedback — Bug / Feature", U, "★", "Feedback -> Type dropdown (Bug/Feature/General) + Submit / Clear", "Type selector present; Submit records; Clear empties", "Verified"),
    ("User Manual", U, "★", "Menu -> User Manual", "Real help content (vouchers, reports, reco, backup, shortcuts) + Download/print", "Verified"),
    ("Company Settings (edit)", W, "", "Company Settings -> change name/address -> Save -> reload", "Values persist", "Ready"),
    ("Users (add/remove)", U, "", "Users -> Add user -> Remove", "Row appears, then removed", "Ready"),
    ("License & Plan", W, "", "Menu -> License", "Plan PRO, key, transactions used, seats, expiry", "Verified"),
    ("Period Locks", W, "", "Period Locks -> lock a range -> try to post inside it", "The post is blocked: 'Period ... is locked'", "Verified"),

    ("NOT TESTABLE YET (honest)", *S),
    ("AI Documents Inbox / Verbal / Auto-Post", "—", "", "—", "Needs an Anthropic API key / credits", "Pending"),
    ("Migration wizard", "—", "", "—", "Needs a fresh zero-voucher company (web is single-company)", "Pending"),
]

r = HR + 1; n = 0
for col, w in {"A": 4, "B": 25, "C": 8, "D": 5, "E": 42, "F": 44, "G": 10, "H": 11}.items():
    ws.column_dimensions[col].width = w
for feature, kind, new, steps, expected, status in rows:
    if steps is None:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = ws.cell(row=r, column=1, value=feature); c.fill = SECT; c.font = Font(bold=True, color="0E7C66", size=12)
        c.border = border; c.alignment = Alignment(vertical="center")
    else:
        n += 1
        for ci, v in enumerate([n, feature, kind, new, steps, expected, status, ""], 1):
            c = ws.cell(row=r, column=ci, value=v); c.border = border
            c.alignment = center if ci in (1, 3, 4, 7) else wrap
        kc = ws.cell(row=r, column=3)
        kc.fill = WIRE if kind == W else NEWU if kind == U else PEND
        kc.font = Font(bold=True, size=9)
        if new:
            nc = ws.cell(row=r, column=4); nc.fill = NEWF; nc.font = Font(bold=True)
        sc = ws.cell(row=r, column=7); st = status.lower()
        sc.fill = VERI if st == "verified" else RDY if st == "ready" else PEND
        sc.font = Font(bold=True)
    r += 1

ws.freeze_panes = "A5"; ws.row_dimensions[HR].height = 20
r += 1
ws.cell(row=r, column=2, value="★ = new/changed since v3").font = Font(bold=True)
ws.cell(row=r, column=5, value="Wiring").fill = WIRE
ws.cell(row=r, column=6, value="= confirm the web fed/showed the proven core engine right").font = Font(size=9)
r += 1
ws.cell(row=r, column=5, value="New UI").fill = NEWU
ws.cell(row=r, column=6, value="= genuinely new web behaviour (test thoroughly)").font = Font(size=9)
wb.save(OUT); print("saved", OUT, "-", n, "test cases")
