"""
Bake operator-config Excel files into Python constants.

Reads:
    config/ai_features.xlsx
    config/pricing.xlsx

Writes:
    core/_baked_config.py
    license_server/_baked_config.py     (identical content; license server
                                         is a separate deployable, must not
                                         import from core.)

Run this AFTER editing the .xlsx files. The generated .py files are
checked into git — that way the running app never needs openpyxl, and a
diff in git review shows exactly what changed in the operator config.

Usage:
    python build/bake_config.py

The build/build.bat installer pipeline should run this as its first step,
so a release build always picks up the latest .xlsx values.
"""
from __future__ import annotations

import datetime as _dt
import sys
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parent.parent
CONFIG_DIR = REPO_ROOT / "config"
CORE_OUT = REPO_ROOT / "core" / "_baked_config.py"
SERVER_OUT = REPO_ROOT / "license_server" / "_baked_config.py"


# ── Excel readers ───────────────────────────────────────────────────────────

def _rows(ws, header_row: int):
    """Yield dicts of {header: cell_value} for every data row below the
    header, stopping at the first row where the first column is blank."""
    headers = [ws.cell(row=header_row, column=c).value
               for c in range(1, ws.max_column + 1)]
    headers = [h for h in headers if h is not None]
    ncols = len(headers)
    for r in range(header_row + 1, ws.max_row + 1):
        first = ws.cell(row=r, column=1).value
        if first is None or (isinstance(first, str) and not first.strip()):
            continue
        row = {}
        for c, name in enumerate(headers, start=1):
            row[name] = ws.cell(row=r, column=c).value
        yield row


def read_ai_features(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb["Features"]
    out = {}
    for row in _rows(ws, header_row=4):
        fid = (row.get("feature_id") or "").strip()
        cls = (row.get("class") or "").strip()
        if not fid:
            continue
        if cls not in ("byok", "ag_key"):
            raise ValueError(
                f"ai_features.xlsx: feature '{fid}' has invalid class "
                f"'{cls}' (must be byok or ag_key)"
            )
        out[fid] = cls
    if not out:
        raise ValueError("ai_features.xlsx: no features found")
    return out


def read_pricing(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)

    # Accounts HQ — unified `ACCOUNTSHQ` sheet (one row per attribute/feature,
    # one column per tier; same layout as RWAHQ/HOAHQ). Replaces the old split
    # `Tiers` + `PlanFeatures` tabs. Per-country prices still live on `Countries`.
    ws = wb["ACCOUNTSHQ"]
    rows = list(_rows(ws, header_row=3))
    headers = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
    reserved = {"row_type", "feature", "id", "upgrade_to", "notes"}
    tier_codes = [h for h in headers if h is not None and h not in reserved]

    def _attr_row(rid):
        for r in rows:
            if (str(r.get("id") or "")).strip() == rid:
                return r
        return {}

    name_r  = _attr_row("tier_name")
    seats_r = _attr_row("seats")
    txn_r   = _attr_row("txn_limit")
    over_r  = _attr_row("overage_rate")
    price_r = _attr_row("price_annual_INR")
    notes_r = _attr_row("tier_notes")

    tiers = []
    for code in tier_codes:
        nm = name_r.get(code)
        nt = notes_r.get(code)
        tiers.append({
            "code":           code,
            "name":           (str(nm).strip() if nm not in (None, "") else code),
            "seats_allowed":  _as_int(seats_r.get(code), default=1),
            "txn_limit":      _as_int(txn_r.get(code), default=0),
            "overage_rate":   _as_float(over_r.get(code), default=0.0),
            "plan_price_INR": _as_float(price_r.get(code), default=0.0),
            "notes":          (str(nt).strip() if nt not in (None, "") else ""),
        })
    if not tiers:
        raise ValueError("pricing.xlsx ACCOUNTSHQ: no tiers found")

    # PlanFeatures matrix — the `feature` rows of the same sheet
    plan_features: dict[str, list[str]] = {c: [] for c in tier_codes}
    feature_upgrade_map: dict[str, str] = {}

    for row in rows:
        if (str(row.get("row_type") or "")).strip() != "feature":
            continue
        fid = (str(row.get("id") or "")).strip()
        if not fid:
            continue
        for code in tier_codes:
            val = row.get(code)
            if isinstance(val, str) and val.strip().upper() in ("Y", "YES", "TRUE", "1", "X"):
                plan_features[code].append(fid)
            elif val is True:
                plan_features[code].append(fid)
        up = (row.get("upgrade_to") or "")
        if isinstance(up, str) and up.strip():
            up = up.strip().upper()
            if up in tier_codes:
                feature_upgrade_map[fid] = up

    # Books HQ shares the `accgenie` product + the same plan tiers, differing
    # only by country (A13). Fold its US-EXCLUSIVE feature flags (sales_tax,
    # form_1099, schedule_c, mileage_tracking, …) into the SAME plan_features
    # map so has_feature() gates them by tier; the CountryProfile gates
    # visibility. Only ids NOT already defined in ACCOUNTSHQ are added — this
    # never loosens an existing India gate.
    if "BOOKSHQ" in wb.sheetnames:
        existing_fids = {
            (str(r.get("id") or "")).strip()
            for r in rows
            if (str(r.get("row_type") or "")).strip() == "feature"
        }
        for row in _rows(wb["BOOKSHQ"], header_row=3):
            if (str(row.get("row_type") or "")).strip() != "feature":
                continue
            fid = (str(row.get("id") or "")).strip()
            if not fid or fid in existing_fids:
                continue
            for code in tier_codes:
                val = row.get(code)
                if (isinstance(val, str) and val.strip().upper() in ("Y", "YES", "TRUE", "1", "X")) or val is True:
                    plan_features[code].append(fid)
            up = (row.get("upgrade_to") or "")
            if isinstance(up, str) and up.strip().upper() in tier_codes:
                feature_upgrade_map[fid] = up.strip().upper()

    # Countries
    ws_c = wb["Countries"]
    countries = []
    for row in _rows(ws_c, header_row=4):
        cc = (row.get("country_code") or "").strip()
        if not cc:
            continue
        tier_prices = {}
        for code in tier_codes:
            tier_prices[code] = _as_float_or_none(row.get(f"price_{code}"))
        active_raw = row.get("active")
        is_active = (
            (isinstance(active_raw, str) and active_raw.strip().upper() in ("Y", "YES", "TRUE", "1"))
            or active_raw is True
        )
        countries.append({
            "country_code":    cc.upper(),
            "country_name":    (row.get("country_name") or cc).strip(),
            "currency_code":   (row.get("currency_code") or "").strip(),
            "currency_symbol": (row.get("currency_symbol") or "").strip(),
            "tier_prices":     tier_prices,
            "active":          is_active,
            "notes":           (row.get("notes") or "").strip(),
        })

    return {
        "tiers": tiers,
        "plan_features": plan_features,
        "feature_upgrade_map": feature_upgrade_map,
        "countries": countries,
    }


def read_rwa(path: Path) -> dict:
    """Read the unified `RWAHQ` sheet — one row per attribute/feature,
    one column per tier. `row_type` says how to read each row:
      price / limit / rate / quota -> numeric per tier
      feature                       -> Y/blank per tier (+ upgrade_to)
    DEMO is derived (desktop trial = mirrors PREMIUM features; server
    never issues DEMO). This replaces the old hand-maintained dicts in
    license_bridge.py + plans.py."""
    wb = load_workbook(path, data_only=True)
    ws = wb["RWAHQ"]
    rows = list(_rows(ws, header_row=3))
    tiers = ["FREE", "STANDARD", "PRO", "PREMIUM"]

    def _row(rid):
        for r in rows:
            if (str(r.get("id") or "")).strip() == rid:
                return r
        return None

    def _nummap(rid, asfloat=False, blank_none=False):
        r = _row(rid)
        out = {}
        for t in tiers:
            v = None if r is None else r.get(t)
            if v is None or v == "":
                out[t] = None if blank_none else (0.0 if asfloat else 0)
            else:
                out[t] = _as_float(v) if asfloat else _as_int(v)
        return out

    features: dict[str, list[str]] = {t: [] for t in tiers}
    upgrade_map: dict[str, str] = {}
    for r in rows:
        if (str(r.get("row_type") or "")).strip() != "feature":
            continue
        fid = (str(r.get("id") or "")).strip()
        if not fid:
            continue
        for t in tiers:
            v = r.get(t)
            if (isinstance(v, str) and v.strip().upper() in ("Y", "YES", "TRUE", "1", "X")) or v is True:
                features[t].append(fid)
        up = str(r.get("upgrade_to") or "").strip().upper()
        if up in tiers:
            upgrade_map[fid] = up

    features["DEMO"] = list(features["PREMIUM"])  # desktop trial mirrors PREMIUM
    txn = _nummap("txn_limit");                 txn["DEMO"] = 50
    flats = _nummap("flats", blank_none=True);  flats["DEMO"] = None
    seats = _nummap("seats");                   seats["DEMO"] = seats["PREMIUM"]
    overage = _nummap("overage_rate", asfloat=True); overage["DEMO"] = 0.0
    notices = _nummap("notices_per_month", blank_none=True); notices["DEMO"] = None
    promos = _nummap("free_promos_per_month"); promos["DEMO"] = promos["PREMIUM"]
    p_annual = _nummap("price_annual_INR");     p_annual["DEMO"] = 0
    p_monthly = _nummap("price_monthly_INR");   p_monthly["DEMO"] = 0

    return {
        "features": features, "upgrade_map": upgrade_map,
        "txn_limit": txn, "flats": flats, "seats": seats, "overage": overage,
        "notices": notices, "promos": promos,
        "price_annual": p_annual, "price_monthly": p_monthly,
    }


def _as_int(v, default=0):
    if v is None or v == "":
        return default
    try:
        return int(v)
    except (TypeError, ValueError):
        return default


def _as_float(v, default=0.0):
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _as_float_or_none(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# ── Python module writer ────────────────────────────────────────────────────

def _py(val, indent=0):
    """Format a value as a Python literal with stable indentation."""
    pad = "    " * indent
    if val is None or isinstance(val, (int, float, bool, str)):
        return repr(val)
    if isinstance(val, list):
        if not val:
            return "[]"
        inner = ",\n".join(pad + "    " + _py(x, indent + 1) for x in val)
        return "[\n" + inner + ",\n" + pad + "]"
    if isinstance(val, dict):
        if not val:
            return "{}"
        items = []
        for k, v in val.items():
            items.append(pad + "    " + repr(k) + ": " + _py(v, indent + 1))
        return "{\n" + ",\n".join(items) + ",\n" + pad + "}"
    raise TypeError(f"Cannot format {type(val)}")


def render_module(ai_features: dict, pricing: dict, rwa: dict) -> str:
    tiers = pricing["tiers"]
    plan_features = pricing["plan_features"]
    upgrade_map = pricing["feature_upgrade_map"]
    countries = pricing["countries"]

    # Derived convenience constants — exposed so consumers don't need to
    # iterate TIERS just to get a single dict.
    plan_codes = [t["code"] for t in tiers]
    plan_limits = {t["code"]: t["txn_limit"] for t in tiers}
    plan_seats = {t["code"]: t["seats_allowed"] for t in tiers}
    plan_user_limits = {
        t["code"]: 999 if "multi_user_unlimited" in plan_features.get(t["code"], [])
        else (5 if "multi_user_5" in plan_features.get(t["code"], [])
              else (2 if "multi_user_2" in plan_features.get(t["code"], []) else 1))
        for t in tiers
    }
    overage_rates = {t["code"]: t["overage_rate"] for t in tiers}
    plan_prices = {t["code"]: t["plan_price_INR"] for t in tiers}

    header = (
        '"""\n'
        'AUTO-GENERATED by build/bake_config.py from:\n'
        '    config/ai_features.xlsx\n'
        '    config/pricing.xlsx\n'
        '\n'
        'DO NOT EDIT BY HAND. Edit the .xlsx files and re-run the baker.\n'
        f'Generated: {_dt.datetime.now().isoformat(timespec="seconds")}\n'
        '"""\n'
        'from __future__ import annotations\n\n'
    )

    body = (
        "AI_FEATURES = " + _py(ai_features) + "\n\n"
        "TIERS = " + _py(tiers) + "\n\n"
        "PLANS = " + _py(plan_codes) + "\n\n"
        "PLAN_FEATURES = " + _py(plan_features) + "\n\n"
        "FEATURE_UPGRADE_MAP = " + _py(upgrade_map) + "\n\n"
        "PLAN_LIMITS = " + _py(plan_limits) + "\n\n"
        "PLAN_SEATS = " + _py(plan_seats) + "\n\n"
        "PLAN_USER_LIMITS = " + _py(plan_user_limits) + "\n\n"
        "OVERAGE_RATES = " + _py(overage_rates) + "\n\n"
        "PLAN_PRICES = " + _py(plan_prices) + "\n\n"
        "COUNTRIES = " + _py(countries) + "\n"
    )

    # ── RWA HQ (rwagenie) — baked from the unified RWAHQ sheet ──
    # Replaces the old hand-maintained dicts in rwagenie/app/license_bridge.py
    # and license_server/plans.py (which now import these).
    rwa_body = (
        "\n\n# ── RWA HQ (rwagenie) — from the RWAHQ sheet ──\n"
        "PLAN_FEATURES_RWA = " + _py(rwa["features"]) + "\n\n"
        "FEATURE_UPGRADE_MAP_RWA = " + _py(rwa["upgrade_map"]) + "\n\n"
        "PLAN_LIMITS_RWA = " + _py(rwa["txn_limit"]) + "\n\n"
        "PLAN_FLATS_LIMIT_RWA = " + _py(rwa["flats"]) + "\n\n"
        "PLAN_SEATS_RWA = " + _py(rwa["seats"]) + "\n\n"
        "OVERAGE_RATES_RWA = " + _py(rwa["overage"]) + "\n\n"
        "RWA_NOTICES_QUOTA = " + _py(rwa["notices"]) + "\n\n"
        "RWA_PROMOS_QUOTA = " + _py(rwa["promos"]) + "\n\n"
        "PLAN_PRICES_RWA_INR = " + _py(rwa["price_annual"]) + "\n\n"
        "PLAN_PRICES_RWA_MONTHLY_INR = " + _py(rwa["price_monthly"]) + "\n"
    )
    return header + body + rwa_body


def main():
    ai_path = CONFIG_DIR / "ai_features.xlsx"
    pr_path = CONFIG_DIR / "pricing.xlsx"
    if not ai_path.exists():
        sys.exit(f"Missing {ai_path}. Run build/make_config_xlsx.py first.")
    if not pr_path.exists():
        sys.exit(f"Missing {pr_path}. Run build/make_config_xlsx.py first.")

    ai_features = read_ai_features(ai_path)
    pricing = read_pricing(pr_path)
    rwa = read_rwa(pr_path)
    text = render_module(ai_features, pricing, rwa)

    CORE_OUT.write_text(text, encoding="utf-8")
    SERVER_OUT.write_text(text, encoding="utf-8")
    print(f"Wrote {CORE_OUT}")
    print(f"Wrote {SERVER_OUT}")
    print(f"  AI features: {len(ai_features)}")
    print(f"  Tiers:       {[t['code'] for t in pricing['tiers']]}")
    print(f"  Countries:   {[c['country_code'] for c in pricing['countries']]}")
    print(f"  RWA features/tier: { {k: len(v) for k, v in rwa['features'].items()} }")


if __name__ == "__main__":
    main()
