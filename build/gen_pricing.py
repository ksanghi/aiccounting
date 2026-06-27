"""
Single-source pricing generator.

Reads config/pricing.xlsx and regenerates, from the sheet only:
  - the pricing-page CARDS in marketing-aic/pricing.html   (price, tier_tag headline, tier_bullets)
  - the checkout CATALOG plans in marketing-aic/checkout.html (price, tier_tag, tier_bullets)
for the products in PRODUCTS. Everything else (hero, tradeHQ, styles, footer)
is left untouched.

Sheet is the one place to edit. Workflow (manual publish gate):
edit sheet -> python build/gen_pricing.py -> review on staging -> deploy.
Idempotent. Reads the sheet even if it's open in Excel (copies it first).
"""
from __future__ import annotations
import re, html, json, shutil, tempfile, os
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "config" / "pricing.xlsx"
PRICING = ROOT / "marketing-aic" / "pricing.html"
CHECKOUT = ROOT / "marketing-aic" / "checkout.html"

# product (checkout slug) -> (sheet tab, pricing-page section comment)
PRODUCTS = {
    "accgenie": ("ACCOUNTSHQ", "Accounts HQ pricing"),
    "rwagenie": ("RWAHQ",      "RWA HQ pricing"),
}
TIER_ORDER = ["FREE", "STANDARD", "PRO", "PREMIUM"]


def _load():
    try:
        return openpyxl.load_workbook(XLSX, data_only=True)
    except PermissionError:
        tmp = Path(tempfile.gettempdir()) / "_gen_pricing_read.xlsx"
        shutil.copy(XLSX, tmp)  # copy works even while Excel holds the file
        return openpyxl.load_workbook(tmp, data_only=True)


def _rows(ws):
    hdr = next(i for i, r in enumerate(ws.iter_rows(values_only=True), 1)
               if r and str(r[0]).strip() == "row_type")
    cols = {ws.cell(hdr, c).value: c for c in range(4, 9) if ws.cell(hdr, c).value}
    data = {}
    for ri in range(1, ws.max_row + 1):
        rid = ws.cell(ri, 3).value
        if rid:
            data[str(rid).strip()] = {t: ws.cell(ri, c).value for t, c in cols.items()}
    return data


def _bullets(raw):
    return [b.strip() for b in str(raw or "").split("|") if b.strip()]


def _money_html(v):
    v = int(v or 0)
    return "Rs. 0" if v == 0 else f'Rs. {v:,}<span class="small">/yr</span>'


# ── pricing-page cards ────────────────────────────────────────────────
def _card(prod, code, name, price, tag, bullets):
    feat = " featured" if code == "STANDARD" else ""
    cta = "Get free" if code == "FREE" else f"Buy {name}"
    lis = "\n".join(f"          <li>{html.escape(b)}</li>" for b in bullets)
    return (
        f'      <div class="price-card{feat}">\n'
        f'        <div class="price-name">{html.escape(name)}</div>\n'
        f'        <div class="price-amt">{_money_html(price)}</div>\n'
        f'        <div class="price-tag">{html.escape(str(tag or ""))}</div>\n'
        f'        <ul class="price-feats">\n{lis}\n        </ul>\n'
        f'        <a class="price-btn" href="checkout.html?product={prod}&amp;plan={code}">{cta}</a>\n'
        f'      </div>'
    )


def _grid(prod, d):
    n, p, t, b = (d.get("tier_name", {}), d.get("price_annual_INR", {}),
                  d.get("tier_tag", {}), d.get("tier_bullets", {}))
    return "\n\n".join(_card(prod, c, n.get(c) or c.title(), p.get(c), t.get(c), _bullets(b.get(c)))
                       for c in TIER_ORDER)


# ── checkout CATALOG plans ────────────────────────────────────────────
def _checkout_plans(d):
    n, p, t, b = (d.get("tier_name", {}), d.get("price_annual_INR", {}),
                  d.get("tier_tag", {}), d.get("tier_bullets", {}))
    out = []
    for c in TIER_ORDER:
        name = n.get(c) or c.title()
        out.append(
            f'      {{ code: "{c}", inr: {int(p.get(c) or 0)}, name: "{name}", '
            f'tag: {json.dumps(str(t.get(c) or ""))},\n'
            f'        bullets: {json.dumps(_bullets(b.get(c)))} }},'
        )
    return "    plans: [\n" + "\n".join(out) + "\n    ],"


def main():
    wb = _load()
    data = {prod: _rows(wb[tab]) for prod, (tab, _) in PRODUCTS.items()}

    # 1) pricing page
    text = PRICING.read_text(encoding="utf-8")
    grid_re = re.compile(r'(<div class="pricing-grid"[^>]*>)(.*?)(\s*</div>\s*</div>\s*</section>)', re.S)
    for prod, (tab, marker) in PRODUCTS.items():
        m = re.search(re.escape(marker), text)
        if not m:
            print("!! pricing marker missing:", prod); continue
        s = m.end(); nxt = re.search(r'<!-- ── ', text[s:])
        e = s + nxt.start() if nxt else len(text)
        seg, n = grid_re.subn(lambda g: g.group(1) + "\n" + _grid(prod, data[prod]) + "\n    " + g.group(3).lstrip(),
                              text[s:e], count=1)
        if n != 1: print("!! pricing grid not matched:", prod); continue
        text = text[:s] + seg + text[e:]
        print("   pricing.html:", prod)
    PRICING.write_text(text, encoding="utf-8")

    # 2) checkout CATALOG plans
    ck = CHECKOUT.read_text(encoding="utf-8")
    for prod in PRODUCTS:
        # replace the plans array inside this product's CATALOG entry
        pat = re.compile(r'(\b' + prod + r':\s*\{.*?)\n    plans:\s*\[.*?\n    \],', re.S)
        def repl(g):
            return g.group(1) + "\n" + _checkout_plans(data[prod])
        ck, n = pat.subn(repl, ck, count=1)
        print("   checkout.html:", prod, "OK" if n == 1 else "!! NOT MATCHED")
    CHECKOUT.write_text(ck, encoding="utf-8")
    print("done.")


if __name__ == "__main__":
    main()
