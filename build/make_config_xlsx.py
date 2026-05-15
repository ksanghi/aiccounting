"""
Create the operator-config Excel templates: config/ai_features.xlsx and
config/pricing.xlsx.

These files are the SOURCE OF TRUTH for AccGenie's per-version operator
config (AI feature classes, tier definitions, plan features, country
pricing). The build pipeline runs build/bake_config.py after these are
edited; that step reads the xlsx files and writes core/_baked_config.py
and license_server/_baked_config.py (plain Python dicts compiled into
the binary).

This script is idempotent — re-running it OVERWRITES the templates back
to their seeded values. Run it once to create the templates; after that,
edit the .xlsx files directly. Only re-run if you want to reset to seed.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parent.parent
CONFIG_DIR = REPO_ROOT / "config"


# ── Seed data ───────────────────────────────────────────────────────────────
#
# Change values here only if you want to reset to a new "factory default".
# Normal edits should be made in the .xlsx files directly.

AI_FEATURES_SEED = [
    # feature_id, class, notes
    ("document_recognition", "byok",   "Reads PDFs/images. Heavy multi-call AI — customer's own key required."),
    ("bank_statement_ai",    "ag_key", "Parses bank statements. Light/structured — AccGenie key + wallet."),
    ("ledger_statement_ai",  "ag_key", "Parses party ledger PDFs."),
    ("sales_ai_fill",        "ag_key", "Suggests sales-voucher line items from a description."),
    ("purchase_ai_fill",     "ag_key", "Suggests purchase-voucher line items."),
    ("ledger_suggest",       "ag_key", "Suggests likely Dr/Cr ledgers for a voucher row."),
    ("verbal_entry",         "ag_key", "Voice-to-voucher one-liner posting."),
]

TIERS_SEED = [
    # code, name, seats_allowed, txn_limit, overage_rate, plan_price_INR, notes
    ("DEMO",     "Demo",     1,    50,       0.00, 0,    "Full-feature trial, hard-capped at txn_limit. Mirrors PREMIUM features."),
    ("FREE",     "Free",     1,    5_000,    0.00, 0,    "Permanent free tier — basic features only."),
    ("STANDARD", "Standard", 2,    20_000,   0.30, 1999, "Reports, exports, reconciliation, 2 users."),
    ("PRO",      "Pro",      5,    50_000,   0.30, 4999, "GST, TDS, AI document reader, 5 users."),
    ("PREMIUM",  "Premium",  10,   100_000,  0.20, 9999, "All features incl. WhatsApp, audit export, API."),
]

# All known feature ids (rows in PlanFeatures sheet) — grouped by category.
# Each entry: (category, feature_id). Category is for visual grouping only.
ALL_FEATURES_V2 = [
    ("Core",           "vouchers"),
    ("Core",           "daybook"),
    ("Core",           "ledger_balances"),
    ("Core",           "backup"),
    ("Users",          "multi_user_2"),
    ("Users",          "multi_user_5"),
    ("Users",          "multi_user_unlimited"),
    ("Reports",        "trial_balance"),
    ("Reports",        "profit_loss"),
    ("Reports",        "balance_sheet"),
    ("Reports",        "cash_book"),
    ("Reports",        "bank_book"),
    ("Reports",        "ledger_account"),
    ("Reports",        "receipts_payments"),
    ("Output",         "export_excel"),
    ("Output",         "export_pdf"),
    ("Output",         "print"),
    ("Output",         "email_report"),
    ("Reconciliation", "bank_reconciliation"),
    ("Reconciliation", "ledger_reconciliation"),
    ("Migration",      "book_migration"),
    ("Tax",            "gst"),
    ("Tax",            "tds"),
    ("AI",             "ai_document_reader"),
    ("AI",             "verbal_entry"),
    ("AI",             "auto_billing"),
    ("Enterprise",     "whatsapp"),
    ("Enterprise",     "audit_export"),
    ("Enterprise",     "api_access"),
    ("Enterprise",     "verticals"),
]
ALL_FEATURES = [fid for _, fid in ALL_FEATURES_V2]
ALL_REPORT_FEATURES = [
    fid for cat, fid in ALL_FEATURES_V2 if cat == "Reports"
]

# Which features each plan includes. Y = included, blank = not included.
PLAN_FEATURE_MATRIX = {
    "DEMO": {
        # DEMO mirrors PREMIUM (everything unlocked) but is capped at txn_limit.
        "vouchers", "daybook", "ledger_balances", "backup",
        "multi_user_unlimited",
        *ALL_REPORT_FEATURES,
        "export_excel", "export_pdf", "print", "email_report",
        "bank_reconciliation", "ledger_reconciliation", "book_migration",
        "gst", "tds",
        "ai_document_reader", "verbal_entry", "auto_billing",
        "whatsapp", "audit_export", "api_access", "verticals",
    },
    "FREE": {
        "vouchers", "daybook", "ledger_balances", "backup",
    },
    "STANDARD": {
        "vouchers", "daybook", "ledger_balances", "backup",
        "multi_user_2",
        *ALL_REPORT_FEATURES,
        "export_excel", "export_pdf", "print", "email_report",
        "bank_reconciliation", "ledger_reconciliation", "book_migration",
    },
    "PRO": {
        "vouchers", "daybook", "ledger_balances", "backup",
        "multi_user_5",
        *ALL_REPORT_FEATURES,
        "export_excel", "export_pdf", "print", "email_report",
        "bank_reconciliation", "ledger_reconciliation", "book_migration",
        "gst", "tds",
        "ai_document_reader", "verbal_entry", "auto_billing",
    },
    "PREMIUM": {
        "vouchers", "daybook", "ledger_balances", "backup",
        "multi_user_unlimited",
        *ALL_REPORT_FEATURES,
        "export_excel", "export_pdf", "print", "email_report",
        "bank_reconciliation", "ledger_reconciliation", "book_migration",
        "gst", "tds",
        "ai_document_reader", "verbal_entry", "auto_billing",
        "whatsapp", "audit_export", "api_access", "verticals",
    },
}

# Which tier a feature first appears in (drives the "Upgrade to ___" prompt
# when the user hits a locked feature). Read by the baker.
FEATURE_UPGRADE_MAP_SEED = {
    **{fid: "STANDARD" for fid in ALL_REPORT_FEATURES},
    "export_excel":          "STANDARD",
    "export_pdf":            "STANDARD",
    "print":                 "STANDARD",
    "email_report":          "STANDARD",
    "bank_reconciliation":   "STANDARD",
    "ledger_reconciliation": "STANDARD",
    "book_migration":        "STANDARD",
    "gst":                   "PRO",
    "tds":                   "PRO",
    "ai_document_reader":    "PRO",
    "verbal_entry":          "PRO",
    "auto_billing":          "PRO",
    "whatsapp":              "PREMIUM",
    "audit_export":          "PREMIUM",
    "api_access":            "PREMIUM",
    "verticals":             "PREMIUM",
}

COUNTRIES_SEED = [
    # country_code, country_name, currency_code, currency_symbol,
    # price_DEMO, price_FREE, price_STANDARD, price_PRO, price_PREMIUM,
    # ai_text_page_cost, ai_scanned_page_cost, ai_per_transaction_cost,
    # active, notes
    ("IN", "India",                "INR", "Rs.",  0, 0,    1999, 4999, 9999,  0.10, 5.00, None, "Y", "Home market. Prices in INR per year."),
    ("US", "United States",        "USD", "$",    0, None, None, None, None,  None, None, None, "Y", "Fill USD prices."),
    ("SG", "Singapore",            "SGD", "S$",   0, None, None, None, None,  None, None, None, "Y", "Fill SGD prices."),
    ("AE", "United Arab Emirates", "AED", "AED",  0, None, None, None, None,  None, None, None, "Y", "Fill AED prices."),
]


# ── Styling helpers ─────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2C5777", end_color="2C5777", fill_type="solid")
NOTE_FONT = Font(italic=True, color="666666", size=9)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_header_row(ws, row_num, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def autosize_columns(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── ai_features.xlsx ────────────────────────────────────────────────────────

def make_ai_features_xlsx(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Features"

    # Title / instructions
    ws["A1"] = "AI Feature Routing — operator config (per AccGenie version)"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:C1")

    ws["A2"] = (
        "byok = customer's own Anthropic key REQUIRED (locked without one). "
        "ag_key = uses AccGenie's key billed to the customer wallet, unless "
        "the customer has supplied their own key."
    )
    ws["A2"].font = NOTE_FONT
    ws["A2"].alignment = LEFT
    ws.merge_cells("A2:C2")
    ws.row_dimensions[2].height = 32

    # Headers on row 4
    headers = ["feature_id", "class", "notes"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, len(headers))

    # Data rows
    for r, (fid, cls, notes) in enumerate(AI_FEATURES_SEED, start=5):
        ws.cell(row=r, column=1, value=fid).border = THIN_BORDER
        c = ws.cell(row=r, column=2, value=cls)
        c.border = THIN_BORDER
        c.alignment = CENTER
        n = ws.cell(row=r, column=3, value=notes)
        n.border = THIN_BORDER
        n.alignment = LEFT

    # Dropdown for class column
    dv = DataValidation(type="list", formula1='"byok,ag_key"', allow_blank=False)
    dv.error = "Must be 'byok' or 'ag_key'"
    dv.errorTitle = "Invalid class"
    last_row = 4 + len(AI_FEATURES_SEED)
    dv.add(f"B5:B{last_row}")
    ws.add_data_validation(dv)

    autosize_columns(ws, [28, 12, 70])
    ws.freeze_panes = "A5"

    wb.save(path)


# ── pricing.xlsx ────────────────────────────────────────────────────────────

def make_pricing_xlsx(path: Path):
    wb = Workbook()

    # Sheet 1: Tiers
    ws = wb.active
    ws.title = "Tiers"
    ws["A1"] = "Plan tiers — one row per tier"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:G1")
    ws["A2"] = (
        "txn_limit = hard cap for DEMO/FREE, soft cap for paid plans (overage charged). "
        "overage_rate = Rs. per extra transaction beyond txn_limit (paid plans only). "
        "plan_price_INR is informational; per-country pricing is on the Countries sheet."
    )
    ws["A2"].font = NOTE_FONT
    ws["A2"].alignment = LEFT
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 32

    headers = ["code", "name", "seats_allowed", "txn_limit",
               "overage_rate", "plan_price_INR", "notes"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, len(headers))

    for r, row_data in enumerate(TIERS_SEED, start=5):
        for c, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN_BORDER
            if c in (3, 4, 5, 6):  # numeric columns
                cell.alignment = CENTER
            elif c == 7:
                cell.alignment = LEFT

    autosize_columns(ws, [12, 12, 14, 12, 14, 16, 60])
    ws.freeze_panes = "A5"

    # Sheet 2: PlanFeatures — row per feature, column per plan, grouped by category
    ws2 = wb.create_sheet("PlanFeatures")
    ws2["A1"] = "Plan ⇄ Feature matrix — Y in a cell means the plan includes that feature"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.merge_cells("A1:H1")
    ws2["A2"] = (
        "Type Y to include a feature in a plan, leave blank to exclude. "
        "'upgrade_to' is the tier shown to the user when they hit a locked "
        "feature. 'category' is for visual grouping only — the baker ignores it."
    )
    ws2["A2"].font = NOTE_FONT
    ws2["A2"].alignment = LEFT
    ws2.merge_cells("A2:H2")
    ws2.row_dimensions[2].height = 32

    plan_codes = [t[0] for t in TIERS_SEED]
    headers2 = ["category", "feature_id"] + plan_codes + ["upgrade_to"]
    for i, h in enumerate(headers2, start=1):
        ws2.cell(row=4, column=i, value=h)
    style_header_row(ws2, 4, len(headers2))

    cat_font = Font(bold=True, color="2C5777", size=10)
    cat_fill = PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid")

    for r, (category, fid) in enumerate(ALL_FEATURES_V2, start=5):
        c1 = ws2.cell(row=r, column=1, value=category)
        c1.border = THIN_BORDER
        c1.alignment = CENTER
        c1.font = cat_font
        c1.fill = cat_fill
        ws2.cell(row=r, column=2, value=fid).border = THIN_BORDER
        for j, plan in enumerate(plan_codes, start=3):
            included = fid in PLAN_FEATURE_MATRIX.get(plan, set())
            cell = ws2.cell(row=r, column=j, value=("Y" if included else None))
            cell.border = THIN_BORDER
            cell.alignment = CENTER
        upgrade_col = 2 + len(plan_codes) + 1
        up = FEATURE_UPGRADE_MAP_SEED.get(fid, "")
        cell = ws2.cell(row=r, column=upgrade_col, value=up if up else None)
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    # Dropdown for plan-column cells: Y or blank
    dv = DataValidation(type="list", formula1='"Y,"', allow_blank=True)
    last_row2 = 4 + len(ALL_FEATURES_V2)
    first_plan_col = 3
    last_plan_col = 2 + len(plan_codes)
    rng = (f"{get_column_letter(first_plan_col)}5:"
           f"{get_column_letter(last_plan_col)}{last_row2}")
    dv.add(rng)
    ws2.add_data_validation(dv)

    # Dropdown for upgrade_to column: must be one of the tier codes (or blank)
    dv2 = DataValidation(type="list", formula1=f'"{",".join(plan_codes)}"', allow_blank=True)
    upgrade_col_letter = get_column_letter(2 + len(plan_codes) + 1)
    dv2.add(f"{upgrade_col_letter}5:{upgrade_col_letter}{last_row2}")
    ws2.add_data_validation(dv2)

    autosize_columns(ws2, [14, 26] + [11] * len(plan_codes) + [14])
    ws2.freeze_panes = "C5"

    # Sheet 3: Countries
    ws3 = wb.create_sheet("Countries")
    ws3["A1"] = "Per-country pricing — one row per country"
    ws3["A1"].font = Font(bold=True, size=13)
    ws3.merge_cells("A1:N1")
    ws3["A2"] = (
        "Tier price columns hold the per-year price in this country's currency. "
        "Leave a tier's price blank if the tier is not sold in this country. "
        "ai_text_page_cost / ai_scanned_page_cost = wallet debit per AI page. "
        "active=Y to expose the country in the upgrade UI."
    )
    ws3["A2"].font = NOTE_FONT
    ws3["A2"].alignment = LEFT
    ws3.merge_cells("A2:N2")
    ws3.row_dimensions[2].height = 44

    tier_price_headers = [f"price_{p}" for p in plan_codes]
    headers3 = (
        ["country_code", "country_name", "currency_code", "currency_symbol"]
        + tier_price_headers
        + ["ai_text_page_cost", "ai_scanned_page_cost",
           "ai_per_transaction_cost", "active", "notes"]
    )
    for i, h in enumerate(headers3, start=1):
        ws3.cell(row=4, column=i, value=h)
    style_header_row(ws3, 4, len(headers3))

    for r, row_data in enumerate(COUNTRIES_SEED, start=5):
        for c, val in enumerate(row_data, start=1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.border = THIN_BORDER
            if 5 <= c <= 4 + len(plan_codes) + 3:
                cell.alignment = CENTER

    # active dropdown
    dv3 = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
    active_col = get_column_letter(len(headers3) - 1)
    last_row3 = 4 + len(COUNTRIES_SEED)
    dv3.add(f"{active_col}5:{active_col}{last_row3}")
    ws3.add_data_validation(dv3)

    widths = ([13, 22, 13, 13] + [13] * len(plan_codes)
              + [18, 20, 22, 9, 36])
    autosize_columns(ws3, widths)
    ws3.freeze_panes = "B5"

    wb.save(path)


# ── Entrypoint ──────────────────────────────────────────────────────────────

def main():
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    ai_path = CONFIG_DIR / "ai_features.xlsx"
    pricing_path = CONFIG_DIR / "pricing.xlsx"
    make_ai_features_xlsx(ai_path)
    make_pricing_xlsx(pricing_path)
    print(f"Wrote {ai_path}")
    print(f"Wrote {pricing_path}")


if __name__ == "__main__":
    main()
