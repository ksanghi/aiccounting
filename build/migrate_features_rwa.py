"""
One-shot migration: append RWA-vertical feature rows to the existing
config/pricing.xlsx PlanFeatures sheet. Idempotent — if any rwa_* row
already exists, the script does nothing.

These features are placeholders for the RWAGenie vertical (see
docs/roadmap.md section 3). The code that implements them lives under
verticals/rwa/ once that work begins. Adding the rows to PlanFeatures
now means lmgr.has_feature("rwa_*") will return the right value the
moment the implementation lands — no schema change needed later.

Run once:
    python build/migrate_features_rwa.py
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX = REPO_ROOT / "config" / "pricing.xlsx"


# Cleaned + deduplicated from File exchange/RWA features.xlsx. Each row:
# (category, feature_id, DEMO, FREE, STANDARD, PRO, PREMIUM, upgrade_to)
# DEMO mirrors PREMIUM (all features unlocked, txn-capped).
RWA_FEATURES = [
    # Core — available on every paid tier (and FREE)
    ("RWA",  "rwa_flat_ledger",          "Y", "Y", "Y", "Y", "Y", ""),
    ("RWA",  "rwa_receipt_tracking",     "Y", "Y", "Y", "Y", "Y", ""),
    ("RWA",  "rwa_member_directory",     "Y", "Y", "Y", "Y", "Y", ""),
    ("RWA",  "rwa_notice_board",         "Y", "Y", "Y", "Y", "Y", ""),
    ("RWA",  "rwa_complaint_tracking",   "Y", "Y", "Y", "Y", "Y", ""),
    ("RWA",  "rwa_broadcast_messaging",  "Y", "Y", "Y", "Y", "Y", ""),
    ("RWA",  "rwa_polls",                "Y", "Y", "Y", "Y", "Y", ""),
    ("RWA",  "rwa_visitor_pass",         "Y", "Y", "Y", "Y", "Y", ""),
    ("RWA",  "rwa_basic_reports",        "Y", "Y", "Y", "Y", "Y", ""),
    # Standard adds
    ("RWA",  "rwa_auto_billing",         "Y", "",  "Y", "Y", "Y", "STANDARD"),
    ("RWA",  "rwa_late_fees",            "Y", "",  "Y", "Y", "Y", "STANDARD"),
    ("RWA",  "rwa_facilities_booking",   "Y", "",  "Y", "Y", "Y", "STANDARD"),
    ("RWA",  "rwa_asset_register",       "Y", "",  "Y", "Y", "Y", "STANDARD"),
    ("RWA",  "rwa_advanced_reports",     "Y", "",  "Y", "Y", "Y", "STANDARD"),
    # Pro adds
    ("RWA",  "rwa_whatsapp_invoices",    "Y", "",  "",  "Y", "Y", "PRO"),
    ("RWA",  "rwa_document_storage",     "Y", "",  "",  "Y", "Y", "PRO"),
    ("RWA",  "rwa_vendor_management",    "Y", "",  "",  "Y", "Y", "PRO"),
]


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2C5777", end_color="2C5777", fill_type="solid")
CAT_FONT = Font(bold=True, color="2C5777", size=10)
CAT_FILL = PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def main():
    if not XLSX.exists():
        sys.exit(f"Not found: {XLSX}")

    wb = load_workbook(XLSX)
    if "PlanFeatures" not in wb.sheetnames:
        sys.exit("PlanFeatures sheet missing from pricing.xlsx — run "
                 "build/make_config_xlsx.py first.")

    ws = wb["PlanFeatures"]

    # Header is at row 4. Confirm we're operating on the v2 layout
    # (category column first).
    h1 = ws.cell(row=4, column=1).value
    if str(h1 or "").strip().lower() != "category":
        sys.exit("Expected 'category' in column A of row 4. Has the "
                 "PlanFeatures sheet been migrated? Run "
                 "build/migrate_features_v2.py first.")

    # Find existing feature_ids to avoid duplicating rows.
    existing_ids: set[str] = set()
    last_data_row = 4
    for r in range(5, ws.max_row + 1):
        fid = ws.cell(row=r, column=2).value
        if fid and str(fid).strip():
            existing_ids.add(str(fid).strip())
            last_data_row = r

    to_add = [row for row in RWA_FEATURES if row[1] not in existing_ids]
    if not to_add:
        print("All RWA features already present in PlanFeatures. Nothing to do.")
        return

    # Append below the last data row.
    start_row = last_data_row + 1
    for i, (cat, fid, demo, free, std, pro, prem, up) in enumerate(to_add):
        r = start_row + i
        c1 = ws.cell(row=r, column=1, value=cat)
        c1.alignment = CENTER
        c1.font = CAT_FONT
        c1.fill = CAT_FILL
        c1.border = THIN_BORDER

        ws.cell(row=r, column=2, value=fid).border = THIN_BORDER

        for j, val in enumerate([demo, free, std, pro, prem], start=3):
            cell = ws.cell(row=r, column=j, value=val or None)
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        cell = ws.cell(row=r, column=8, value=up or None)
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    wb.save(XLSX)
    print(f"Added {len(to_add)} RWA feature row(s) to {XLSX}.")
    print("Run `python build/bake_config.py` to regenerate "
          "core/_baked_config.py + license_server/_baked_config.py.")


if __name__ == "__main__":
    main()
