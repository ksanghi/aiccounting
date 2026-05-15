"""
One-shot migration: rewrite the PlanFeatures sheet in config/pricing.xlsx
with the v2 layout — category column on the left, reports split per-report,
print + email_report added, rows grouped by functionality.

Preserves existing Y/y/etc. marks where the feature_id is unchanged. For
the 7 new per-report rows, copies the value from the old "reports" row.
For new print/email_report rows, copies the value from the old "reports"
row as a sensible default.

Idempotent if re-run: detects the new layout by the presence of a
"category" column and bails with a message instead of overwriting.

Run once:
    python build/migrate_features_v2.py
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX = REPO_ROOT / "config" / "pricing.xlsx"


# v2 feature list — (category, feature_id, default upgrade_to)
# Order in this list = order of rows in the sheet, grouped by category.
V2_FEATURES = [
    # Core — always available
    ("Core",           "vouchers",              ""),
    ("Core",           "daybook",               ""),
    ("Core",           "ledger_balances",       ""),
    ("Core",           "backup",                ""),
    # Users
    ("Users",          "multi_user_2",          ""),
    ("Users",          "multi_user_5",          ""),
    ("Users",          "multi_user_unlimited",  ""),
    # Reports — one row per report, replaces old "reports" umbrella
    ("Reports",        "trial_balance",         "STANDARD"),
    ("Reports",        "profit_loss",           "STANDARD"),
    ("Reports",        "balance_sheet",         "STANDARD"),
    ("Reports",        "cash_book",             "STANDARD"),
    ("Reports",        "bank_book",             "STANDARD"),
    ("Reports",        "ledger_account",        "STANDARD"),
    ("Reports",        "receipts_payments",     "STANDARD"),
    # Output
    ("Output",         "export_excel",          "STANDARD"),
    ("Output",         "export_pdf",            "STANDARD"),
    ("Output",         "print",                 "STANDARD"),
    ("Output",         "email_report",          "STANDARD"),
    # Reconciliation
    ("Reconciliation", "bank_reconciliation",   "STANDARD"),
    ("Reconciliation", "ledger_reconciliation", "STANDARD"),
    # Migration
    ("Migration",      "book_migration",        "STANDARD"),
    # Tax
    ("Tax",            "gst",                   "PRO"),
    ("Tax",            "tds",                   "PRO"),
    # AI
    ("AI",             "ai_document_reader",    "PRO"),
    ("AI",             "verbal_entry",          "PRO"),
    ("AI",             "auto_billing",          "PRO"),
    # Enterprise
    ("Enterprise",     "whatsapp",              "PREMIUM"),
    ("Enterprise",     "audit_export",          "PREMIUM"),
    ("Enterprise",     "api_access",            "PREMIUM"),
    ("Enterprise",     "verticals",             "PREMIUM"),
]

# Features whose initial plan-column values should be COPIED from the old
# "reports" row when migrating (per-report split + print + email_report).
COPY_FROM_REPORTS = {
    "trial_balance", "profit_loss", "balance_sheet",
    "cash_book", "bank_book", "ledger_account", "receipts_payments",
    "print", "email_report",
}


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2C5777", end_color="2C5777", fill_type="solid")
CAT_FONT = Font(bold=True, color="2C5777", size=10)
CAT_FILL = PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid")
NOTE_FONT = Font(italic=True, color="666666", size=9)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def main():
    if not XLSX.exists():
        sys.exit(f"Not found: {XLSX}")

    wb = load_workbook(XLSX)
    if "PlanFeatures" not in wb.sheetnames:
        sys.exit("PlanFeatures sheet missing from pricing.xlsx")

    old_ws = wb["PlanFeatures"]

    # Detect existing layout: header row is row 4.
    headers = [old_ws.cell(row=4, column=c).value
               for c in range(1, old_ws.max_column + 1)]
    headers = [h for h in headers if h is not None]
    if "category" in [str(h).strip().lower() for h in headers]:
        print("PlanFeatures already in v2 layout (has 'category' column). "
              "Nothing to do.")
        return

    # Tier codes come from the Tiers sheet — single source of truth.
    tiers_ws = wb["Tiers"]
    tier_codes = []
    for r in range(5, tiers_ws.max_row + 1):
        code = tiers_ws.cell(row=r, column=1).value
        if code and str(code).strip():
            tier_codes.append(str(code).strip())
    if not tier_codes:
        sys.exit("No tier codes found in Tiers sheet")

    # Read existing rows into a dict.
    existing: dict[str, dict] = {}
    for r in range(5, old_ws.max_row + 1):
        fid = old_ws.cell(row=r, column=1).value
        if not fid or not str(fid).strip():
            continue
        fid = str(fid).strip()
        row_data = {}
        for c, h in enumerate(headers, start=1):
            row_data[h] = old_ws.cell(row=r, column=c).value
        existing[fid] = row_data

    old_reports = existing.get("reports", {})

    # Delete the old sheet and create a fresh one.
    del wb["PlanFeatures"]
    # Re-create in the same position (sheet 2).
    new_ws = wb.create_sheet("PlanFeatures", index=1)

    # Title rows
    new_ws["A1"] = (
        "Plan ⇄ Feature matrix — Y in a cell means the plan includes that feature"
    )
    new_ws["A1"].font = Font(bold=True, size=13)
    new_ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=4 + len(tier_codes))
    new_ws["A2"] = (
        "Type Y to include a feature in a plan, leave blank to exclude. "
        "'upgrade_to' is the tier shown to the user when they hit a locked "
        "feature. 'category' is for visual grouping only — the baker ignores it."
    )
    new_ws["A2"].font = NOTE_FONT
    new_ws["A2"].alignment = LEFT
    new_ws.merge_cells(start_row=2, start_column=1,
                       end_row=2, end_column=4 + len(tier_codes))
    new_ws.row_dimensions[2].height = 32

    # Header row 4
    headers_out = ["category", "feature_id"] + tier_codes + ["upgrade_to"]
    for i, h in enumerate(headers_out, start=1):
        cell = new_ws.cell(row=4, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Data rows, grouped by category
    r = 5
    last_cat = None
    for category, fid, default_up in V2_FEATURES:
        # Lookup existing values
        if fid in existing:
            src = existing[fid]
        elif fid in COPY_FROM_REPORTS and old_reports:
            src = old_reports
        else:
            src = {}

        # Write row
        c1 = new_ws.cell(row=r, column=1, value=category)
        c1.alignment = CENTER
        c1.font = CAT_FONT
        c1.fill = CAT_FILL
        c1.border = THIN_BORDER

        c2 = new_ws.cell(row=r, column=2, value=fid)
        c2.alignment = LEFT
        c2.border = THIN_BORDER

        for j, tier in enumerate(tier_codes, start=3):
            val = src.get(tier)
            # Strip whitespace, keep lowercase or uppercase Y intact.
            if isinstance(val, str):
                val = val.strip() or None
            cell = new_ws.cell(row=r, column=j, value=val)
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        up_col = 2 + len(tier_codes) + 1
        existing_up = src.get("upgrade_to")
        if isinstance(existing_up, str) and existing_up.strip():
            up_val = existing_up.strip().upper()
        else:
            up_val = default_up or None
        cell = new_ws.cell(row=r, column=up_col, value=up_val)
        cell.alignment = CENTER
        cell.border = THIN_BORDER

        r += 1
        last_cat = category

    # Data validation: plan columns accept Y or blank
    last_row = 4 + len(V2_FEATURES)
    dv = DataValidation(type="list", formula1='"Y,"', allow_blank=True)
    first_plan_col = 3
    last_plan_col = 2 + len(tier_codes)
    rng = (f"{get_column_letter(first_plan_col)}5:"
           f"{get_column_letter(last_plan_col)}{last_row}")
    dv.add(rng)
    new_ws.add_data_validation(dv)

    # Data validation: upgrade_to is one of the tier codes (or blank)
    dv2 = DataValidation(type="list",
                         formula1=f'"{",".join(tier_codes)}"',
                         allow_blank=True)
    upgrade_col_letter = get_column_letter(len(tier_codes) + 3)
    dv2.add(f"{upgrade_col_letter}5:{upgrade_col_letter}{last_row}")
    new_ws.add_data_validation(dv2)

    # Column widths
    widths = [14, 26] + [11] * len(tier_codes) + [14]
    for i, w in enumerate(widths, start=1):
        new_ws.column_dimensions[get_column_letter(i)].width = w
    new_ws.freeze_panes = "C5"

    wb.save(XLSX)
    print(f"Migrated {XLSX} -> v2 layout.")
    print(f"  Categories: {sorted(set(c for c, _, _ in V2_FEATURES))}")
    print(f"  Total features: {len(V2_FEATURES)}")
    if old_reports:
        kept = {k: v for k, v in old_reports.items() if k in tier_codes}
        print(f"  Old 'reports' row values copied to per-report rows: {kept}")


if __name__ == "__main__":
    main()
