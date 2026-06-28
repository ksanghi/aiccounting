"""
Exporters — Excel (openpyxl) and PDF (reportlab) export for all reports.
Both classes degrade gracefully if the library is not installed.
"""
import os


def _fmt(v: float) -> str:
    return f"{v:,.2f}"


# ── Excel ─────────────────────────────────────────────────────────────────────

class ExcelExporter:

    def __init__(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            self._xl   = openpyxl
            self._Font = Font
            self._Fill = PatternFill
            self._Aln  = Alignment
            self.available = True
        except ImportError:
            self.available = False

    # shared helpers
    def _wb(self, title: str):
        wb = self._xl.Workbook()
        ws = wb.active
        ws.title = title
        return wb, ws

    def _header(self, ws, company: dict, subtitle: str, row: int = 1):
        ws.cell(row, 1, company.get("name", "Company")).font = self._Font(bold=True, size=13)
        ws.cell(row+1, 1, subtitle).font = self._Font(bold=True, size=10)
        return row + 3  # next usable row

    def _col_hdrs(self, ws, row: int, headers: list):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row, c, h)
            cell.font = self._Font(bold=True, color="FFFFFF")
            cell.fill = self._Fill("solid", fgColor="1A3A5C")
        return row + 1

    def _set_widths(self, ws, widths: list):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[
                self._xl.utils.get_column_letter(i)
            ].width = w

    # ── reports ──────────────────────────────────────────────────────────────

    def trial_balance(self, data: list, company: dict, as_of: str, path: str):
        wb, ws = self._wb("Trial Balance")
        r = self._header(ws, company, f"Trial Balance as on {as_of}")
        r = self._col_hdrs(ws, r, [
            "#","Ledger","Group","Nature","Op Dr","Op Cr","Txn Dr","Txn Cr","Cl Dr","Cl Cr"
        ])
        for i, d in enumerate(data, 1):
            ws.append([i, d["ledger"], d["group"], d["nature"],
                       d["opening_dr"], d["opening_cr"],
                       d["txn_dr"],    d["txn_cr"],
                       d["closing_dr"],d["closing_cr"]])
        ws.append(["","TOTAL","","",
            sum(d["opening_dr"] for d in data), sum(d["opening_cr"] for d in data),
            sum(d["txn_dr"]     for d in data), sum(d["txn_cr"]     for d in data),
            sum(d["closing_dr"] for d in data), sum(d["closing_cr"] for d in data),
        ])
        self._set_widths(ws, [4,30,22,10,14,14,14,14,14,14])
        wb.save(path)

    def simple_table(self, title: str, subtitle: str, headers: list,
                     rows: list, company: dict, path: str, widths: list = None):
        """Generic single-table sheet — used by report pages that don't have a
        bespoke exporter (e.g. Schedule C, 1099, Mileage)."""
        wb, ws = self._wb(title[:31])
        r = self._header(ws, company, subtitle or title)
        r = self._col_hdrs(ws, r, headers)
        for row in rows:
            ws.append(list(row))
        self._set_widths(ws, widths or [20] * len(headers))
        wb.save(path)

    def profit_and_loss(self, data: dict, company: dict, path: str):
        wb, ws = self._wb("P and L")
        r = self._header(ws, company, f"Profit & Loss: {data['from_date']} to {data['to_date']}")
        r = self._col_hdrs(ws, r, ["Section","Ledger","Amount (Rs)"])
        for d in data["income"]:
            ws.append(["Income", d["ledger"], d["amount"]])
        ws.append(["","Total Income", data["total_income"]])
        for d in data["expenses"]:
            ws.append(["Expense", d["ledger"], d["amount"]])
        ws.append(["","Total Expenses", data["total_expense"]])
        ws.append(["","Net Profit / (Loss)", data["net_profit"]])
        self._set_widths(ws, [12,36,16])
        wb.save(path)

    def balance_sheet(self, data: dict, company: dict, path: str):
        wb, ws = self._wb("Balance Sheet")
        r = self._header(ws, company, f"Balance Sheet as on {data['as_of']}")
        r = self._col_hdrs(ws, r, ["Assets","Amount","Liabilities","Amount"])
        assets = data["assets"]
        liabs  = data["liabilities"]
        for i in range(max(len(assets), len(liabs))):
            a = assets[i] if i < len(assets) else {}
            l = liabs[i]  if i < len(liabs)  else {}
            # A balance on the side opposite its column's natural side (e.g. a
            # Cr balance sitting under Assets) is shown NEGATIVE so the column
            # visibly reconciles with the signed total below.
            a_amt = (a["balance"] if a["side"] == "Dr" else -a["balance"]) if a else ""
            l_amt = (l["balance"] if l["side"] == "Cr" else -l["balance"]) if l else ""
            ws.append([
                a.get("ledger",""), a_amt,
                l.get("ledger",""), l_amt,
            ])
        ws.append(["Total Assets", data["total_assets"], "Total Liabilities", data["total_liabilities"]])
        self._set_widths(ws, [30,14,30,14])
        wb.save(path)

    def ledger_book(self, data: dict, company: dict, title: str, path: str):
        wb = self._xl.Workbook()
        wb.remove(wb.active)
        for book in data["books"]:
            ws = wb.create_sheet(book["ledger"][:28])
            r = self._header(ws, company, f"{title}: {book['ledger']} ({data['from_date']} to {data['to_date']})")
            ws.append(["Opening Balance","","","","","","", book["opening"]])
            r = self._col_hdrs(ws, r+1, ["Date","Voucher","Type","Particulars","Narration","Ref","Dr","Cr","Balance"])
            for t in book["transactions"]:
                ws.append([t["date"], t["voucher_no"], t["voucher_type"],
                           t.get("party",""), t["narration"], t["reference"],
                           t["dr"] or "", t["cr"] or "", t["balance"]])
            ws.append(["Closing Balance","","","","","","","", book["closing"]])
            self._set_widths(ws, [12,14,8,26,26,12,12,12,14])
        wb.save(path)

    def receipts_payments(self, data: dict, company: dict, path: str):
        wb, ws = self._wb("Receipts and Payments")
        r = self._header(ws, company, f"Receipts & Payments: {data['from_date']} to {data['to_date']}")
        r = self._col_hdrs(ws, r, ["Voucher Type","Count","Amount (Rs)"])
        for key, label in [
            ("receipts","Receipts"), ("payments","Payments"), ("sales","Sales"),
            ("purchases","Purchases"), ("journals","Journal Entries"), ("contras","Contra"),
        ]:
            d = data[key]
            ws.append([label, d["count"], d["total"]])
        self._set_widths(ws, [24,10,18])
        wb.save(path)

    def gst_summary(self, data: dict, company: dict, path: str):
        wb, ws = self._wb("GST Summary")
        r = self._header(ws, company, f"GST Summary: {data['from_date']} to {data['to_date']}")
        r = self._col_hdrs(ws, r, ["Tax Type","Rate %","Output Tax","Input Tax (ITC)"])
        for row in data["tax_lines"]:
            ws.append([row["tax_type"], row["tax_rate"], row["output_tax"], row["input_tax"]])
        ws.append(["Total","", data["total_output"], data["total_input"]])
        ws.append(["Net GST Payable","","", data["net_gst_payable"]])
        ws.append([])
        ws.append(["Sales Base Value", data["sales_base"]])
        ws.append(["Purchase Base Value", data["purchase_base"]])
        self._set_widths(ws, [14,10,18,18])
        wb.save(path)

    def tds_report(self, data: dict, company: dict, path: str):
        wb, ws = self._wb("TDS Report")
        r = self._header(ws, company, f"TDS Report: {data['from_date']} to {data['to_date']}")
        r = self._col_hdrs(ws, r, ["Rate %","TDS Amount","Voucher Count"])
        for row in data["tds_lines"]:
            ws.append([row["tax_rate"], row["tds_amount"], row["voucher_count"]])
        ws.append(["Total TDS", data["total_tds"], ""])
        self._set_widths(ws, [12,18,14])
        wb.save(path)

    def gstr3b(self, data: dict, company: dict, path: str):
        wb, ws = self._wb("GSTR-3B")
        r = self._header(ws, company, f"GSTR-3B: {data['from_date']} to {data['to_date']}")
        r = self._col_hdrs(ws, r, ["Section","IGST","CGST","SGST","Total"])
        ow, itc, net = data["outward"], data["itc"], data["net_payable"]
        ws.append(["3.1 Outward tax", ow["IGST"], ow["CGST"], ow["SGST"], data["total_output"]])
        ws.append(["4. Eligible ITC", itc["IGST"], itc["CGST"], itc["SGST"], data["total_itc"]])
        ws.append(["Net payable",     net["IGST"], net["CGST"], net["SGST"], data["total_payable"]])
        ws.append([])
        ws.append(["Outward taxable value", ow["taxable"]])
        self._set_widths(ws, [22,14,14,14,16])
        wb.save(path)

    def gstr1(self, data: dict, company: dict, path: str):
        wb, ws = self._wb("GSTR-1")
        r = self._header(ws, company, f"GSTR-1: {data['from_date']} to {data['to_date']}")
        r = self._col_hdrs(ws, r, ["Invoice","Date","Party","GSTIN","POS","Cat",
                                   "Taxable","CGST","SGST","IGST"])
        for inv in data["invoices"]:
            ws.append([inv["invoice_no"], inv["invoice_date"], inv["party"], inv["gstin"],
                       inv["pos"], inv["category"], inv["taxable"],
                       inv["cgst"], inv["sgst"], inv["igst"]])
        b, c = data["b2b"], data["b2c"]
        ws.append([])
        ws.append(["B2B Total","","","","","", b["taxable"], b["cgst"], b["sgst"], b["igst"]])
        ws.append(["B2C Total","","","","","", c["taxable"], c["cgst"], c["sgst"], c["igst"]])
        self._set_widths(ws, [16,12,26,20,8,8,14,12,12,12])
        wb.save(path)

    def gstr2b_recon(self, data: dict, company: dict, path: str):
        wb, ws = self._wb("GSTR-2B Reconciliation")
        r = self._header(ws, company,
                         f"GSTR-2B Reconciliation: {data['from_date']} to {data['to_date']}")
        r = self._col_hdrs(ws, r, ["Status","GSTIN","Invoice","Party",
                                   "Book Taxable","2B Taxable","Book Tax","2B Tax","Diff"])
        def emit(status, items):
            for x in items:
                ws.append([status, x.get("gstin",""), x.get("invoice_no",""), x.get("party",""),
                           x.get("book_taxable",""), x.get("b2b_taxable",""),
                           x.get("book_tax",""), x.get("b2b_tax",""), x.get("diff","")])
        emit("Matched", data["matched"]); emit("Mismatch", data["mismatch"])
        emit("In books, not in 2B", data["only_books"]); emit("In 2B, not in books", data["only_2b"])
        ws.append([])
        ws.append(["ITC Matched", data["itc_matched"]])
        ws.append(["ITC at Risk (in books, not in 2B)", data["itc_at_risk"]])
        self._set_widths(ws, [22,20,16,24,14,14,12,12,10])
        wb.save(path)

    def tds_register(self, data: dict, company: dict, path: str):
        wb, ws = self._wb("TDS Register")
        r = self._header(ws, company, f"TDS Register: {data['from_date']} to {data['to_date']}")
        r = self._col_hdrs(ws, r, ["Party","PAN","Section","Nature","Rate %","Gross Paid","TDS","Txns"])
        for p in data["parties"]:
            ws.append([p["party"], p["pan"], p["section"], p["section_desc"],
                       p["rate"], p["gross"], p["tds"], p["count"]])
        ws.append(["Total","","","","", data["total_gross"], data["total_tds"], ""])
        self._set_widths(ws, [24,16,10,28,8,14,14,8])
        wb.save(path)

    def hsn_summary(self, data: dict, company: dict, path: str):
        wb, ws = self._wb("HSN Summary")
        r = self._header(ws, company, f"HSN Summary: {data['from_date']} to {data['to_date']}")
        r = self._col_hdrs(ws, r, ["HSN/SAC","Taxable","CGST","SGST","IGST","Total Tax"])
        for x in data["rows"]:
            ws.append([x["hsn"], x["taxable"], x["cgst"], x["sgst"], x["igst"],
                       round(x["cgst"]+x["sgst"]+x["igst"], 2)])
        ws.append(["Total", data["total_taxable"], "", "", "", data["total_tax"]])
        self._set_widths(ws, [22,16,14,14,14,14])
        wb.save(path)

    def receivables_aging(self, data: dict, company: dict, path: str):
        wb, ws = self._wb("Receivables Aging")
        r = self._header(ws, company,
                         f"Receivables Aging as on {data['as_of']}")
        r = self._col_hdrs(ws, r, ["Customer", "0-30 days", "31-60 days",
                                   "61-90 days", "90+ days", "Total"])
        for d in data["rows"]:
            ws.append([d["ledger"], d["b0_30"], d["b31_60"],
                       d["b61_90"], d["b90p"], d["total"]])
        t = data["totals"]
        ws.append(["TOTAL", t["b0_30"], t["b31_60"], t["b61_90"],
                   t["b90p"], round(sum(t.values()), 2)])
        self._set_widths(ws, [30,14,14,14,14,16])
        wb.save(path)

    def bill_wise_aging(self, data: dict, company: dict, path: str):
        wb, ws = self._wb("Bill-wise Outstanding")
        r = self._header(ws, company,
                         f"Bill-wise Outstanding ({data['group']}) as on {data['as_of']}")
        r = self._col_hdrs(ws, r, ["Party", "Bill No", "Bill Date",
                                   "Age (days)", "Bill Amount", "Pending"])
        for d in data["rows"]:
            ws.append([d["party"], d["bill_number"], d["bill_date"],
                       d["age_days"], d["bill_amount"], d["pending_amount"]])
        b = data["buckets"]
        ws.append([])
        ws.append(["Buckets", "0-30", b["b0_30"], "31-60", b["b31_60"], ""])
        ws.append(["", "61-90", b["b61_90"], "90+", b["b90p"], ""])
        ws.append(["TOTAL", "", "", "", "", data["total"]])
        self._set_widths(ws, [28,16,14,12,16,16])
        wb.save(path)

    def payables_aging(self, data: dict, company: dict, path: str):
        wb, ws = self._wb("Payables Aging")
        r = self._header(ws, company,
                         f"Payables Aging as on {data['as_of']}")
        r = self._col_hdrs(ws, r, ["Supplier", "0-30 days", "31-60 days",
                                   "61-90 days", "90+ days", "Total"])
        for d in data["rows"]:
            ws.append([d["ledger"], d["b0_30"], d["b31_60"],
                       d["b61_90"], d["b90p"], d["total"]])
        t = data["totals"]
        ws.append(["TOTAL", t["b0_30"], t["b31_60"], t["b61_90"],
                   t["b90p"], round(sum(t.values()), 2)])
        self._set_widths(ws, [30,14,14,14,14,16])
        wb.save(path)

    def cash_flow(self, data: dict, company: dict, path: str):
        wb, ws = self._wb("Cash-Flow Planning")
        r = self._header(ws, company,
                         f"Cash-Flow Planning (assisted) as on {data['as_of']}")
        r = self._col_hdrs(ws, r, ["Period", "Expected In", "Expected Out",
                                   "Net", "Projected Cash"])
        ws.append(["Opening cash", "", "", "", data["opening"]])
        for d in data["rows"]:
            ws.append([d["label"], d["inflow"], d["outflow"], d["net"],
                       d["closing"]])
        self._set_widths(ws, [22,16,16,16,16])
        wb.save(path)


# ── PDF ───────────────────────────────────────────────────────────────────────

class PDFExporter:

    _HDR_COLOR  = (0x1A/255, 0x3A/255, 0x5C/255)
    _ROW_COLORS = ((1,1,1), (0.95,0.97,1.0))

    def __init__(self):
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle,
                Paragraph, Spacer
            )
            self._A4         = A4
            self._landscape  = landscape
            self._colors     = colors
            self._styles     = getSampleStyleSheet()
            self._cm         = cm
            self._Doc        = SimpleDocTemplate
            self._Table      = Table
            self._TStyle     = TableStyle
            self._Para       = Paragraph
            self._Spacer     = Spacer
            self.available   = True
        except ImportError:
            self.available   = False

    def _doc(self, path: str, wide: bool = False):
        size = self._landscape(self._A4) if wide else self._A4
        cm = self._cm
        return self._Doc(path, pagesize=size,
                         topMargin=1.5*cm, bottomMargin=1.5*cm,
                         leftMargin=1.5*cm, rightMargin=1.5*cm)

    def _tbl_style(self):
        c = self._colors
        return self._TStyle([
            ("BACKGROUND",  (0,0), (-1,0),  c.HexColor("#1A3A5C")),
            ("TEXTCOLOR",   (0,0), (-1,0),  c.white),
            ("FONTNAME",    (0,0), (-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[c.white, c.HexColor("#EFF3F8")]),
            ("GRID",        (0,0), (-1,-1), 0.25, c.HexColor("#C0C8D0")),
            ("TOPPADDING",  (0,0), (-1,-1), 3),
            ("BOTTOMPADDING",(0,0),(-1,-1), 3),
            ("LEFTPADDING", (0,0), (-1,-1), 5),
        ])

    def _heading(self, company: dict, subtitle: str) -> list:
        s = self._styles
        return [
            self._Para(company.get("name","Company"), s["Title"]),
            self._Para(subtitle, s["Normal"]),
            self._Spacer(1, 0.35*self._cm),
        ]

    def trial_balance(self, data: list, company: dict, as_of: str, path: str):
        cm = self._cm
        doc = self._doc(path, wide=True)
        elems = self._heading(company, f"Trial Balance as on {as_of}")
        hdrs = ["#","Ledger","Group","Nat","Op Dr","Op Cr","Txn Dr","Txn Cr","Cl Dr","Cl Cr"]
        rows = [hdrs]
        for i, d in enumerate(data, 1):
            rows.append([str(i), d["ledger"][:26], d["group"][:18], d["nature"][:3],
                _fmt(d["opening_dr"]), _fmt(d["opening_cr"]),
                _fmt(d["txn_dr"]),    _fmt(d["txn_cr"]),
                _fmt(d["closing_dr"]),_fmt(d["closing_cr"])])
        rows.append(["","TOTAL","","",
            _fmt(sum(d["opening_dr"] for d in data)), _fmt(sum(d["opening_cr"] for d in data)),
            _fmt(sum(d["txn_dr"]     for d in data)), _fmt(sum(d["txn_cr"]     for d in data)),
            _fmt(sum(d["closing_dr"] for d in data)), _fmt(sum(d["closing_cr"] for d in data)),
        ])
        cw = [0.5*cm,4.2*cm,3.2*cm,1.2*cm,2.5*cm,2.5*cm,2.5*cm,2.5*cm,2.5*cm,2.5*cm]
        t = self._Table(rows, colWidths=cw)
        t.setStyle(self._tbl_style())
        elems.append(t)
        doc.build(elems)

    def profit_and_loss(self, data: dict, company: dict, path: str):
        cm = self._cm
        doc = self._doc(path)
        elems = self._heading(company, f"P&L: {data['from_date']} to {data['to_date']}")
        rows = [["Section","Ledger","Amount (Rs)"]]
        for d in data["income"]:
            rows.append(["Income", d["ledger"][:38], _fmt(d["amount"])])
        rows.append(["","Total Income", _fmt(data["total_income"])])
        for d in data["expenses"]:
            rows.append(["Expense", d["ledger"][:38], _fmt(d["amount"])])
        rows.append(["","Total Expenses", _fmt(data["total_expense"])])
        rows.append(["","Net Profit / (Loss)", _fmt(data["net_profit"])])
        cw = [2.5*cm, 9*cm, 4*cm]
        t = self._Table(rows, colWidths=cw)
        t.setStyle(self._tbl_style())
        elems.append(t)
        doc.build(elems)

    def balance_sheet(self, data: dict, company: dict, path: str):
        cm = self._cm
        doc = self._doc(path, wide=True)
        elems = self._heading(company, f"Balance Sheet as on {data['as_of']}")
        rows = [["Assets","Amount","Liabilities","Amount"]]
        a_list = data["assets"]
        l_list = data["liabilities"]
        for i in range(max(len(a_list), len(l_list))):
            a = a_list[i] if i < len(a_list) else {}
            l = l_list[i] if i < len(l_list) else {}
            # Contra-side balance (e.g. a Cr under Assets) shown negative so the
            # column reconciles with the signed total.
            a_amt = (a["balance"] if a["side"] == "Dr" else -a["balance"]) if a else None
            l_amt = (l["balance"] if l["side"] == "Cr" else -l["balance"]) if l else None
            rows.append([
                a.get("ledger","")[:30], _fmt(a_amt) if a else "",
                l.get("ledger","")[:30], _fmt(l_amt) if l else "",
            ])
        rows.append(["Total Assets", _fmt(data["total_assets"]),
                     "Total Liabilities", _fmt(data["total_liabilities"])])
        cw = [6*cm, 3.5*cm, 6*cm, 3.5*cm]
        t = self._Table(rows, colWidths=cw)
        t.setStyle(self._tbl_style())
        elems.append(t)
        doc.build(elems)

    def ledger_book(self, data: dict, company: dict, title: str, path: str):
        cm = self._cm
        doc = self._doc(path, wide=True)
        elems = self._heading(company, f"{title}: {data['from_date']} to {data['to_date']}")
        for book in data["books"]:
            elems.append(self._Para(book["ledger"], self._styles["Heading2"]))
            rows = [["Date","Voucher","Type","Particulars","Narration","Ref","Dr","Cr","Balance"]]
            rows.append(["Opening","","","","","","","",_fmt(book["opening"])])
            for t in book["transactions"]:
                rows.append([
                    t["date"], t["voucher_no"], t["voucher_type"][:3],
                    (t.get("party","") or "")[:24],
                    t["narration"][:22], t["reference"][:10],
                    _fmt(t["dr"]) if t["dr"] else "",
                    _fmt(t["cr"]) if t["cr"] else "",
                    _fmt(t["balance"]),
                ])
            rows.append(["Closing","","","","","","","",_fmt(book["closing"])])
            cw = [2*cm,2.6*cm,1.1*cm,4.6*cm,4.2*cm,2*cm,2.3*cm,2.3*cm,2.6*cm]
            tbl = self._Table(rows, colWidths=cw)
            tbl.setStyle(self._tbl_style())
            elems += [tbl, self._Spacer(1, 0.5*cm)]
        doc.build(elems)

    def receipts_payments(self, data: dict, company: dict, path: str):
        cm = self._cm
        doc = self._doc(path)
        elems = self._heading(company, f"Receipts & Payments: {data['from_date']} to {data['to_date']}")
        rows = [["Voucher Type","Count","Amount (Rs)"]]
        for key, label in [
            ("receipts","Receipts"), ("payments","Payments"), ("sales","Sales"),
            ("purchases","Purchases"), ("journals","Journal Entries"), ("contras","Contra"),
        ]:
            d = data[key]
            rows.append([label, str(d["count"]), _fmt(d["total"])])
        cw = [6*cm, 3*cm, 5*cm]
        t = self._Table(rows, colWidths=cw)
        t.setStyle(self._tbl_style())
        elems.append(t)
        doc.build(elems)

    def gst_summary(self, data: dict, company: dict, path: str):
        cm = self._cm
        doc = self._doc(path)
        elems = self._heading(company, f"GST Summary: {data['from_date']} to {data['to_date']}")
        rows = [["Tax Type","Rate %","Output Tax","ITC (Input)"]]
        for r in data["tax_lines"]:
            rows.append([r["tax_type"], f"{r['tax_rate']}%",
                         _fmt(r["output_tax"]), _fmt(r["input_tax"])])
        rows.append(["Total","",_fmt(data["total_output"]),_fmt(data["total_input"])])
        rows.append(["Net GST Payable","","",_fmt(data["net_gst_payable"])])
        rows.append(["Sales Base",_fmt(data["sales_base"]),"",""])
        rows.append(["Purchase Base",_fmt(data["purchase_base"]),"",""])
        cw = [3.5*cm,2.5*cm,5*cm,5*cm]
        t = self._Table(rows, colWidths=cw)
        t.setStyle(self._tbl_style())
        elems.append(t)
        doc.build(elems)

    def tds_report(self, data: dict, company: dict, path: str):
        cm = self._cm
        doc = self._doc(path)
        elems = self._heading(company, f"TDS Report: {data['from_date']} to {data['to_date']}")
        rows = [["Rate %","TDS Amount","Voucher Count"]]
        for r in data["tds_lines"]:
            rows.append([f"{r['tax_rate']}%", _fmt(r["tds_amount"]), str(r["voucher_count"])])
        rows.append(["Total TDS", _fmt(data["total_tds"]), ""])
        cw = [3*cm, 5*cm, 4*cm]
        t = self._Table(rows, colWidths=cw)
        t.setStyle(self._tbl_style())
        elems.append(t)
        doc.build(elems)

    def simple_table(self, title: str, subtitle: str, headers: list,
                     rows: list, company: dict, path: str,
                     widths: list = None, wide: bool = False):
        """Generic single-table PDF — used by report pages without a bespoke
        exporter (Schedule C, 1099, Mileage)."""
        doc = self._doc(path, wide=wide)
        elems = self._heading(company, subtitle or title)
        data = [list(headers)] + [list(r) for r in rows]
        t = self._Table(data, colWidths=widths)
        t.setStyle(self._tbl_style())
        elems.append(t)
        doc.build(elems)

    def gstr3b(self, data: dict, company: dict, path: str):
        cm = self._cm
        doc = self._doc(path)
        elems = self._heading(company, f"GSTR-3B: {data['from_date']} to {data['to_date']}")
        ow, itc, net = data["outward"], data["itc"], data["net_payable"]
        rows = [["Section","IGST","CGST","SGST","Total"],
                ["3.1 Outward tax", _fmt(ow["IGST"]), _fmt(ow["CGST"]), _fmt(ow["SGST"]), _fmt(data["total_output"])],
                ["4. Eligible ITC", _fmt(itc["IGST"]), _fmt(itc["CGST"]), _fmt(itc["SGST"]), _fmt(data["total_itc"])],
                ["Net payable",     _fmt(net["IGST"]), _fmt(net["CGST"]), _fmt(net["SGST"]), _fmt(data["total_payable"])],
                ["Outward taxable value", _fmt(ow["taxable"]), "", "", ""]]
        t = self._Table(rows, colWidths=[5*cm,3*cm,3*cm,3*cm,3.5*cm])
        t.setStyle(self._tbl_style())
        elems.append(t)
        doc.build(elems)

    def gstr1(self, data: dict, company: dict, path: str):
        cm = self._cm
        doc = self._doc(path, wide=True)
        elems = self._heading(company, f"GSTR-1: {data['from_date']} to {data['to_date']}")
        rows = [["Invoice","Date","Party","GSTIN","Cat","Taxable","CGST","SGST","IGST"]]
        for inv in data["invoices"]:
            rows.append([inv["invoice_no"], inv["invoice_date"], (inv["party"] or "")[:18],
                         inv["gstin"], inv["category"], _fmt(inv["taxable"]),
                         _fmt(inv["cgst"]), _fmt(inv["sgst"]), _fmt(inv["igst"])])
        b, c = data["b2b"], data["b2c"]
        rows.append(["B2B Total","","","","",_fmt(b["taxable"]),_fmt(b["cgst"]),_fmt(b["sgst"]),_fmt(b["igst"])])
        rows.append(["B2C Total","","","","",_fmt(c["taxable"]),_fmt(c["cgst"]),_fmt(c["sgst"]),_fmt(c["igst"])])
        cw = [3*cm,2*cm,4*cm,4*cm,1.5*cm,2.5*cm,2.2*cm,2.2*cm,2.2*cm]
        t = self._Table(rows, colWidths=cw)
        t.setStyle(self._tbl_style())
        elems.append(t)
        doc.build(elems)

    def gstr2b_recon(self, data: dict, company: dict, path: str):
        cm = self._cm
        doc = self._doc(path, wide=True)
        elems = self._heading(company,
                              f"GSTR-2B Reconciliation: {data['from_date']} to {data['to_date']}")
        rows = [["Status","GSTIN","Invoice","Party","Book Tax","2B Tax","Diff"]]
        def emit(status, items):
            for x in items:
                rows.append([status, x.get("gstin",""), x.get("invoice_no",""),
                             (x.get("party","") or "")[:16],
                             _fmt(x["book_tax"]) if "book_tax" in x else "",
                             _fmt(x["b2b_tax"]) if "b2b_tax" in x else "",
                             _fmt(x["diff"]) if "diff" in x else ""])
        emit("Matched", data["matched"]); emit("Mismatch", data["mismatch"])
        emit("Books not 2B", data["only_books"]); emit("2B not books", data["only_2b"])
        rows.append(["ITC at Risk","","","","","",_fmt(data["itc_at_risk"])])
        cw = [3.4*cm,4*cm,3*cm,4*cm,2.5*cm,2.5*cm,2.2*cm]
        t = self._Table(rows, colWidths=cw)
        t.setStyle(self._tbl_style())
        elems.append(t)
        doc.build(elems)

    def tds_register(self, data: dict, company: dict, path: str):
        cm = self._cm
        doc = self._doc(path, wide=True)
        elems = self._heading(company, f"TDS Register: {data['from_date']} to {data['to_date']}")
        rows = [["Party","PAN","Section","Nature","Rate %","Gross","TDS","Txns"]]
        for p in data["parties"]:
            rows.append([(p["party"] or "")[:18], p["pan"], p["section"],
                         (p["section_desc"] or "")[:18], f"{p['rate']}%",
                         _fmt(p["gross"]), _fmt(p["tds"]), str(p["count"])])
        rows.append(["Total","","","","",_fmt(data["total_gross"]),_fmt(data["total_tds"]),""])
        cw = [3.5*cm,3*cm,1.6*cm,3.5*cm,1.6*cm,2.5*cm,2.5*cm,1.4*cm]
        t = self._Table(rows, colWidths=cw)
        t.setStyle(self._tbl_style())
        elems.append(t)
        doc.build(elems)

    def hsn_summary(self, data: dict, company: dict, path: str):
        cm = self._cm
        doc = self._doc(path)
        elems = self._heading(company, f"HSN Summary: {data['from_date']} to {data['to_date']}")
        rows = [["HSN/SAC","Taxable","CGST","SGST","IGST","Total Tax"]]
        for x in data["rows"]:
            rows.append([x["hsn"], _fmt(x["taxable"]), _fmt(x["cgst"]), _fmt(x["sgst"]),
                         _fmt(x["igst"]), _fmt(round(x["cgst"]+x["sgst"]+x["igst"], 2))])
        rows.append(["Total", _fmt(data["total_taxable"]),"","","",_fmt(data["total_tax"])])
        t = self._Table(rows, colWidths=[3.5*cm,3*cm,2.6*cm,2.6*cm,2.6*cm,2.8*cm])
        t.setStyle(self._tbl_style())
        elems.append(t)
        doc.build(elems)

    def receivables_aging(self, data: dict, company: dict, path: str):
        cm = self._cm
        doc = self._doc(path, wide=True)
        elems = self._heading(company,
                              f"Receivables Aging as on {data['as_of']}")
        rows = [["Customer", "0-30", "31-60", "61-90", "90+", "Total"]]
        for d in data["rows"]:
            rows.append([d["ledger"], _fmt(d["b0_30"]), _fmt(d["b31_60"]),
                         _fmt(d["b61_90"]), _fmt(d["b90p"]), _fmt(d["total"])])
        t = data["totals"]
        rows.append(["TOTAL", _fmt(t["b0_30"]), _fmt(t["b31_60"]),
                     _fmt(t["b61_90"]), _fmt(t["b90p"]),
                     _fmt(round(sum(t.values()), 2))])
        cw = [7*cm, 3.4*cm, 3.4*cm, 3.4*cm, 3.4*cm, 3.8*cm]
        tbl = self._Table(rows, colWidths=cw)
        tbl.setStyle(self._tbl_style())
        elems.append(tbl)
        doc.build(elems)

    def bill_wise_aging(self, data: dict, company: dict, path: str):
        cm = self._cm
        doc = self._doc(path, wide=True)
        elems = self._heading(
            company,
            f"Bill-wise Outstanding ({data['group']}) as on {data['as_of']}")
        rows = [["Party", "Bill No", "Bill Date", "Age", "Bill Amt", "Pending"]]
        for d in data["rows"]:
            rows.append([(d["party"] or "")[:22], (d["bill_number"] or "")[:14],
                         d["bill_date"], str(d["age_days"]),
                         _fmt(d["bill_amount"]), _fmt(d["pending_amount"])])
        rows.append(["TOTAL", "", "", "", "", _fmt(data["total"])])
        cw = [4.6*cm, 3*cm, 2.6*cm, 1.6*cm, 3*cm, 3*cm]
        t = self._Table(rows, colWidths=cw)
        t.setStyle(self._tbl_style())
        elems.append(t)
        doc.build(elems)

    def payables_aging(self, data: dict, company: dict, path: str):
        cm = self._cm
        doc = self._doc(path, wide=True)
        elems = self._heading(company,
                              f"Payables Aging as on {data['as_of']}")
        rows = [["Supplier", "0-30", "31-60", "61-90", "90+", "Total"]]
        for d in data["rows"]:
            rows.append([d["ledger"], _fmt(d["b0_30"]), _fmt(d["b31_60"]),
                         _fmt(d["b61_90"]), _fmt(d["b90p"]), _fmt(d["total"])])
        t = data["totals"]
        rows.append(["TOTAL", _fmt(t["b0_30"]), _fmt(t["b31_60"]),
                     _fmt(t["b61_90"]), _fmt(t["b90p"]),
                     _fmt(round(sum(t.values()), 2))])
        cw = [7*cm, 3.4*cm, 3.4*cm, 3.4*cm, 3.4*cm, 3.8*cm]
        tbl = self._Table(rows, colWidths=cw)
        tbl.setStyle(self._tbl_style())
        elems.append(tbl)
        doc.build(elems)

    def cash_flow(self, data: dict, company: dict, path: str):
        cm = self._cm
        doc = self._doc(path, wide=True)
        elems = self._heading(
            company, f"Cash-Flow Planning (assisted) as on {data['as_of']}")
        rows = [["Period", "Expected In", "Expected Out", "Net", "Projected Cash"]]
        rows.append(["Opening cash", "", "", "", _fmt(data["opening"])])
        for d in data["rows"]:
            rows.append([d["label"], _fmt(d["inflow"]), _fmt(d["outflow"]),
                         _fmt(d["net"]), _fmt(d["closing"])])
        cw = [5.5*cm, 3.2*cm, 3.2*cm, 3.2*cm, 3.4*cm]
        tbl = self._Table(rows, colWidths=cw)
        tbl.setStyle(self._tbl_style())
        elems.append(tbl)
        doc.build(elems)
