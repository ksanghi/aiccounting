"""
GSTR-2B importer — parse the portal's 2B export (Excel/CSV) into normalised
invoice rows so the reports engine can reconcile them against the company's
purchase/ITC records.

This is REPORTS-side: the user downloads the 2B from the GST portal and
uploads it here. No GSP, no filing. We only read the file and match.

The portal's B2B sheet carries header noise (title rows) above the real
column header. We scan for the header row by keyword, map columns
flexibly (so minor portal-format changes don't break us), and read down.
"""
from __future__ import annotations

import csv as _csv
from pathlib import Path


def _num(v) -> float:
    if v is None:
        return 0.0
    s = str(v).strip().replace(",", "").replace("₹", "")
    if not s:
        return 0.0
    try:
        return round(float(s), 2)
    except ValueError:
        return 0.0


def _h(v) -> str:
    return str(v or "").strip().lower()


# header-keyword -> normalised field. First matching column wins.
def _build_colmap(header: list) -> dict:
    cm: dict[str, int] = {}
    for i, cell in enumerate(header):
        h = _h(cell)
        if not h:
            continue
        if "gstin" in h and "gstin" not in cm:
            cm["gstin"] = i
        elif "invoice" in h and "date" in h and "invoice_date" not in cm:
            cm["invoice_date"] = i
        elif "invoice" in h and ("number" in h or "no" in h) and "value" not in h and "invoice_no" not in cm:
            cm["invoice_no"] = i
        elif "taxable" in h and "taxable" not in cm:
            cm["taxable"] = i
        elif "integrated" in h and "igst" not in cm:
            cm["igst"] = i
        elif "central" in h and "cgst" not in cm:
            cm["cgst"] = i
        elif ("state" in h or "ut tax" in h) and "tax" in h and "sgst" not in cm:
            cm["sgst"] = i
    return cm


def _read_rows(path: str) -> list[list]:
    p = Path(path)
    if p.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        # Prefer a sheet whose name hints B2B; else the first sheet.
        ws = next((wb[s] for s in wb.sheetnames if "b2b" in s.lower()), wb.worksheets[0])
        return [list(r) for r in ws.iter_rows(values_only=True)]
    # CSV
    with open(path, newline="", encoding="utf-8-sig") as f:
        return [row for row in _csv.reader(f)]


def parse_gstr2b(path: str) -> list[dict]:
    """Return a list of {gstin, invoice_no, invoice_date, taxable,
    igst, cgst, sgst} from a portal 2B Excel/CSV. Raises ValueError if no
    recognisable header row is found."""
    rows = _read_rows(path)
    hdr_idx, colmap = None, {}
    for i, r in enumerate(rows[:25]):
        cm = _build_colmap(r)
        if "gstin" in cm and "invoice_no" in cm and "taxable" in cm:
            hdr_idx, colmap = i, cm
            break
    if hdr_idx is None:
        raise ValueError(
            "Couldn't find a GSTR-2B header row. The file needs columns for "
            "GSTIN, Invoice Number and Taxable Value (the portal's B2B sheet)."
        )

    def cell(r, key):
        idx = colmap.get(key)
        if idx is None or idx >= len(r):
            return None
        return r[idx]

    out: list[dict] = []
    for r in rows[hdr_idx + 1:]:
        gstin = cell(r, "gstin")
        inv = cell(r, "invoice_no")
        if not gstin or not inv or "gstin" in _h(gstin):
            continue
        out.append({
            "gstin":        str(gstin).strip().upper(),
            "invoice_no":   str(inv).strip(),
            "invoice_date": str(cell(r, "invoice_date") or "").strip(),
            "taxable":      _num(cell(r, "taxable")),
            "igst":         _num(cell(r, "igst")),
            "cgst":         _num(cell(r, "cgst")),
            "sgst":         _num(cell(r, "sgst")),
        })
    return out
