"""
Excel chart-of-accounts parser.

Accepts the user-prepared spreadsheet in two common shapes:

  Shape A — two named sheets:
      "Groups"  : Name, Parent, Nature
      "Ledgers" : Name, Group, Opening Balance, Opening Type [Dr/Cr],
                  GSTIN, PAN, State Code, Account Number, IFSC, Bank Name,
                  TDS Section, TDS Rate

  Shape B — single sheet with a "Type" column distinguishing rows:
      Type=Group rows define groups; Type=Ledger rows define ledgers.

Header sniffing is case-insensitive and tolerant — common synonyms
("ledger name", "account name", "balance", "dr/cr", "ifsc code",
"bank acc no", "tax id" …) are accepted. Missing optional columns just
leave the corresponding LedgerSpec field blank.
"""
from __future__ import annotations

from pathlib import Path
from typing import Optional

from .payload import (
    MigrationPayload, GroupSpec, LedgerSpec, CompanySpec,
)


# ── Header synonyms (case- and whitespace-insensitive) ──────────────────────

_LEDGER_HEADERS = {
    "name":            ("name", "ledger name", "account name", "ledger",
                        "account", "particulars"),
    "group":           ("group", "group name", "under group", "parent group",
                        "category", "type", "account type"),
    "opening_balance": ("opening balance", "opening", "balance",
                        "opening bal", "ob", "amount"),
    "opening_type":    ("dr/cr", "type (dr/cr)", "opening type", "drcr",
                        "balance type", "side"),
    "gstin":           ("gstin", "gst number", "gst no", "gst id"),
    "pan":             ("pan", "pan number", "pan no"),
    "state_code":      ("state code", "state", "state code (gst)"),
    "account_number":  ("account number", "account no", "account no.",
                        "a/c no", "a/c number", "bank account number",
                        "bank acc no"),
    "ifsc":            ("ifsc", "ifsc code", "ifsc no"),
    "bank_name":       ("bank name", "bank"),
    "tds_section":     ("tds section", "tds", "section"),
    "tds_rate":        ("tds rate", "tds %", "tds rate %"),
}

_GROUP_HEADERS = {
    "name":   ("name", "group name", "group"),
    "parent": ("parent", "parent group", "under", "under group"),
    "nature": ("nature", "type", "category", "kind"),
}


def _norm(s) -> str:
    return str(s or "").strip().lower()


def _pick(headers: list[str], options: tuple) -> Optional[int]:
    norm = [_norm(h) for h in headers]
    for o in options:
        if o in norm:
            return norm.index(o)
    # substring fallback
    for o in options:
        for i, h in enumerate(norm):
            if o in h:
                return i
    return None


def _parse_amount(raw) -> float:
    if raw is None:
        return 0.0
    s = str(raw).strip()
    if not s:
        return 0.0
    import re
    s = re.sub(r"[^\d.\-+]", "", s.replace(",", ""))
    if not s or s in ("-", "+", "."):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _norm_dr_cr(raw) -> str:
    s = _norm(raw)
    if s in ("dr", "debit", "d"):
        return "Dr"
    if s in ("cr", "credit", "c"):
        return "Cr"
    return "Dr"     # default


# ── Top-level parser ────────────────────────────────────────────────────────

def parse_excel_coa(file_path: str) -> MigrationPayload:
    from openpyxl import load_workbook
    from .migrator import Migrator

    path = Path(file_path)
    if not path.exists():
        raise ValueError(f"File not found: {file_path}")

    wb = load_workbook(path, data_only=True, read_only=True)
    sheets = {ws.title.strip().lower(): ws for ws in wb.worksheets}

    payload = MigrationPayload(
        source_type="EXCEL_COA",
        source_label="Excel chart of accounts",
        file_name=path.name,
        file_hash=Migrator.sha256(path),
    )

    # Shape A — two named sheets
    if "groups" in sheets and "ledgers" in sheets:
        payload.groups = _parse_group_sheet(sheets["groups"])
        payload.ledgers = _parse_ledger_sheet(sheets["ledgers"])
        return payload

    # Try first sheet as ledger sheet (Shape B-ish)
    first = wb.worksheets[0]
    rows = _read_rows(first)
    if not rows:
        raise ValueError("Workbook is empty.")

    # If "Type" column distinguishes group vs ledger rows, split them.
    headers = rows[0]
    type_col = _pick(headers, ("type", "row type", "kind"))
    if type_col is not None:
        groups, ledgers = [], []
        for row in rows[1:]:
            if type_col >= len(row):
                continue
            kind = _norm(row[type_col])
            if "group" in kind:
                g = _row_to_group(headers, row, type_col)
                if g:
                    groups.append(g)
            elif "ledger" in kind or "account" in kind:
                l = _row_to_ledger(headers, row)
                if l:
                    ledgers.append(l)
        payload.groups = groups
        payload.ledgers = ledgers
        return payload

    # Otherwise treat the only sheet as a Ledgers sheet (the most common case)
    payload.ledgers = _parse_ledger_sheet(first)
    return payload


# ── Per-sheet parsers ───────────────────────────────────────────────────────

def _parse_group_sheet(ws) -> list[GroupSpec]:
    rows = _read_rows(ws)
    if not rows:
        return []
    headers = rows[0]
    out: list[GroupSpec] = []
    for row in rows[1:]:
        g = _row_to_group(headers, row)
        if g:
            out.append(g)
    return out


def _parse_ledger_sheet(ws) -> list[LedgerSpec]:
    rows = _read_rows(ws)
    if not rows:
        return []
    headers = rows[0]
    out: list[LedgerSpec] = []
    for row in rows[1:]:
        l = _row_to_ledger(headers, row)
        if l:
            out.append(l)
    return out


def _row_to_group(headers, row, type_col: int | None = None) -> Optional[GroupSpec]:
    name_idx   = _pick(headers, _GROUP_HEADERS["name"])
    parent_idx = _pick(headers, _GROUP_HEADERS["parent"])
    nature_idx = _pick(headers, _GROUP_HEADERS["nature"])
    if name_idx is None or name_idx >= len(row):
        return None
    name = str(row[name_idx] or "").strip()
    if not name:
        return None
    return GroupSpec(
        name=name,
        parent_name=(
            str(row[parent_idx] or "").strip() or None
            if parent_idx is not None and parent_idx < len(row) else None
        ),
        nature=(
            _norm_nature(row[nature_idx])
            if nature_idx is not None and nature_idx < len(row) else ""
        ),
    )


def _norm_nature(raw) -> str:
    s = _norm(raw)
    if s in ("asset", "assets"):
        return "ASSET"
    if s in ("liability", "liabilities"):
        return "LIABILITY"
    if s in ("income", "revenue"):
        return "INCOME"
    if s in ("expense", "expenses"):
        return "EXPENSE"
    return s.upper() if s else ""


def _row_to_ledger(headers, row) -> Optional[LedgerSpec]:
    name_idx   = _pick(headers, _LEDGER_HEADERS["name"])
    if name_idx is None or name_idx >= len(row):
        return None
    name = str(row[name_idx] or "").strip()
    if not name:
        return None

    def cell(key) -> str:
        idx = _pick(headers, _LEDGER_HEADERS[key])
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return "" if v is None else str(v).strip()

    return LedgerSpec(
        name=name,
        group_name=cell("group"),
        opening_balance=_parse_amount(cell("opening_balance")),
        opening_type=_norm_dr_cr(cell("opening_type")),
        gstin=cell("gstin") or None,
        pan=cell("pan") or None,
        state_code=cell("state_code") or None,
        account_number=cell("account_number") or None,
        ifsc=cell("ifsc") or None,
        bank_name=cell("bank_name") or None,
        tds_section=cell("tds_section") or None,
        tds_rate=_parse_amount(cell("tds_rate")) or None,
        is_tds_applicable=bool(cell("tds_section")),
    )


def _read_rows(ws) -> list[list]:
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else c for c in row])
    return rows
