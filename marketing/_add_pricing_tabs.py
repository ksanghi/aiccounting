"""
One-shot: add product-specific tabs to config/pricing.xlsx so it
becomes the single canonical source for marketing pages across AHQ,
RWA HQ, HOA HQ, and tradeHQ.

After this runs, the operator edits prices.xlsx and tells the agent
to regenerate pricing.html / checkout.html — the marketing flow
stays manual but the source-of-truth is unambiguous.

Source data:
  - RWA HQ : license_server/plans.py (PLAN_PRICES_RWA_INR, PLAN_FLATS_LIMIT_RWA, PLAN_FEATURES_RWA)
  - HOA HQ : same features as RWA HQ; prices left blank (international TBD)
  - tradeHQ: license_server/plans.py (PLAN_PRICES_THQ_INR; features empty per Phase 1)

We DO NOT touch the existing Tiers / PlanFeatures / Countries sheets
(those drive bake_config.py for the AHQ desktop app and changing
them would break the baker).
"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "config" / "pricing.xlsx"

HDR_FONT  = Font(bold=True)
TITLE_FONT = Font(bold=True, size=12)
DESC_FONT  = Font(italic=True, color="666666")
HDR_FILL  = PatternFill("solid", fgColor="E2E8F0")


def _write_section_header(ws, title: str, description: str):
    """Row 1 title, row 2 description, row 3 blank, row 4 reserved for column headers."""
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=2, column=1, value=description).font = DESC_FONT


def _write_tiers(ws, tiers: list[dict], capacity_label: str):
    """Schema: code | name | seats_allowed | capacity_limit | capacity_unit |
    overage_rate | plan_price_INR | notes  (row 4 = headers, row 5+ = data).
    A separate capacity_unit column makes the sheet self-describing —
    AHQ uses 'transactions', RWA HQ uses 'flats', tradeHQ uses
    'family members'."""
    headers = ["code", "name", "seats_allowed", "capacity_limit",
               "capacity_unit", "overage_rate", "plan_price_INR", "notes"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL

    for i, t in enumerate(tiers, start=5):
        ws.cell(row=i, column=1, value=t["code"])
        ws.cell(row=i, column=2, value=t["name"])
        ws.cell(row=i, column=3, value=t["seats"])
        ws.cell(row=i, column=4, value=t["capacity"])
        ws.cell(row=i, column=5, value=capacity_label)
        ws.cell(row=i, column=6, value=t["overage"])
        ws.cell(row=i, column=7, value=t["price_inr"])
        ws.cell(row=i, column=8, value=t["notes"])

    widths = [12, 14, 16, 16, 16, 14, 16, 70]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + c)].width = w


def _write_plan_features(ws, features: list[dict], tier_codes: list[str]):
    """Schema: category | feature_id | <one column per tier> | upgrade_to.
    Mirrors the AG PlanFeatures sheet exactly so a single bake_config.py
    update later can read all four products with one helper."""
    headers = ["category", "feature_id"] + tier_codes + ["upgrade_to"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL

    for i, f in enumerate(features, start=5):
        ws.cell(row=i, column=1, value=f["category"])
        ws.cell(row=i, column=2, value=f["id"])
        for j, code in enumerate(tier_codes, start=3):
            ws.cell(row=i, column=j, value="Y" if code in f["tiers"] else "")
        ws.cell(row=i, column=len(headers), value=f.get("upgrade_to", ""))

    widths = [16, 28] + [11] * len(tier_codes) + [14]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + c)].width = w


def _write_comms_quota(ws, rows):
    """Schema: tier | notices_per_month | free_promos_per_month | notes
    (row 4 = headers, row 5+ = data).

    notices_per_month   = society broadcast (notice board) volume cap.
    free_promos_per_month = promotional sends a society may make for FREE
        before it must buy promo rights like any promoter (or act as our
        partner/dealer). Beyond-quota promo + the platform-wide per-RESIDENT
        promo frequency cap are governed CENTRALLY by the platform — the
        society approves nothing. Blank cell = unlimited."""
    headers = ["tier", "notices_per_month", "free_promos_per_month", "notes"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
    for i, r in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=r["tier"])
        ws.cell(row=i, column=2, value=r["notices"])
        ws.cell(row=i, column=3, value=r["free_promos"])
        ws.cell(row=i, column=4, value=r["notes"])
    widths = [14, 20, 24, 64]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + c)].width = w


# ─── RWA HQ ──────────────────────────────────────────────────────────────────
# (FREE flats cap inferred from pricing.html — plans.py has the formal cap)
RWAHQ_TIERS = [
    {"code": "FREE",     "name": "Free",     "seats": 1,  "capacity": 300,
     "overage": 0, "price_inr": 0,
     "notes": "Up to 300 flats. Member directory, receipt tracking, notice board, polls, visitor passes, basic reports."},
    {"code": "STANDARD", "name": "Standard", "seats": 2,  "capacity": 1000,
     "overage": 0, "price_inr": 2999,
     "notes": "Up to 1,000 flats. Adds auto-billing + late fees, facilities booking, asset register, advanced reports."},
    {"code": "PRO",      "name": "Pro",      "seats": 5,  "capacity": 2500,
     "overage": 0, "price_inr": 5999,
     "notes": "Up to 2,500 flats. Adds WhatsApp invoices, document storage, vendor management."},
    {"code": "PREMIUM",  "name": "Premium",  "seats": 10, "capacity": None,
     "overage": 0, "price_inr": 14999,
     "notes": "Unlimited flats. Multi-block / phase, priority email support."},
]

# Feature matrix mirrors PLAN_FEATURES_RWA in license_server/plans.py
RWAHQ_FEATURES = [
    {"category": "Core",   "id": "rwa_flat_ledger",         "tiers": ["FREE","STANDARD","PRO","PREMIUM"]},
    {"category": "Core",   "id": "rwa_receipt_tracking",    "tiers": ["FREE","STANDARD","PRO","PREMIUM"]},
    {"category": "Core",   "id": "rwa_member_directory",    "tiers": ["FREE","STANDARD","PRO","PREMIUM"]},
    {"category": "Core",   "id": "rwa_notice_board",        "tiers": ["FREE","STANDARD","PRO","PREMIUM"]},
    {"category": "Core",   "id": "rwa_complaint_tracking",  "tiers": ["FREE","STANDARD","PRO","PREMIUM"]},
    {"category": "Core",   "id": "rwa_broadcast_messaging", "tiers": ["FREE","STANDARD","PRO","PREMIUM"]},
    {"category": "Core",   "id": "rwa_polls",               "tiers": ["FREE","STANDARD","PRO","PREMIUM"]},
    {"category": "Core",   "id": "rwa_visitor_pass",        "tiers": ["FREE","STANDARD","PRO","PREMIUM"]},
    {"category": "Core",   "id": "rwa_basic_reports",       "tiers": ["FREE","STANDARD","PRO","PREMIUM"]},
    {"category": "Billing","id": "rwa_auto_billing",        "tiers": ["STANDARD","PRO","PREMIUM"], "upgrade_to": "STANDARD"},
    {"category": "Billing","id": "rwa_late_fees",           "tiers": ["STANDARD","PRO","PREMIUM"], "upgrade_to": "STANDARD"},
    {"category": "Ops",    "id": "rwa_facilities_booking",  "tiers": ["STANDARD","PRO","PREMIUM"], "upgrade_to": "STANDARD"},
    {"category": "Ops",    "id": "rwa_asset_register",      "tiers": ["STANDARD","PRO","PREMIUM"], "upgrade_to": "STANDARD"},
    {"category": "Reports","id": "rwa_advanced_reports",    "tiers": ["STANDARD","PRO","PREMIUM"], "upgrade_to": "STANDARD"},
    {"category": "Comms",  "id": "rwa_whatsapp_invoices",   "tiers": ["PRO","PREMIUM"], "upgrade_to": "PRO"},
    {"category": "Ops",    "id": "rwa_document_storage",    "tiers": ["PRO","PREMIUM"], "upgrade_to": "PRO"},
    {"category": "Ops",    "id": "rwa_vendor_management",   "tiers": ["PRO","PREMIUM"], "upgrade_to": "PRO"},
]

# Communications volume quotas (per society / month). Confirmed 2026-06-03.
# Blank notices = unlimited. Promotions BEYOND free_promos = society buys promo
# rights like any promoter, OR becomes our partner/dealer. A platform-wide
# per-RESIDENT promo frequency cap (3 promos/week across ALL sources) controls
# inbox congestion and is enforced CENTRALLY — society approves nothing.
# See messenger + marketing-platform design memory.
RWAHQ_COMMS_QUOTA = [
    {"tier": "FREE",     "notices": 50,   "free_promos": 1,
     "notes": "50 society notices/mo; 1 free promo/mo."},
    {"tier": "STANDARD", "notices": 150,  "free_promos": 3,
     "notes": "150 notices/mo; 3 free promos/mo."},
    {"tier": "PRO",      "notices": 400,  "free_promos": 6,
     "notes": "400 notices/mo; 6 free promos/mo."},
    {"tier": "PREMIUM",  "notices": None, "free_promos": 12,
     "notes": "Unlimited notices; 12 free promos/mo."},
]

# ─── HOA HQ (international brand of RWA HQ) ──────────────────────────────────
# Same codebase, same features, country-bound pricing TBD.
# Prices left blank — operator fills when international pricing is decided.
HOAHQ_TIERS = [
    {"code": "FREE",     "name": "Free",     "seats": 1,  "capacity": 300,
     "overage": 0, "price_inr": None,
     "notes": "Up to 300 units. Same feature set as RWA HQ Free; international pricing TBD."},
    {"code": "STANDARD", "name": "Standard", "seats": 2,  "capacity": 1000,
     "overage": 0, "price_inr": None,
     "notes": "Up to 1,000 units. International pricing TBD — fill price_INR (or add Countries_HOAHQ tab) when set."},
    {"code": "PRO",      "name": "Pro",      "seats": 5,  "capacity": 2500,
     "overage": 0, "price_inr": None,
     "notes": "Up to 2,500 units. International pricing TBD."},
    {"code": "PREMIUM",  "name": "Premium",  "seats": 10, "capacity": None,
     "overage": 0, "price_inr": None,
     "notes": "Unlimited units. International pricing TBD."},
]
HOAHQ_FEATURES = RWAHQ_FEATURES  # identical — same product

# ─── tradeHQ ─────────────────────────────────────────────────────────────────
# Phase 1: only FREE + STANDARD sold; PRO/PREMIUM rows omitted (not in market).
TRADEHQ_TIERS = [
    {"code": "FREE",     "name": "Free",     "seats": 1,  "capacity": 1,
     "overage": 0, "price_inr": 0,
     "notes": "1 broker, manual pull, dashboard & holdings, 1 family member."},
    {"code": "STANDARD", "name": "Standard", "seats": 6,  "capacity": None,
     "overage": 0, "price_inr": 2400,
     "notes": "All brokers auto-pull via TOTP, full family wealth view, AI news (cached daily), AHQ cash-flow bridge, up to 6 family members. Rs.200/month."},
]
# Phase 1 has no gated features in code (per plans.py) — empty matrix on purpose.
TRADEHQ_FEATURES: list[dict] = []


def main():
    wb = load_workbook(XLSX)

    # Build the 6 new tabs. If any already exist (re-run), replace cleanly
    # rather than appending duplicates.
    to_build = [
        ("Tiers_RWAHQ",
            "RWA HQ — plan tiers (India)",
            "Source of truth for RWA HQ tier limits and INR prices. Mirrors license_server/plans.py.",
            lambda ws: _write_tiers(ws, RWAHQ_TIERS, "flats")),
        ("PlanFeatures_RWAHQ",
            "RWA HQ — Plan ⇄ Feature matrix",
            "Y = included in that tier. upgrade_to = tier shown to user when they hit a locked feature.",
            lambda ws: _write_plan_features(ws, RWAHQ_FEATURES,
                                            ["FREE","STANDARD","PRO","PREMIUM"])),
        ("CommsQuota_RWAHQ",
            "RWA HQ — communications volume quotas (per society / month)",
            "notices = society broadcast volume; free_promos = promotional sends allowed "
            "FREE before buying promo rights. Beyond-quota promo + a platform-wide "
            "per-RESIDENT promo frequency cap (3/week across ALL sources) are controlled "
            "CENTRALLY by the platform, not the society. Blank = unlimited.",
            lambda ws: _write_comms_quota(ws, RWAHQ_COMMS_QUOTA)),
        ("Tiers_HOAHQ",
            "HOA HQ — plan tiers (international)",
            "Same product as RWA HQ; international brand. plan_price_INR is BLANK — fill when international pricing is decided (or add a Countries_HOAHQ tab for per-currency pricing).",
            lambda ws: _write_tiers(ws, HOAHQ_TIERS, "units")),
        ("PlanFeatures_HOAHQ",
            "HOA HQ — Plan ⇄ Feature matrix",
            "Identical to PlanFeatures_RWAHQ — same codebase. If HOA HQ ever forks, update only this sheet.",
            lambda ws: _write_plan_features(ws, HOAHQ_FEATURES,
                                            ["FREE","STANDARD","PRO","PREMIUM"])),
        ("Tiers_tradeHQ",
            "tradeHQ — plan tiers (India)",
            "Phase 1 ships only FREE + STANDARD. PRO / PREMIUM rows are deliberately omitted.",
            lambda ws: _write_tiers(ws, TRADEHQ_TIERS, "family members")),
        ("PlanFeatures_tradeHQ",
            "tradeHQ — Plan ⇄ Feature matrix",
            "Phase 1 has no code-gated features (license_manager.has_feature() is dormant for tradeHQ). Populate when metered features are added.",
            lambda ws: _write_plan_features(ws, TRADEHQ_FEATURES,
                                            ["FREE","STANDARD"])),
    ]

    for name, title, desc, fill in to_build:
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        _write_section_header(ws, title, desc)
        fill(ws)

    wb.save(XLSX)
    print(f"Updated: {XLSX}")
    print("Sheets now present:")
    for s in wb.sheetnames:
        print(f"  - {s}")


if __name__ == "__main__":
    main()
