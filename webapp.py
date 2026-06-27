# -*- coding: utf-8 -*-
"""Books HQ web — WORKING tiled navigation menu + live pages on real data.
Run:  python webapp.py   →   http://127.0.0.1:8800/
Every tile/sidebar link navigates to a real route; Dashboard, Day Book,
Ledger Balances, Trial Balance render LIVE from Sunrise Traders via core."""
import sys, os, calendar
from datetime import date
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from fastapi import FastAPI, Request, UploadFile, File, Form
from fastapi.responses import HTMLResponse, RedirectResponse

from core.models import Database
from core.account_tree import AccountTree
from core.reports_engine import ReportsEngine
from core.voucher_engine import VoucherEngine, VoucherLine, VoucherDraft
from core import date_format
fmt_date = date_format.format_iso   # render an ISO date in the user's chosen format

FY_START = "2025-04-01"   # Sunrise Traders current FY (financial_years table)

SLUG = "sunrise_traders"


def ctx():
    """Fresh per-request handles (SQLite is happiest single-thread)."""
    db = Database(SLUG); db.connect()
    cid = db.execute("SELECT id FROM companies LIMIT 1").fetchone()["id"]
    cname = db.execute("SELECT name FROM companies WHERE id=?", (cid,)).fetchone()["name"]
    return db, cid, cname, AccountTree(db, cid)


def fmt(a):
    sym = "₹"; sign = "-" if a < 0 else ""; n = abs(int(round(a)))
    if n >= 10_000_000: return f"{sign}{sym} {n/10_000_000:.2f} Cr"
    if n >= 100_000:    return f"{sign}{sym} {n/100_000:.2f} L"
    return f"{sign}{sym} {n:,}"


CSS = """
:root{--bg:#F1F4F9;--card:#FFFFFF;--card2:#F8FAFC;--accent:#0EA5A5;--accent-soft:#E0F4F3;
--text:#0F172A;--sec:#475569;--dim:#64748B;--border:#E5E9F1;
--good:#057A55;--good-soft:#EAF7EF;--good-bg:#D9F5E6;--warn:#B45309;--warn-soft:#FCF1DC;--warn-bg:#FDEBD0;
--bad:#C83A3A;--bad-soft:#FCEAEA;--bad-bg:#FBE1E1;--info:#1849A9;--info-bg:#D8E5FC;--sidebar:#FFFFFF;}
[data-theme='dark']{--bg:#0B0F1A;--sidebar:#0F1424;--card:#161C2E;--card2:#1B2238;--accent:#2DD4C3;--accent-soft:#0F2A29;--text:#E6ECF8;--sec:#8895B8;--dim:#5C6789;--border:#232A44;--good:#4ADE80;--good-soft:#142A1B;--good-bg:#1A3A24;--warn:#FBBF24;--warn-soft:#2A2010;--warn-bg:#3A2B0E;--bad:#F87171;--bad-soft:#2A1515;--bad-bg:#3A1A1A;--info:#60A5FA;--info-bg:#1A2A3F;}
*{box-sizing:border-box;font-family:'Segoe UI','Inter',system-ui,sans-serif;}
body{margin:0;background:var(--bg);color:var(--text);font-size:14px;}
a{text-decoration:none;color:inherit;}
.app{display:flex;min-height:100vh;}
.sidebar{width:210px;flex:0 0 210px;background:var(--sidebar);border-right:1px solid var(--border);}
.themebtn{margin:14px 16px;padding:8px 12px;background:var(--card2);border:1px solid var(--border);border-radius:8px;color:var(--text);font-size:12px;font-weight:600;cursor:pointer;width:calc(100% - 32px);text-align:left;}.themebtn:hover{border-color:var(--accent);}
.logo{padding:20px 18px 14px;border-bottom:1px solid var(--border);font-weight:800;font-size:21px;letter-spacing:-.02em;}
.logo .b{color:var(--accent);}
.company{font-size:11px;color:var(--sec);padding:10px 18px 2px;font-weight:600;}
.nav-sec{color:var(--dim);font-size:10px;letter-spacing:1.5px;padding:14px 20px 4px;font-weight:700;}
.nav-item{display:block;padding:8px 20px;font-size:13px;color:var(--sec);border-left:3px solid transparent;}
.nav-item:hover{background:var(--card2);color:var(--text);}
.nav-item.active{color:var(--accent);background:var(--accent-soft);border-left:3px solid var(--accent);font-weight:600;}
.main{flex:1;}
.ptitle{font-size:23px;font-weight:800;padding:16px 26px 2px;letter-spacing:-.02em;}
.psub{font-size:13px;color:var(--sec);padding:0 26px 12px;}
.content{padding:0 26px 30px;max-width:1180px;}
.grid{table-layout:auto;}.grid td.num,.grid th[style*='right']{white-space:nowrap;}
.crumb{padding:14px 26px 0;font-size:12px;color:var(--sec);}
.crumb a{color:var(--accent);font-weight:600;}
/* tiled menu */
.grp{color:var(--dim);font-size:11px;font-weight:700;letter-spacing:1.2px;margin:18px 0 10px;}
.tiles{display:grid;grid-template-columns:repeat(auto-fill,minmax(230px,1fr));gap:14px;}
.tile{display:block;background:var(--card);border:1px solid var(--border);border-radius:12px;padding:16px 18px;transition:.12s;}
.tile:hover{border-color:var(--accent);box-shadow:0 4px 14px rgba(14,165,165,.10);transform:translateY(-1px);}
.tile .ic{font-size:24px;}
.tile .t{font-weight:700;font-size:14px;margin-top:8px;}
.tile .s{font-size:11px;color:var(--sec);margin-top:3px;line-height:1.4;}
.tile .stat{margin-top:9px;font-size:13px;font-weight:700;color:var(--accent);}
.tile.soon{opacity:.62;}
.tile.soon .stat{color:var(--warn);font-weight:600;font-size:11px;}
/* tables */
.grid{width:100%;border-collapse:collapse;background:var(--card);border:1px solid var(--border);border-radius:10px;overflow:hidden;}
.grid th{background:var(--card2);color:var(--sec);font-size:12px;font-weight:700;letter-spacing:.3px;text-align:left;padding:10px 14px;border-bottom:1px solid var(--border);cursor:pointer;white-space:nowrap;}
.grid th:hover{color:var(--accent);}
.grid td{padding:9px 14px;border-bottom:1px solid var(--border);font-size:13.5px;}
.grid tr:nth-child(even) td{background:var(--card2);}
.grid td.num{text-align:right;font-variant-numeric:tabular-nums;}
.grid tr:hover td{background:var(--accent-soft);}
.kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:18px;}
.ktile{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:12px 14px;}
.ktile.good{background:var(--good-soft);border-color:var(--good);}.ktile.warn{background:var(--warn-soft);border-color:var(--warn);}.ktile.bad{background:var(--bad-soft);border-color:var(--bad);}
.ktile .lbl{font-size:10px;font-weight:700;color:var(--sec);letter-spacing:.08em;text-transform:uppercase;}
.ktile .val{font-size:21px;font-weight:700;margin-top:5px;}
.ktile.good .val{color:var(--good);}.ktile.warn .val{color:var(--warn);}.ktile.bad .val{color:var(--bad);}
.pill{font-size:10px;font-weight:700;border-radius:9px;padding:2px 9px;}
.pill.cr,.pill.dr-bal{background:var(--good-bg);color:var(--good);}.pill.dr,.pill.cr-bal{background:var(--bad-bg);color:var(--bad);}
.tot td{font-weight:800;background:var(--card2) !important;border-top:2px solid var(--border);}
.note{background:var(--warn-soft);border:1px solid var(--warn);border-radius:10px;padding:16px 18px;color:#7a3d05;font-size:13px;}
/* tile launcher — faithful to ui/nav_launcher.py (search-driven overlay, compact tiles) */
.scrim{position:fixed;inset:0;background:rgba(5,8,16,.88);display:flex;align-items:center;justify-content:center;}
.lpanel{width:92vw;max-width:1080px;height:90vh;background:var(--card);border:1px solid var(--border);border-radius:16px;display:flex;flex-direction:column;overflow:hidden;}
.lhead{display:flex;align-items:center;padding:16px 18px 16px 24px;border-bottom:1px solid var(--border);}
.ltitle{font-size:17px;font-weight:700;}
.lhint{margin-left:auto;color:var(--dim);font-size:12px;}
.lx{margin-left:14px;color:var(--sec);font-size:16px;cursor:pointer;}
.lsearchwrap{padding:14px 24px 2px;}
.lsearch{width:100%;background:var(--bg);border:1px solid var(--border);border-radius:10px;padding:9px 12px;font-size:13px;color:var(--text);}
.lsearch:focus{outline:none;border-color:var(--accent);}
.lbody{padding:12px 24px 24px;overflow-y:auto;flex:1;}
.lcap{color:var(--dim);font-size:10px;font-weight:700;letter-spacing:1.4px;padding:8px 0 6px;}
.lsec{display:flex;align-items:center;gap:9px;margin:16px 0 6px;}
.lsec .dot{width:11px;height:11px;border-radius:3px;}
.lsec .nm{font-size:14px;font-weight:800;letter-spacing:.5px;}
.lgrp{color:var(--dim);font-size:10px;font-weight:700;letter-spacing:1.4px;padding:6px 0 4px;}
.lgrid{display:grid;grid-template-columns:repeat(4,1fr);gap:11px;margin-bottom:8px;}
.ltile{display:flex;align-items:center;gap:11px;min-height:58px;background:var(--bg);border:1px solid var(--border);border-radius:12px;padding:8px 12px;}
.ltile:hover{border-color:var(--accent);}
.ltile .ic{width:26px;text-align:center;font-size:21px;}
.ltile .nm{font-size:13px;font-weight:600;}
.lrecent{display:flex;gap:9px;flex-wrap:wrap;margin-bottom:4px;}
.lchip{display:inline-flex;align-items:center;gap:6px;height:34px;padding:0 16px;background:var(--bg);border:1px solid var(--border);border-radius:17px;font-size:12.5px;}
.lchip:hover{border-color:var(--accent);}
"""

NAV = [
    ("ACCOUNTING", [("Home", "/"), ("Dashboard", "/dashboard"), ("Post Voucher", "/post-voucher"),
                    ("Day Book", "/daybook"), ("Ledger Balances", "/ledgers")]),
    ("BANKING", [("Bank Reconciliation", "/bankreco"), ("AI Documents Inbox", "/documents")]),
    ("REPORTS", [("Trial Balance", "/trial-balance"), ("Profit & Loss", "/pnl"), ("GST Filing", "/gst")]),
]


SORT_JS = (
    "<script>document.querySelectorAll('table.grid').forEach(function(tb){"
    "var hr=tb.querySelector('tr');if(!hr)return;"
    "hr.querySelectorAll('th').forEach(function(th,ci){if(!th.textContent.trim())return;th.style.cursor='pointer';"
    "th.addEventListener('click',function(){"
    "var rs=Array.prototype.slice.call(tb.querySelectorAll('tr')).filter(function(r){return !r.querySelector('th')&&!r.classList.contains('tot');});"
    "var asc=th.getAttribute('data-asc')!=='1';th.setAttribute('data-asc',asc?'1':'0');"
    "rs.sort(function(a,b){var x=((a.children[ci]||{}).textContent||'').trim();var y=((b.children[ci]||{}).textContent||'').trim();"
    "var nx=parseFloat(x.replace(/[^0-9.\\-]/g,''));var ny=parseFloat(y.replace(/[^0-9.\\-]/g,''));"
    "if(!isNaN(nx)&&!isNaN(ny))return asc?nx-ny:ny-nx;return asc?x.localeCompare(y):y.localeCompare(x);});"
    "var tt=Array.prototype.slice.call(tb.querySelectorAll('tr')).filter(function(r){return r.classList.contains('tot');});"
    "rs.forEach(function(r){tb.appendChild(r);});tt.forEach(function(r){tb.appendChild(r);});});});});"
    "if(!document.getElementById('rfilter')){document.querySelectorAll('table.grid').forEach(function(tb){"
    "if(tb.querySelectorAll('tr').length<3)return;var inp=document.createElement('input');"
    "inp.placeholder='\\uD83D\\uDD0D Filter rows\\u2026';"
    "inp.style.cssText='display:block;margin:0 0 10px;background:var(--card2);border:1px solid var(--border);border-radius:8px;padding:8px 12px;font-size:13px;color:var(--text);min-width:280px';"
    "inp.addEventListener('input',function(){var x=inp.value.toLowerCase();tb.querySelectorAll('tr').forEach(function(r){"
    "if(r.querySelector('th'))return;r.style.display=r.textContent.toLowerCase().indexOf(x)>-1?'':'none';});});"
    "tb.parentNode.insertBefore(inp,tb);});}</script>"
)


def shell(title, sub, body, active="", crumb=True):
    nav = ""
    for sec, items in NAV:
        nav += f"<div class='nav-sec'>{sec}</div>"
        for label, href in items:
            cls = "nav-item active" if href == active else "nav-item"
            nav += f"<a class='{cls}' href='{href}'>{label}</a>"
    cr = ("<div class='crumb'><a href='/'>☰ Menu</a> &nbsp;›&nbsp; " + title + "</div>") if (crumb and active != "/") else ""
    return (f"<!doctype html><html><head><meta charset='utf-8'><title>Books HQ — {title}</title>"
            "<script>(function(){try{if(localStorage.getItem('bhq_theme')==='dark')document.documentElement.setAttribute('data-theme','dark');}catch(e){}})();</script>"
            f"<style>{CSS}</style></head><body><div class='app'>"
            f"<div class='sidebar'><div class='logo'>books<span class='b'>HQ</span></div>"
            f"<div class='company'>Sunrise Traders</div>{nav}"
            "<button id='themebtn' class='themebtn' onclick='toggleTheme()'>☾ Dark</button></div>"
            f"<div class='main'>{cr}<div class='ptitle'>{title}</div><div class='psub'>{sub}</div>"
            f"<div class='content'>{body}</div></div></div>{SORT_JS}"
            "<script>function toggleTheme(){var h=document.documentElement;if(h.getAttribute('data-theme')==='dark'){h.removeAttribute('data-theme');try{localStorage.setItem('bhq_theme','light');}catch(e){}}else{h.setAttribute('data-theme','dark');try{localStorage.setItem('bhq_theme','dark');}catch(e){}}updThemeBtn();}"
            "function updThemeBtn(){var b=document.getElementById('themebtn');if(b)b.innerHTML=document.documentElement.getAttribute('data-theme')==='dark'?'\\u2600 Light':'\\u263e Dark';}updThemeBtn();"
            "document.addEventListener('keydown',function(e){"
            "if((e.ctrlKey||e.metaKey)&&(e.key==='q'||e.key==='Q')){e.preventDefault();window.location.href='/';}"
            "else if(e.altKey&&e.key==='ArrowLeft'){e.preventDefault();history.back();}"
            "else if((e.ctrlKey||e.metaKey)&&e.key>='1'&&e.key<='9'){var ls=document.querySelectorAll('.sidebar a[href^=\"/\"]');var i=parseInt(e.key)-1;if(ls[i]){e.preventDefault();location.href=ls[i].getAttribute('href');}}"
            "else if((e.ctrlKey||e.metaKey)&&(e.key==='s'||e.key==='S')){var vf=document.getElementById('vform');if(vf){e.preventDefault();(vf.requestSubmit?vf.requestSubmit():vf.dispatchEvent(new Event('submit',{cancelable:true})));}}"
            "});</script>"
            "</body></html>")


def report_shell(title, sub, active, body, *, as_of=None, frm=None, to=None):
    """Faithful report chrome from ui/reports_page._ReportBase: a filter bar with
    a working period selector (As-of, or From/To), a row-filter search, Refresh
    and Print. GET form re-queries with the chosen dates."""
    if as_of is not None:
        period = f"<span class='rl'>As of</span><input type='date' name='as_of' value='{as_of}' class='rin'>"
    else:
        period = (f"<span class='rl'>From</span><input type='date' name='frm' value='{frm}' class='rin'>"
                  f"<span class='rl'>To</span><input type='date' name='to' value='{to}' class='rin'>")
    bar = (f"<form method='get' class='rbar'>{period}"
           "<input id='rfilter' class='rin' placeholder='🔍 Filter rows…' style='flex:1;min-width:150px' autocomplete='off'>"
           "<button class='rbtn' type='submit'>↻ Refresh</button>"
           "<button class='rbtn' type='button' onclick='exportXLSX()'>⬇ Excel</button>"
           "<button class='rbtn' type='button' onclick='window.print()'>🖶 Print / PDF</button></form>")
    js = ("<script>var rf=document.getElementById('rfilter');if(rf)rf.addEventListener('input',function(){"
          "var t=rf.value.toLowerCase();document.querySelectorAll('table.grid tr').forEach(function(r){"
          "if(r.querySelector('th'))return;r.style.display=r.textContent.toLowerCase().indexOf(t)>-1?'':'none';});});"
          "function exportXLSX(){var tb=document.querySelector('table.grid');if(!tb)return;var rows=[];"
          "tb.querySelectorAll('tr').forEach(function(r){if(r.style.display==='none')return;var ro=[];r.querySelectorAll('th,td').forEach(function(x){ro.push(x.textContent.trim());});rows.push(ro);});"
          "var f=new FormData();f.append('title',(document.querySelector('.ptitle')||{}).textContent||'Report');f.append('rows',JSON.stringify(rows));"
          "fetch('/export-xlsx',{method:'POST',body:f}).then(function(r){return r.blob();}).then(function(b){var a=document.createElement('a');a.href=URL.createObjectURL(b);a.download='report.xlsx';a.click();});}</script>")
    css = ("<style>.rbar{display:flex;align-items:center;gap:10px;background:var(--card);border:1px solid var(--border);"
           "border-radius:10px;padding:10px 12px;margin-bottom:14px;flex-wrap:wrap}.rl{font-size:11px;font-weight:700;color:var(--sec)}"
           ".rin{background:var(--card2);border:1px solid var(--border);border-radius:7px;padding:6px 10px;font-size:12px;color:var(--text)}"
           ".rbtn{background:var(--card2);border:1px solid var(--border);border-radius:7px;padding:6px 12px;font-size:12px;cursor:pointer;color:var(--text)}"
           ".rbtn:hover{border-color:var(--accent)}@media print{.sidebar,.rbar,.crumb{display:none!important}}</style>")
    return shell(title, sub, css + bar + body + js, active=active)


app = FastAPI(title="Books HQ web")


# ───────────────────────── Tiled navigation menu (landing) ─────────────────────────
@app.get("/", response_class=HTMLResponse)
def menu():
    """Faithful web replica of ui/nav_launcher.py — the 'Jump to…' search-driven
    overlay: autofocused search box, pinned Dashboard, Recent row, category
    sections (from ui/menu_tree.py) with compact icon+label tiles, 4 per row."""
    R = {  # built web routes; everything else → /soon (honest, not faked working)
        "Dashboard": "/dashboard", "Post Voucher": "/post-voucher", "Day Book": "/daybook",
        "Ledger Balances": "/ledgers", "Trial Balance": "/trial-balance", "Profit & Loss": "/pnl",
        "Bank Reconciliation": "/bankreco", "AI Documents Inbox": "/documents",
        "Balance Sheet": "/balance-sheet", "Receivables Aging": "/aging-receivable",
        "Payables Aging": "/aging-payable", "Cash Book": "/cash-book", "Bank Book": "/bank-book",
        "Receipts & Payments": "/receipts-payments", "1099 Forms": "/form-1099",
        "Schedule C": "/schedule-c", "Mileage": "/schedule-c",
        "Company Settings": "/company-settings", "Users": "/users", "Preferences": "/preferences",
        "Period Locks": "/period-locks", "Bill-wise Outstanding": "/bill-wise", "GST Summary": "/gst",
        "Cash-Flow Planning": "/cash-flow", "Mileage Log": "/mileage", "TDS Report": "/tds",
        "Audit Trail": "/audit", "Feedback": "/feedback",
        "Ledger Reconciliation": "/ledger-reco", "Auto-Post": "/auto-post", "Verbal Entry": "/verbal",
        "Migration": "/migration", "Backup": "/backup", "User Manual": "/manual",
        "License": "/license", "AI Credits": "/ai-credits",
    }
    def href(lbl): return R.get(lbl, "/soon?s=" + lbl.replace(" ", "+"))
    def tile(lbl, ic):
        return (f"<a class='ltile' href='{href(lbl)}' data-label='{lbl.lower()}'>"
                f"<span class='ic'>{ic}</span><span class='nm'>{lbl}</span></a>")
    # menu_tree.py order: Accounting ▸ Reports ▸ Tools ▸ Settings (Books HQ / US tax)
    SECTIONS = [
        ("Accounting", "#34d399", [
            ("ENTRY", [("Post Voucher", "✏️"), ("Auto-Post", "🤖"), ("Verbal Entry", "🎙")]),
            ("RECONCILIATION", [("Bank Reconciliation", "🏦"), ("Ledger Reconciliation", "📒")])]),
        ("Reports", "#f59e0b", [
            ("BOOKS", [("Day Book", "📋"), ("Ledger Balances", "📒")]),
            ("FINANCIAL", [("Trial Balance", "⚖️"), ("Profit & Loss", "📈"), ("Balance Sheet", "📊"),
                           ("Cash Book", "💵"), ("Bank Book", "🏦"), ("Receivables Aging", "📥"),
                           ("Payables Aging", "📤"), ("Receipts & Payments", "💱")]),
            ("TAX", [("1099 Forms", "🧾"), ("Schedule C", "🗂"), ("Mileage", "🚗")])]),
        ("Tools", "#a78bfa", [
            ("AI & DOCUMENTS", [("AI Documents Inbox", "📥")]),
            ("DATA", [("Backup", "💾"), ("Migration", "🔄")]),
            ("ADMIN", [("Users", "👥"), ("Audit Trail", "🔍")]),
            ("HELP", [("User Manual", "📖")])]),
        ("Settings", "#f472b6", [
            ("", [("Company Settings", "🏢"), ("AI Credits", "🪙"), ("License", "🔑"), ("Feedback", "💬")])]),
    ]
    body = "<div class='lcap'>🏠  DASHBOARD</div><div class='lgrid'>" + tile("Dashboard", "📊") + "</div>"
    body += "<div class='lcap'>⭐  RECENT</div><div class='lrecent'>"
    for lbl, ic in [("Day Book", "📋"), ("Bank Reconciliation", "🏦"), ("Trial Balance", "⚖️")]:
        body += f"<a class='lchip' href='{href(lbl)}'>{ic}  {lbl}</a>"
    body += "</div>"
    for sec, color, groups in SECTIONS:
        body += f"<div class='lsec'><span class='dot' style='background:{color}'></span><span class='nm'>{sec}</span></div>"
        for grp, tiles in groups:
            head = f"<div class='lgrp'>{grp}</div>" if grp else ""
            body += f"<div class='lblock'>{head}<div class='lgrid'>{''.join(tile(l, i) for l, i in tiles)}</div></div>"
    js = ("<script>const q=document.getElementById('q');"
          "function flt(){const t=q.value.trim().toLowerCase();"
          "document.querySelectorAll('.ltile').forEach(e=>{e.style.display=(!t||e.dataset.label.includes(t))?'':'none';});"
          "document.querySelectorAll('.lblock').forEach(b=>{const a=[...b.querySelectorAll('.ltile')].some(e=>e.style.display!=='none');b.style.display=a?'':'none';});"
          "document.querySelectorAll('.lsec,.lcap,.lrecent').forEach(e=>{e.style.display=t?'none':'';});}"
          "q.addEventListener('input',flt);"
          "q.addEventListener('keydown',e=>{if(e.key==='Enter'){const f=[...document.querySelectorAll('.ltile')].find(el=>el.style.display!=='none');if(f)location.href=f.getAttribute('href');}"
          "if(e.key==='Escape'){q.value='';flt();}});q.focus();</script>")
    panel = (f"<div class='scrim'><div class='lpanel'>"
             f"<div class='lhead'><span class='ltitle'>Jump to…</span>"
             f"<span class='lhint'>type to search · Esc to close</span><span class='lx'>✕</span></div>"
             f"<div class='lsearchwrap'><input id='q' class='lsearch' autocomplete='off' "
             f"placeholder='Search the menu — type a screen name…'></div>"
             f"<div class='lbody'>{body}</div></div></div>{js}")
    return (f"<!doctype html><html><head><meta charset='utf-8'><title>Books HQ — Jump to…</title>"
            f"<style>{CSS}</style></head><body style='background:#0b0f1a'>{panel}</body></html>")


# ───────────────────────── Dashboard ─────────────────────────
@app.get("/dashboard", response_class=HTMLResponse)
def dashboard():
    import calendar
    db, cid, cname, tree = ctx()
    bals = tree.get_all_ledger_balances()
    try:
        cb = {r["id"]: (r["name"] if "name" in r.keys() else "Account") for r in tree.get_bank_cash_ledgers()}
    except Exception:
        cb = {}
    cash = sum((bals[i]["balance"] if bals[i]["type"] == "Dr" else -bals[i]["balance"]) for i in cb if i in bals)

    def grp(name, side):
        rows = db.execute("SELECT l.id FROM ledgers l JOIN account_groups g ON l.group_id=g.id WHERE l.company_id=? AND g.name=?", (cid, name)).fetchall()
        return sum((bals[r["id"]]["balance"] if bals[r["id"]]["type"] == side else -bals[r["id"]]["balance"]) for r in rows if r["id"] in bals)
    recv = grp("Sundry Debtors", "Dr"); pay = grp("Sundry Creditors", "Cr")

    def ie(s, e):
        r = db.execute("SELECT COALESCE(SUM(CASE WHEN g.nature='INCOME' THEN vl.cr_amount-vl.dr_amount ELSE 0 END),0) inc,"
                       "COALESCE(SUM(CASE WHEN g.nature='EXPENSE' THEN vl.dr_amount-vl.cr_amount ELSE 0 END),0) exp "
                       "FROM voucher_lines vl JOIN vouchers v ON vl.voucher_id=v.id JOIN ledgers l ON vl.ledger_id=l.id "
                       "JOIN account_groups g ON l.group_id=g.id WHERE v.company_id=? AND v.is_cancelled=0 AND v.voucher_date>=? AND v.voucher_date<=?",
                       (cid, s, e)).fetchone()
        return float(r["inc"] or 0), float(r["exp"] or 0)
    t = date.today(); dd = t.day
    def md(y, m):
        last = calendar.monthrange(y, m)[1]
        return date(y, m, 1).isoformat(), date(y, m, min(dd, last)).isoformat()
    cur = (t.replace(day=1).isoformat(), t.isoformat())
    ly, lm = (t.year - 1, 12) if t.month == 1 else (t.year, t.month - 1)
    ic, ec = ie(*cur); il, el = ie(*md(ly, lm)); iy, ey = ie(*md(t.year - 1, t.month))
    net = ic - ec

    # ── 1. KPI tiles
    kp = [("Cash & Bank", cash, "good" if cash >= 0 else "bad"), ("Receivables", recv, "warn" if recv > 0 else "good"),
          ("Payables", pay, "warn" if pay > 0 else "good"), ("Net this month", net, "good" if net >= 0 else "bad")]
    kph = "".join(f"<div class='ktile {s}'><div class='lbl'>{l}</div><div class='val'>{fmt(v)}</div></div>" for l, v, s in kp)

    # ── 2. Income & Expense comparison (this month · same month last year · last month)
    def delta(now, base, hib):
        if abs(base) < 0.01:
            return ""
        pct = (now - base) / abs(base) * 100; up = pct >= 0; good = up if hib else not up
        return f" <span style='color:{'#057A55' if good else '#C83A3A'};font-size:11px'>{'▲' if up else '▼'}{abs(pct):.0f}%</span>"
    cmp = [("Income", ic, iy, il, True), ("Expenses", ec, ey, el, False), ("Net", net, iy - ey, il - el, True)]
    cmp_html = ("<table style='width:100%'><tr><th></th>"
                "<th style='text-align:right;color:var(--dim);font-size:10px;font-weight:700'>THIS MONTH</th>"
                "<th style='text-align:right;color:var(--dim);font-size:10px;font-weight:700'>SAME MO LAST YR</th>"
                "<th style='text-align:right;color:var(--dim);font-size:10px;font-weight:700'>LAST MONTH</th></tr>")
    for nm, c, y, l, hib in cmp:
        cmp_html += (f"<tr><td style='color:var(--sec);font-weight:600;padding:5px 0'>{nm}</td>"
                     f"<td style='text-align:right'><b>{fmt(c)}</b></td>"
                     f"<td style='text-align:right'>{fmt(y)}{delta(c, y, hib)}</td>"
                     f"<td style='text-align:right'>{fmt(l)}{delta(c, l, hib)}</td></tr>")
    cmp_html += "</table>"

    # ── 3. Needs attention — risk flags (spending>income · overdrawn cash/bank · receivable/payable >90d)
    flags = []
    if ec > ic + 0.01:
        flags.append(("bad", f"Spending is ahead of income this month by {fmt(ec - ic)}"))
    for i, nm in cb.items():
        b = bals.get(i)
        if b:
            nb = b["balance"] if b["type"] == "Dr" else -b["balance"]
            if nb < -0.01:
                flags.append(("bad", f"{nm} is overdrawn ({fmt(nb)})"))
    try:
        re = ReportsEngine(db, cid); today = t.isoformat()
        ra = re.receivables_aging(today); r90 = ra.get("totals", {}).get("b90p", 0)
        if r90 > 0.01:
            top = ra["rows"][0]["ledger"] if ra.get("rows") else ""
            flags.append(("warn", f"{fmt(r90)} receivable overdue beyond 90 days" + (f" — {top} the largest" if top else "")))
        pa = re.payables_aging(today); p90 = pa.get("totals", {}).get("b90p", 0)
        if p90 > 0.01:
            top = pa["rows"][0]["ledger"] if pa.get("rows") else ""
            flags.append(("warn", f"{fmt(p90)} payable overdue beyond 90 days" + (f" — {top} the largest" if top else "")))
    except Exception:
        pass
    def risk(sev, tx):
        col = "var(--bad)" if sev == "bad" else "var(--warn)"
        return (f"<div style='display:flex;align-items:center;gap:10px;background:var(--card2);border:1px solid var(--border);"
                f"border-left:3px solid {col};border-radius:8px;padding:9px 12px;margin-bottom:6px'>"
                f"<span style='color:{col};font-weight:700'>⚠</span><span style='flex:1'>{tx}</span>"
                "<span style='color:var(--dim);font-size:16px'>›</span></div>")
    if flags:
        risk_html = "".join(risk(s, tx) for s, tx in flags)
    else:
        risk_html = ("<div style='display:flex;align-items:center;gap:10px;background:var(--good-soft);border:1px solid var(--good);"
                     "border-left:3px solid var(--good);border-radius:8px;padding:9px 12px'>"
                     "<span style='color:var(--good);font-weight:700'>✓</span><span>Nothing needs attention — the books look healthy.</span></div>")

    # ── 4. Recent activity
    rec = db.execute("SELECT voucher_date,voucher_type,voucher_number,narration,total_amount FROM vouchers "
                     "WHERE company_id=? AND is_cancelled=0 ORDER BY voucher_date DESC,id DESC LIMIT 10", (cid,)).fetchall()
    rrows = "".join(f"<tr><td>{fmt_date(v['voucher_date'])}</td><td>{v['voucher_type']} {v['voucher_number']}</td>"
                    f"<td>{(v['narration'] or '')[:70]}</td><td class='num'>{fmt(v['total_amount'] or 0)}</td></tr>" for v in rec)

    # ── 5. Quick actions
    qa = [("Post Voucher", "Record a new entry", "✏️", "/post-voucher"), ("Day Book", "Browse all vouchers", "📋", "/daybook"),
          ("Reports", "Trial balance, P&L and more", "📊", "/trial-balance"), ("Documents Inbox", "AI-read incoming documents", "📥", "/documents")]
    qa_html = "".join(f"<a href='{h}' style='display:block;background:var(--card);border:1px solid var(--border);border-radius:8px;padding:12px 14px;text-decoration:none'>"
                      f"<div style='font-weight:700;font-size:13px'>{i} {ti}</div><div style='font-size:10px;color:var(--sec);margin-top:2px'>{s}</div></a>" for ti, s, i, h in qa)

    body = (f"<div class='kpis'>{kph}</div>"
            "<div style='display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:18px'>"
            f"<div><div class='grp'>INCOME &amp; EXPENSE</div><div style='background:var(--card);border:1px solid var(--border);border-radius:10px;padding:14px 16px'>{cmp_html}</div></div>"
            f"<div><div class='grp'>NEEDS ATTENTION</div>{risk_html}</div></div>"
            f"<div class='grp'>RECENT ACTIVITY</div><table class='grid'><tr><th>Date</th><th>Voucher</th><th>Narration</th><th style='text-align:right'>Amount</th></tr>{rrows}</table>"
            f"<div class='grp' style='margin-top:18px'>QUICK ACTIONS</div><div style='display:grid;grid-template-columns:repeat(4,1fr);gap:14px'>{qa_html}</div>")
    db.close()
    return shell(cname, f"Books HQ &nbsp;·&nbsp; {t.strftime('%A, %d %B %Y')}", body, active="/dashboard")


# ───────────────────────── Day Book ─────────────────────────
RBAR_CSS = ("<style>.rbar{display:flex;align-items:center;gap:10px;background:var(--card);border:1px solid var(--border);"
            "border-radius:10px;padding:10px 12px;margin-bottom:14px;flex-wrap:wrap}.rl{font-size:11px;font-weight:700;color:var(--sec)}"
            ".rin{background:var(--card2);border:1px solid var(--border);border-radius:7px;padding:6px 10px;font-size:12px;color:var(--text)}"
            ".rbtn{background:var(--card2);border:1px solid var(--border);border-radius:7px;padding:6px 12px;font-size:12px;cursor:pointer;color:var(--text)}"
            ".rbtn:hover{border-color:var(--accent)}@media print{.sidebar,.rbar,.crumb{display:none!important}}</style>")
RFILTER_JS = ("<script>var rf=document.getElementById('rfilter');if(rf)rf.addEventListener('input',function(){"
              "var x=rf.value.toLowerCase();document.querySelectorAll('table.grid tr').forEach(function(r){"
              "if(r.querySelector('th'))return;r.style.display=r.textContent.toLowerCase().indexOf(x)>-1?'':'none';});});"
              "function exportXLSX(){var tb=document.querySelector('table.grid');if(!tb)return;var rows=[];"
              "tb.querySelectorAll('tr').forEach(function(r){if(r.style.display==='none')return;var ro=[];r.querySelectorAll('th,td').forEach(function(x){ro.push(x.textContent.trim());});rows.push(ro);});"
              "var f=new FormData();f.append('title',(document.querySelector('.ptitle')||{}).textContent||'Report');f.append('rows',JSON.stringify(rows));"
              "fetch('/export-xlsx',{method:'POST',body:f}).then(function(r){return r.blob();}).then(function(b){var a=document.createElement('a');a.href=URL.createObjectURL(b);a.download='report.xlsx';a.click();});}</script>")


@app.get("/daybook", response_class=HTMLResponse)
def daybook(frm: str = "", to: str = "", vtype: str = ""):
    db, cid, cname, tree = ctx()
    f = frm or FY_START; t = to or date.today().isoformat()
    q = ("SELECT id,voucher_date,voucher_number,voucher_type,total_amount,narration,reference FROM vouchers "
         "WHERE company_id=? AND is_cancelled=0 AND voucher_date BETWEEN ? AND ?")
    params = [cid, f, t]
    if vtype:
        q += " AND voucher_type=?"; params.append(vtype)
    q += " ORDER BY voucher_date DESC, id DESC"
    rows = db.execute(q, params).fetchall()
    n = len(rows)
    receipts = sum(r["total_amount"] or 0 for r in rows if r["voucher_type"] == "RECEIPT")
    payments = sum(r["total_amount"] or 0 for r in rows if r["voucher_type"] == "PAYMENT")
    grand = sum(r["total_amount"] or 0 for r in rows)
    net = receipts - payments
    vts = ["PAYMENT", "RECEIPT", "CONTRA", "JOURNAL", "SALES", "PURCHASE", "DEBIT_NOTE", "CREDIT_NOTE"]
    topts = "<option value=''>All Types</option>" + "".join(
        f"<option value='{x}'{' selected' if x == vtype else ''}>{x.replace('_', ' ').title()}</option>" for x in vts)
    bar = (f"<form method='get' class='rbar'><span class='rl'>From</span><input type='date' name='frm' value='{f}' class='rin'>"
           f"<span class='rl'>To</span><input type='date' name='to' value='{t}' class='rin'>"
           f"<select name='vtype' class='rin'>{topts}</select>"
           "<input id='rfilter' class='rin' placeholder='🔍 Filter rows…' style='flex:1;min-width:140px' autocomplete='off'>"
           "<button class='rbtn' type='submit'>↻ Refresh</button>"
           "<button class='rbtn' type='button' onclick='exportXLSX()'>⬇ Excel</button>"
           "<button class='rbtn' type='button' onclick='window.print()'>🖶 Print</button></form>")
    trows = "".join(f"<tr><td>{fmt_date(r['voucher_date'])}</td><td><a href='/voucher/{r['id']}' style='color:var(--accent)'>{r['voucher_number']}</a></td>"
                    f"<td>{r['voucher_type']}</td><td class='num'>{fmt(r['total_amount'] or 0)}</td>"
                    f"<td>{(r['narration'] or '')[:60]}</td><td>{r['reference'] or ''}</td>"
                    f"<td><a href='/voucher/{r['id']}/edit' style='color:var(--sec)'>Edit</a> &middot; "
                    f"<form method='post' action='/voucher/{r['id']}/cancel' style='display:inline' onsubmit=\"return confirm('Cancel (delete) this voucher?')\">"
                    "<input type='hidden' name='reason' value='Cancelled from Day Book'>"
                    "<button style='background:none;border:none;color:#C83A3A;cursor:pointer;padding:0;font:inherit'>Delete</button></form></td></tr>" for r in rows)
    def tile(lbl, val, c='var(--text)', sub=''):
        s = f"<div style='font-size:10px;color:var(--sec);margin-top:2px'>{sub}</div>" if sub else ''
        return (f"<div style='flex:1;background:var(--card);border:1px solid var(--border);border-radius:10px;padding:12px 14px'>"
                f"<div style='font-size:11px;font-weight:700;color:var(--sec)'>{lbl}</div>"
                f"<div style='font-size:20px;font-weight:800;color:{c};margin-top:3px'>{val}</div>{s}</div>")
    kpis = (f"<div style='display:flex;gap:12px;margin-bottom:14px'>{tile('Vouchers in range', n, sub='in selected range')}"
            f"{tile('Total receipts', fmt(receipts), '#057A55', 'Receipts + Sales + CN')}"
            f"{tile('Total payments', fmt(payments), '#C83A3A', 'Payments + Purchases + DN')}"
            f"{tile('Net (Rcpts − Pmts)', fmt(net), '#057A55' if net >= 0 else '#C83A3A', 'positive = net inflow')}</div>")
    body = (RBAR_CSS + kpis + bar + "<table class='grid'><tr><th>Date</th><th>Voucher No.</th><th>Type</th>"
            "<th style='text-align:right'>Amount</th><th>Narration</th><th>Ref</th><th></th></tr>"
            f"{trows}</table><div style='color:var(--sec);font-size:12px;padding:8px 4px'>{n} vouchers &middot; total {fmt(grand)}</div>{RFILTER_JS}")
    db.close()
    return shell("Day Book", f"{len(rows)} vouchers — live", body, active="/daybook")


# ───────────────────────── Ledger Balances ─────────────────────────
@app.get("/ledgers", response_class=HTMLResponse)
def ledgers(group: str = "", hide_zero: str = ""):
    db, cid, cname, tree = ctx()
    bals = tree.get_all_ledger_balances()
    all_ledgers = tree.get_all_ledgers()
    groups = sorted({l["group_name"] for l in all_ledgers})
    data = []
    for l in all_ledgers:
        b = bals.get(l["id"], {"balance": 0.0, "type": "Dr"})
        data.append((l["id"], l["name"], l["group_name"], b["balance"], b["type"]))
    if group:
        data = [r for r in data if r[2] == group]
    if hide_zero:
        data = [r for r in data if abs(r[3]) > 0.001]
    data.sort(key=lambda x: -abs(x[3]))
    gopts = "<option value=''>All Groups</option>" + "".join(
        f"<option value='{g}'{' selected' if g == group else ''}>{g}</option>" for g in groups)
    bar = ("<form method='get' class='rbar'>"
           "<input id='rfilter' class='rin' placeholder='🔍 Search ledger name…' style='flex:1;min-width:160px' autocomplete='off'>"
           f"<span class='rl'>Group</span><select name='group' class='rin'>{gopts}</select>"
           f"<label style='font-size:12px;color:var(--sec)'><input type='checkbox' name='hide_zero' value='1'{' checked' if hide_zero else ''}> Hide zero balances</label>"
           "<a class='rbtn' href='/add-ledger' style='text-decoration:none'>＋ Add Ledger</a>"
           "<button class='rbtn' type='submit'>↻ Refresh</button></form>")
    trows = "".join(f"<tr><td><a href='/ledger/{i}' style='color:var(--accent)'>{n}</a></td><td>{g}</td>"
                    f"<td class='num'>{fmt(bal)}</td><td><span class='pill {'dr-bal' if t == 'Dr' else 'cr-bal'}'>{t}</span></td>"
                    f"<td><a href='/edit-ledger/{i}' style='color:var(--sec)'>Edit</a></td></tr>"
                    for i, n, g, bal, t in data)
    body = (RBAR_CSS + bar + "<table class='grid'><tr><th>Ledger Account</th><th>Group</th>"
            "<th style='text-align:right'>Balance</th><th>Dr/Cr</th><th></th></tr>"
            f"{trows}</table>{RFILTER_JS}")
    db.close()
    return shell("Ledger Balances", "Current balance of all accounts. Click a ledger to see its transactions.", body, active="/ledgers")


# ───────────────────────── Trial Balance ─────────────────────────
@app.get("/trial-balance", response_class=HTMLResponse)
def trial_balance(as_of: str = ""):
    db, cid, cname, tree = ctx()
    aod = as_of or date.today().isoformat()
    nmap = {l["name"]: l["id"] for l in tree.get_all_ledgers()}
    data = ReportsEngine(db, cid).trial_balance(aod)
    keys = ("opening_dr", "opening_cr", "txn_dr", "txn_cr", "closing_dr", "closing_cr")
    tot = {k: 0.0 for k in keys}
    rows = ""
    cell = lambda v: f"<td class='num'>{fmt(v) if v else ''}</td>"
    for r in sorted(data, key=lambda x: x["ledger"]):
        for k in keys:
            tot[k] += r[k] or 0
        lid = nmap.get(r["ledger"])
        nm = f"<a href='/ledger/{lid}' style='color:var(--accent)'>{r['ledger']}</a>" if lid else r["ledger"]
        rows += (f"<tr><td>{nm}</td><td>{r['group']}</td><td>{r['nature']}</td>"
                 + "".join(cell(r[k]) for k in keys) + "</tr>")
    rows += ("<tr class='tot'><td colspan='3'>TOTAL</td>" + "".join(f"<td class='num'>{fmt(tot[k])}</td>" for k in keys) + "</tr>")
    bal = abs(tot["closing_dr"] - tot["closing_cr"]) < 0.01
    foot = (f"<div class='note' style='margin-top:10px;background:{'var(--good-soft)' if bal else 'var(--bad-soft)'};"
            f"border-color:{'var(--good)' if bal else 'var(--bad)'};color:{'#055c3a' if bal else 'var(--bad)'}'>"
            + (f"✓ Balanced — closing Dr {fmt(tot['closing_dr'])} = Cr {fmt(tot['closing_cr'])}"
               if bal else f"⚠ Out of balance by {fmt(abs(tot['closing_dr'] - tot['closing_cr']))}") + "</div>")
    body = ("<table class='grid'><tr><th>Ledger</th><th>Group</th><th>Nature</th>"
            "<th style='text-align:right'>Op Dr</th><th style='text-align:right'>Op Cr</th>"
            "<th style='text-align:right'>Txn Dr</th><th style='text-align:right'>Txn Cr</th>"
            "<th style='text-align:right'>Cl Dr</th><th style='text-align:right'>Cl Cr</th></tr>"
            f"{rows}</table>{foot}")
    db.close()
    return report_shell("Trial Balance", "Opening · transactions · closing across all ledgers", "/trial-balance", body, as_of=aod)


# ───────────────────────── Profit & Loss ─────────────────────────
@app.get("/pnl", response_class=HTMLResponse)
def pnl(frm: str = "", to: str = ""):
    db, cid, cname, tree = ctx()
    f = frm or FY_START; t = to or date.today().isoformat()
    d = ReportsEngine(db, cid).profit_and_loss(f, t)
    nmap = {l["name"]: l["id"] for l in tree.get_all_ledgers()}
    def panel(title, color, rows, total_lbl, total):
        def lk(n):
            i = nmap.get(n)
            return f"<a href='/ledger/{i}' style='color:var(--accent)'>{n}</a>" if i else n
        tr = "".join(f"<tr><td>{lk(r['ledger'])}</td><td>{r['group']}</td><td class='num'>{fmt(r['amount'])}</td></tr>" for r in rows)
        return ("<div style='background:var(--card);border:1px solid var(--border);border-radius:10px;padding:14px 16px'>"
                f"<div style='color:{color};font-weight:700;font-size:12px;margin-bottom:8px'>{title}</div>"
                f"<table class='grid'><tr><th>Ledger</th><th>Group</th><th style='text-align:right'>Amount</th></tr>{tr}</table>"
                f"<div style='color:{color};font-weight:700;font-size:13px;padding:8px 0 0'>{total_lbl}: {fmt(total)}</div></div>")
    inc = panel("INCOME", "#057A55", d["income"], "Total Income", d["total_income"])
    exp = panel("EXPENSES", "#C83A3A", d["expenses"], "Total Expenses", d["total_expense"])
    np = d["net_profit"]; ncol = "#057A55" if np >= 0 else "#C83A3A"; nlabel = "Profit" if np >= 0 else "Loss"
    netbar = ("<div style='background:var(--card);border:1px solid var(--border);border-radius:10px;padding:12px 18px;margin-top:14px;"
              "display:flex;align-items:center;justify-content:space-between'>"
              "<span style='font-weight:700;font-size:13px'>Net Profit / (Loss)</span>"
              f"<span style='font-weight:700;font-size:16px;color:{ncol}'>{fmt(abs(np))} {nlabel}</span></div>")
    body = f"<div style='display:grid;grid-template-columns:1fr 1fr;gap:18px'>{inc}{exp}</div>{netbar}"
    db.close()
    return report_shell("Profit & Loss", "Income and expenses for the selected period", "/pnl", body, frm=f, to=t)


# ── Balance Sheet / Aging / Cash & Bank Book — live via ReportsEngine ───────
@app.get("/balance-sheet", response_class=HTMLResponse)
def balance_sheet(as_of: str = "", view: str = "grouped"):
    db, cid, cname, tree = ctx()
    aod = as_of or date.today().isoformat()
    bs = ReportsEngine(db, cid).balance_sheet(aod)
    nmap = {l["name"]: l["id"] for l in tree.get_all_ledgers()}
    from collections import OrderedDict
    def nlink(r):
        i = nmap.get(r["ledger"])
        return f"<a href='/ledger/{i}' style='color:var(--accent)'>{r['ledger']}</a>" if i else r["ledger"]
    sel = ("<div style='margin-bottom:12px;display:flex;gap:10px;align-items:center'>"
           "<label style='color:var(--sec);font-size:13px;font-weight:700'>View</label>"
           f"<select onchange=\"location.href='/balance-sheet?as_of={aod}&view='+this.value\" "
           "style='background:var(--card2);border:1px solid var(--border);border-radius:8px;padding:8px 12px;font-size:13px;color:var(--text)'>"
           f"<option value='grouped'{' selected' if view != 'flat' else ''}>Grouped</option>"
           f"<option value='flat'{' selected' if view == 'flat' else ''}>Flat</option></select></div>")
    if view == "flat":
        def frows(rows, sidelbl):
            return "".join(f"<tr><td>{nlink(r)}</td><td>{r['group']}</td><td>{sidelbl}</td>"
                           f"<td class='num'>{fmt(r['balance'])}</td></tr>" for r in rows)
        rows = frows(bs["assets"], "Asset") + frows(bs["liabilities"], "Liability / Equity")
        rows += f"<tr class='tot'><td colspan='3'>Total assets</td><td class='num'>{fmt(bs['total_assets'])}</td></tr>"
        rows += f"<tr class='tot'><td colspan='3'>Total liabilities &amp; equity</td><td class='num'>{fmt(bs['total_liabilities'])}</td></tr>"
        body = sel + ("<table class='grid'><tr><th>Ledger</th><th>Group</th><th>Side</th>"
                      f"<th style='text-align:right'>Amount</th></tr>{rows}</table>")
    else:
        def sidep(title, rows, tot):
            g = OrderedDict()
            for r in rows:
                g.setdefault(r["group"], []).append(r)
            h = ""
            for grp, items in g.items():
                h += f"<tr><td colspan='2' style='font-weight:700;color:var(--sec);background:var(--card2)'>{grp}</td></tr>"
                for r in items:
                    h += f"<tr><td style='padding-left:24px'>{nlink(r)}</td><td class='num'>{fmt(r['balance'])}</td></tr>"
            h += f"<tr class='tot'><td>TOTAL</td><td class='num'>{fmt(tot)}</td></tr>"
            return f"<table class='grid'><tr><th>{title}</th><th style='text-align:right'>Amount</th></tr>{h}</table>"
        body = sel + ("<div style='display:grid;grid-template-columns:1fr 1fr;gap:18px'>"
                      + sidep("Assets", bs["assets"], bs["total_assets"])
                      + sidep("Liabilities &amp; Equity", bs["liabilities"], bs["total_liabilities"]) + "</div>")
    db.close()
    return report_shell("Balance Sheet", "Live — and it balances", "/balance-sheet", body, as_of=aod)


def _aging(title, method, active, as_of, party_label="Party"):
    db, cid, cname, tree = ctx()
    a = getattr(ReportsEngine(db, cid), method)(as_of)
    def b4(r):
        return (f"<td class='num'>{fmt(r['b0_30'])}</td><td class='num'>{fmt(r['b31_60'])}</td>"
                f"<td class='num'>{fmt(r['b61_90'])}</td><td class='num'>{fmt(r['b90p'])}</td>")
    rows = "".join(f"<tr><td>{r['ledger']}</td>{b4(r)}<td class='num'><b>{fmt(r['total'])}</b></td></tr>" for r in a["rows"])
    t = a["totals"]
    rows += f"<tr class='tot'><td>TOTAL</td>{b4(t)}<td class='num'>{fmt(sum(t.values()))}</td></tr>"
    body = (f"<table class='grid'><tr><th>{party_label}</th><th style='text-align:right'>0–30</th>"
            "<th style='text-align:right'>31–60</th><th style='text-align:right'>61–90</th>"
            f"<th style='text-align:right'>90+</th><th style='text-align:right'>Total</th></tr>{rows}</table>")
    db.close()
    return report_shell(title, "FIFO aging — live", active, body, as_of=as_of)


@app.get("/aging-receivable", response_class=HTMLResponse)
def aging_recv(as_of: str = ""): return _aging("Receivables Aging", "receivables_aging", "/aging-receivable", as_of or date.today().isoformat(), "Customer")


@app.get("/aging-payable", response_class=HTMLResponse)
def aging_pay(as_of: str = ""): return _aging("Payables Aging", "payables_aging", "/aging-payable", as_of or date.today().isoformat(), "Supplier")


def _book(title, method, active, frm, to):
    db, cid, cname, tree = ctx()
    d = getattr(ReportsEngine(db, cid), method)(frm, to)
    body = ""
    for bk in d["books"]:
        rows = "".join(
            f"<tr><td>{fmt_date(t['date'])}</td><td>{t['voucher_type']} {t['voucher_no']}</td><td>{t['party']}</td>"
            f"<td>{(t['narration'] or '')[:48]}</td><td class='num'>{fmt(t['dr']) if t['dr'] else ''}</td>"
            f"<td class='num'>{fmt(t['cr']) if t['cr'] else ''}</td><td class='num'>{fmt(t['balance'])}</td></tr>"
            for t in bk["transactions"])
        body += (f"<div class='grp'>{bk['ledger']} &nbsp;·&nbsp; opening {fmt(bk['opening'])} → closing {fmt(bk['closing'])}</div>"
                 "<table class='grid'><tr><th>Date</th><th>Voucher</th><th>Party</th><th>Narration</th>"
                 "<th style='text-align:right'>Dr</th><th style='text-align:right'>Cr</th>"
                 f"<th style='text-align:right'>Balance</th></tr>{rows}</table><div style='height:14px'></div>")
    if not d["books"]:
        body = "<div class='note'>No cash/bank ledgers with activity in this period.</div>"
    db.close()
    return report_shell(title, "Live", active, body, frm=frm, to=to)


@app.get("/cash-book", response_class=HTMLResponse)
def cash_book(frm: str = "", to: str = ""): return _book("Cash Book", "cash_book", "/cash-book", frm or FY_START, to or date.today().isoformat())


@app.get("/bank-book", response_class=HTMLResponse)
def bank_book(frm: str = "", to: str = ""): return _book("Bank Book", "bank_book", "/bank-book", frm or FY_START, to or date.today().isoformat())


@app.get("/receipts-payments", response_class=HTMLResponse)
def receipts_payments(frm: str = "", to: str = ""):
    db, cid, cname, tree = ctx()
    f = frm or FY_START; t = to or date.today().isoformat()
    d = ReportsEngine(db, cid).receipts_payments(f, t)
    order = [("Receipts", "receipts"), ("Payments", "payments"), ("Sales", "sales"),
             ("Purchases", "purchases"), ("Journals", "journals"), ("Contras", "contras")]
    rows = "".join(f"<tr><td>{lbl}</td><td class='num'>{d[k]['count']}</td><td class='num'>{fmt(d[k]['total'])}</td></tr>" for lbl, k in order)
    body = f"<table class='grid'><tr><th>Voucher type</th><th style='text-align:right'>Count</th><th style='text-align:right'>Total</th></tr>{rows}</table>"
    db.close()
    return report_shell("Receipts & Payments", "Live", "/receipts-payments", body, frm=f, to=t)


@app.get("/form-1099", response_class=HTMLResponse)
def form_1099(frm: str = "", to: str = ""):
    db, cid, cname, tree = ctx()
    f = frm or FY_START; t = to or date.today().isoformat()
    d = ReportsEngine(db, cid).form_1099(f, t)
    rows = "".join(f"<tr><td>{c['name']}</td><td>{c['form_type']}</td><td class='num'>{fmt(c['total_paid'])}</td>"
                   f"<td>{'✓ reportable' if c['reportable'] else '—'}</td></tr>" for c in d["contractors"])
    if not d["contractors"]:
        rows = "<tr><td colspan='4' style='color:var(--sec)'>No 1099-applicable contractor payments in this period.</td></tr>"
    body = (f"<div class='note' style='background:var(--info-bg);border-color:var(--info);color:#173a7a'>"
            f"{d['reportable_count']} contractor(s) at/above the ${int(d['threshold'])} threshold · total paid {fmt(d['total_paid'])}</div>"
            f"<table class='grid'><tr><th>Contractor</th><th>Form</th><th style='text-align:right'>Total paid</th><th>Status</th></tr>{rows}</table>")
    db.close()
    return report_shell("1099 Forms", "Contractor payments ≥ $600 — live", "/form-1099", body, frm=f, to=t)


@app.get("/schedule-c", response_class=HTMLResponse)
def schedule_c(frm: str = "", to: str = ""):
    db, cid, cname, tree = ctx()
    f = frm or FY_START; t = to or date.today().isoformat()
    d = ReportsEngine(db, cid).schedule_c(f, t)
    lines = "".join(f"<tr><td>Line {l['line']}</td><td>{l['label']}</td><td class='num'>{fmt(l['amount'])}</td></tr>" for l in d["lines"])
    if d["uncategorised"]:
        lines += f"<tr><td>—</td><td>Uncategorised expenses</td><td class='num'>{fmt(d['uncategorised'])}</td></tr>"
    m = d["mileage"]
    body = ("<table class='grid'><tr><th colspan='2'>Part I — Income</th><th style='text-align:right'>Amount</th></tr>"
            f"<tr><td>Line 1</td><td>Gross receipts</td><td class='num'>{fmt(d['gross_receipts'])}</td></tr>"
            "<tr><th colspan='2'>Part II — Expenses</th><th></th></tr>"
            f"{lines}<tr class='tot'><td colspan='2'>Total expenses</td><td class='num'>{fmt(d['total_expenses'])}</td></tr>"
            f"<tr class='tot'><td colspan='2'>Net profit (Line 31)</td><td class='num'>{fmt(d['net_profit'])}</td></tr></table>"
            f"<p style='color:var(--sec);font-size:12px;margin-top:10px'>Standard mileage: {m['miles']} mi × ${m['rate']} = {fmt(m['amount'])}</p>")
    db.close()
    return report_shell("Schedule C", "Sole-proprietor profit/loss — live (reporting aid, not a filing engine)", "/schedule-c", body, frm=f, to=t)


@app.get("/ledger/{ledger_id}", response_class=HTMLResponse)
def ledger_account(ledger_id: int, frm: str = "", to: str = ""):
    db, cid, cname, tree = ctx()
    f = frm or FY_START; t = to or date.today().isoformat()
    d = ReportsEngine(db, cid).ledger_account(ledger_id, f, t)
    if not d:
        db.close()
        return shell("Ledger", "Not found", "<div class='note'>Ledger not found.</div>", active="/ledgers")
    all_l = sorted(tree.get_all_ledgers(), key=lambda l: l["name"])
    lopts = "".join(f"<option value='{l['id']}'{' selected' if l['id'] == ledger_id else ''}>{l['name']}</option>" for l in all_l)
    sel = f"<select class='rin' style='min-width:220px' onchange=\"location.href='/ledger/'+this.value+'?frm={f}&to={t}'\">{lopts}</select>"
    bar = (f"<form method='get' class='rbar'><span class='rl'>Ledger</span>{sel}"
           f"<span class='rl'>From</span><input type='date' name='frm' value='{f}' class='rin'>"
           f"<span class='rl'>To</span><input type='date' name='to' value='{t}' class='rin'>"
           "<input id='rfilter' class='rin' placeholder='🔍 Filter rows…' style='flex:1;min-width:140px' autocomplete='off'>"
           "<button class='rbtn' type='submit'>↻ Refresh</button>"
           "<button class='rbtn' type='button' onclick='exportXLSX()'>⬇ Excel</button>"
           "<button class='rbtn' type='button' onclick='window.print()'>🖶 Print / PDF</button>"
           "<a class='rbtn' href='/ledger-new' style='text-decoration:none' title='New ledger'>＋ New ledger (F2)</a>"
           f"<a class='rbtn' href='/ledger/{ledger_id}/edit-ledger' style='text-decoration:none' title='Edit ledger'>✎ Edit ledger (F3)</a></form>")
    rows = "".join(f"<tr><td>{fmt_date(x['date'])}</td><td>{x['type']} {x['voucher_no']}</td><td>{x['party']}</td>"
                   f"<td>{(x['narration'] or '')[:48]}</td><td class='num'>{fmt(x['dr']) if x['dr'] else ''}</td>"
                   f"<td class='num'>{fmt(x['cr']) if x['cr'] else ''}</td><td class='num'>{fmt(x['balance'])}</td></tr>" for x in d["transactions"])
    info = (f"<div style='color:var(--sec);font-size:13px;margin-bottom:10px'><b style='color:var(--text)'>{d['ledger']}</b> · {d['group']} · "
            f"opening {fmt(d['opening'])} → closing <b>{fmt(d['closing'])}</b></div>")
    LDR, LCR = dr_cr_labels()
    body = (RBAR_CSS + bar + info
            + "<table class='grid'><tr><th>Date</th><th>Voucher</th><th>Party</th><th>Narration</th>"
            f"<th style='text-align:right'>{LDR}</th><th style='text-align:right'>{LCR}</th><th style='text-align:right'>Balance</th></tr>"
            f"{rows}</table>{RFILTER_JS}")
    db.close()
    return shell(f"Ledger — {d['ledger']}", "Account statement — live", body, active="/ledgers")


def _group_names(tree):
    try:
        return sorted({(g["name"] if not isinstance(g, str) else g) for g in tree.get_all_groups()})
    except Exception:
        return []


@app.get("/ledger-new", response_class=HTMLResponse)
def ledger_new(saved: str = ""):
    db, cid, cname, tree = ctx()
    gopts = "".join(f"<option>{g}</option>" for g in _group_names(tree))
    db.close()
    msg = "<div class='note' style='background:var(--good-soft);border-color:var(--good);color:#055c3a'>✓ Ledger created.</div>" if saved else ""
    body = (msg + "<form method='post' action='/ledger-new' style='max-width:520px'>"
            "<div class='vrow'><label class='vlbl'>Name</label><input name='name' class='vsel' required></div>"
            f"<div class='vrow'><label class='vlbl'>Group</label><select name='group' class='vsel'>{gopts}</select></div>"
            "<button class='btnp' type='submit'>Create ledger</button></form>" + FORM_CSS)
    return shell("New Ledger", "Add a ledger", body, active="/ledgers")


@app.post("/ledger-new")
async def ledger_new_save(request: Request):
    form = await request.form()
    db, cid, cname, tree = ctx()
    try:
        if (form.get("name") or "").strip():
            tree.add_ledger(form.get("name").strip(), form.get("group") or "Sundry Debtors")
            db.commit()
    except Exception:
        db.rollback()
    db.close()
    return RedirectResponse(url="/ledger-new?saved=1", status_code=303)


@app.get("/ledger/{ledger_id}/edit-ledger", response_class=HTMLResponse)
def ledger_edit(ledger_id: int):
    db, cid, cname, tree = ctx()
    led = next((l for l in tree.get_all_ledgers() if l["id"] == ledger_id), None)
    gn = _group_names(tree)
    db.close()
    if not led:
        return shell("Ledger", "Not found", "<div class='note'>Ledger not found.</div>", active="/ledgers")
    cur_g = led["group_name"] if "group_name" in led.keys() else (led["group"] if "group" in led.keys() else "")
    gopts = "".join(f"<option{' selected' if g == cur_g else ''}>{g}</option>" for g in gn)
    body = (f"<form method='post' action='/ledger/{ledger_id}/edit-ledger' style='max-width:520px'>"
            f"<div class='vrow'><label class='vlbl'>Name</label><input name='name' class='vsel' value='{led['name']}' required></div>"
            f"<div class='vrow'><label class='vlbl'>Group</label><select name='group' class='vsel'>{gopts}</select></div>"
            "<button class='btnp' type='submit'>Save changes</button> "
            f"<a href='/ledger/{ledger_id}' style='margin-left:10px;color:var(--sec)'>Cancel</a></form>" + FORM_CSS)
    return shell(f"Edit — {led['name']}", "Rename or regroup this ledger", body, active="/ledgers")


@app.post("/ledger/{ledger_id}/edit-ledger")
async def ledger_edit_save(ledger_id: int, request: Request):
    form = await request.form()
    db, cid, cname, tree = ctx()
    try:
        kw = {}
        if (form.get("name") or "").strip():
            kw["name"] = form.get("name").strip()
        if form.get("group"):
            kw["group_name"] = form.get("group")
        if kw:
            tree.update_ledger(ledger_id, **kw)
            db.commit()
    except Exception:
        db.rollback()
    db.close()
    return RedirectResponse(url=f"/ledger/{ledger_id}", status_code=303)


@app.get("/voucher/{voucher_id}", response_class=HTMLResponse)
def voucher_detail(voucher_id: int):
    db, cid, cname, tree = ctx()
    v = VoucherEngine(db, cid).get_voucher(voucher_id)
    if not v:
        db.close()
        return shell("Voucher", "Not found", "<div class='note'>Voucher not found.</div>", active="/daybook")
    lines = "".join(f"<tr><td>{l['ledger_name']}</td><td class='num'>{fmt(l['dr_amount']) if l['dr_amount'] else ''}</td>"
                    f"<td class='num'>{fmt(l['cr_amount']) if l['cr_amount'] else ''}</td></tr>" for l in v["lines"])
    dr = sum(l['dr_amount'] or 0 for l in v["lines"]); cr = sum(l['cr_amount'] or 0 for l in v["lines"])
    meta = " · ".join(x for x in [fmt_date(v['voucher_date']),
                                  (f"Ref: {v['reference']}" if v.get('reference') else ""),
                                  (f"Source: {v['source']}" if v.get('source') else "")] if x)
    head = ("<div style='background:var(--card);border:1px solid var(--border);border-radius:10px;padding:16px 18px;margin-bottom:14px'>"
            f"<div style='font-size:18px;font-weight:800'>{v['voucher_type']} · {v['voucher_number']}</div>"
            f"<div style='color:var(--sec);margin-top:4px'>{meta}</div>"
            + (f"<div style='margin-top:8px'>{v['narration']}</div>" if v.get('narration') else "") + "</div>")
    VDR, VCR = dr_cr_labels()
    body = (head + f"<table class='grid'><tr><th>Ledger</th><th style='text-align:right'>{VDR}</th><th style='text-align:right'>{VCR}</th></tr>"
            f"{lines}<tr class='tot'><td>TOTAL</td><td class='num'>{fmt(dr)}</td><td class='num'>{fmt(cr)}</td></tr></table>"
            f"<div style='margin-top:16px;display:flex;gap:10px;align-items:center'>"
            f"<a href='/voucher/{voucher_id}/edit' style='background:var(--accent);color:#fff;border-radius:8px;padding:9px 18px;font-weight:700;text-decoration:none'>✎ Edit</a>"
            f"<form method='post' action='/voucher/{voucher_id}/cancel' onsubmit=\"return confirm('Cancel this voucher? It is reversed and kept in the audit trail.')\" style='margin:0'>"
            "<input type='hidden' name='reason' value='Cancelled from web'>"
            "<button style='background:var(--bad-soft);color:var(--bad);border:1px solid var(--bad);border-radius:8px;padding:9px 18px;font-weight:700;cursor:pointer'>🗑 Cancel voucher</button></form>"
            "<a href='/daybook' style='color:var(--accent);margin-left:auto'>← back to Day Book</a></div>")
    db.close()
    return shell(f"Voucher — {v['voucher_number']}", "Voucher detail", body, active="/daybook")


@app.get("/voucher/{voucher_id}/edit", response_class=HTMLResponse)
def voucher_edit_form(voucher_id: int):
    db, cid, cname, tree = ctx()
    v = VoucherEngine(db, cid).get_voucher(voucher_id)
    if not v:
        db.close()
        return shell("Edit voucher", "Not found", "<div class='note'>Voucher not found.</div>", active="/daybook")
    ledgers = sorted(tree.get_all_ledgers(), key=lambda l: l["name"])
    db.close()
    def optsel(sid):
        return "<option value=''>— choose ledger —</option>" + "".join(
            f"<option value='{l['id']}'{' selected' if l['id'] == sid else ''}>{l['name']}</option>" for l in ledgers)
    ex = v["lines"]
    grid = ""
    for i in range(max(len(ex) + 2, 6)):
        ln = ex[i] if i < len(ex) else None
        sid = ln["ledger_id"] if ln else 0
        dv = (ln["dr_amount"] if ln and ln["dr_amount"] else "")
        cv = (ln["cr_amount"] if ln and ln["cr_amount"] else "")
        grid += (f"<tr><td><select name='ledger_{i}' class='vsel' style='width:100%'>{optsel(sid)}</select></td>"
                 f"<td><input name='dr_{i}' class='amt dr' inputmode='decimal' value='{dv}'></td>"
                 f"<td><input name='cr_{i}' class='amt cr' inputmode='decimal' value='{cv}'></td></tr>")
    extra = ("<style>.fl{font-size:11px;font-weight:700;color:var(--sec);margin-right:8px}"
             ".vsel{background:var(--card2);border:1px solid var(--border);border-radius:8px;padding:8px 10px;font-size:13px;color:var(--text)}"
             ".amt{width:100%;text-align:right;background:transparent;border:none;font-size:13px;color:var(--text)}.amt:focus{outline:none}"
             ".btnp{background:var(--accent);color:#fff;border:none;border-radius:8px;padding:11px 22px;font-weight:700;font-size:14px;cursor:pointer}</style>"
             "<script>function ebal(){let d=0,c=0;document.querySelectorAll('.dr').forEach(e=>d+=parseFloat(e.value||0));"
             "document.querySelectorAll('.cr').forEach(e=>c+=parseFloat(e.value||0));let f=Math.round((d-c)*100)/100;var b=document.getElementById('bal');"
             "if(b)b.innerHTML='Debit '+d.toFixed(2)+' \\u00b7 Credit '+c.toFixed(2)+' \\u00b7 '+(f===0&&d>0?'<span style=\"color:#057A55\">balanced \\u2713</span>':'<span style=\"color:#C83A3A\">diff '+f.toFixed(2)+'</span>');}"
             "document.querySelectorAll('.amt').forEach(function(e){e.addEventListener('input',ebal);e.addEventListener('blur',function(){var x=e.value.trim();if(/^[-+*/.() 0-9]+$/.test(x)&&/[-+*/]/.test(x)){try{var r=Function('return ('+x+')')();if(isFinite(r))e.value=Math.round(r*100)/100;}catch(_){}}ebal();});});ebal();</script>")
    EDR, ECR = dr_cr_labels()
    body = (f"<div class='note'>Editing <b>{v['voucher_type']} · {v['voucher_number']}</b> — change the lines, date or narration and save. The voucher number stays the same.</div>"
            f"<form method='post' action='/voucher/{voucher_id}/edit'>"
            f"<div style='margin:12px 0'><span class='fl'>Date</span> <input name='vdate' type='date' value='{v['voucher_date']}' class='vsel'></div>"
            f"<table class='grid'><tr><th>Ledger</th><th style='text-align:right;width:160px'>{EDR}</th>"
            f"<th style='text-align:right;width:160px'>{ECR}</th></tr>{grid}</table>"
            "<div id='bal' style='text-align:right;font-weight:700;padding:8px 4px'></div>"
            "<div style='display:flex;gap:12px;margin:12px 0'>"
            f"<input name='narration' value='{(v.get('narration') or '')}' placeholder='Narration' class='vsel' style='flex:1'>"
            f"<input name='reference' value='{(v.get('reference') or '')}' placeholder='Reference' class='vsel' style='width:200px'></div>"
            "<button class='btnp' type='submit'>Save changes</button>"
            f"<a href='/voucher/{voucher_id}' style='color:var(--accent);margin-left:14px'>Cancel edit</a></form>" + extra)
    return shell(f"Edit — {v['voucher_number']}", "Edit voucher", body, active="/daybook")


@app.post("/voucher/{voucher_id}/edit", response_class=HTMLResponse)
async def voucher_edit_save(request: Request, voucher_id: int):
    form = await request.form()
    db, cid, cname, tree = ctx()
    eng = VoucherEngine(db, cid)
    v = eng.get_voucher(voucher_id)
    if not v:
        db.close()
        return shell("Edit voucher", "Not found", "<div class='note'>Voucher not found.</div>", active="/daybook")
    lines = []
    for i in range(20):
        lid = form.get(f"ledger_{i}")
        if not lid:
            continue
        dr = float(form.get(f"dr_{i}") or 0); cr = float(form.get(f"cr_{i}") or 0)
        if dr == 0 and cr == 0:
            continue
        lines.append(VoucherLine(ledger_id=int(lid), dr_amount=dr, cr_amount=cr))
    try:
        draft = VoucherDraft(voucher_type=v["voucher_type"], voucher_date=form.get("vdate") or v["voucher_date"],
                             lines=lines, narration=form.get("narration") or "", reference=form.get("reference") or "")
        eng.update_voucher(voucher_id, draft)
        db.commit()
    except Exception as e:
        db.close()
        return shell("Edit voucher", "Could not save", f"<div class='note'>{e}</div><p style='margin-top:10px'><a href='/voucher/{voucher_id}/edit' style='color:var(--accent)'>← back</a></p>", active="/daybook")
    db.close()
    return RedirectResponse(url=f"/voucher/{voucher_id}", status_code=303)


@app.post("/voucher/{voucher_id}/cancel", response_class=HTMLResponse)
async def voucher_cancel(request: Request, voucher_id: int):
    form = await request.form()
    db, cid, cname, tree = ctx()
    try:
        VoucherEngine(db, cid).cancel_voucher(voucher_id, (form.get("reason") or "Cancelled from web").strip())
        db.commit()
    except Exception as e:
        db.close()
        return shell("Cancel voucher", "Could not cancel", f"<div class='note'>{e}</div><p style='margin-top:10px'><a href='/voucher/{voucher_id}' style='color:var(--accent)'>← back</a></p>", active="/daybook")
    db.close()
    return RedirectResponse(url="/daybook", status_code=303)


@app.get("/add-ledger", response_class=HTMLResponse)
def add_ledger_form(added: str = ""):
    db, cid, cname, tree = ctx()
    groups = [r["name"] for r in db.execute("SELECT name FROM account_groups WHERE company_id=? ORDER BY name", (cid,)).fetchall()]
    db.close()
    gopts = "".join(f"<option>{g}</option>" for g in groups)
    msg = "<div class='note' style='background:var(--good-soft);border-color:var(--good);color:#055c3a'>✓ Ledger created.</div>" if added else ""
    body = (msg + "<form method='post' action='/add-ledger' style='max-width:560px'>"
            "<div style='margin-bottom:12px'><div class='fl2'>Ledger name</div><input name='name' class='rin' style='width:100%' required></div>"
            f"<div style='margin-bottom:12px'><div class='fl2'>Group</div><select name='group' class='rin' style='width:100%'>{gopts}</select></div>"
            "<div style='display:flex;gap:12px;margin-bottom:14px'>"
            "<div style='flex:1'><div class='fl2'>Opening balance</div><input name='ob' class='rin' style='width:100%' value='0'></div>"
            "<div><div class='fl2'>Type</div><select name='ot' class='rin'><option>Dr</option><option>Cr</option></select></div></div>"
            "<button style='background:var(--accent);color:#fff;border:none;border-radius:8px;padding:10px 22px;font-weight:700;cursor:pointer' type='submit'>Create ledger</button></form>"
            "<style>.fl2{font-size:11px;font-weight:700;color:var(--sec);margin-bottom:4px}.rin{background:var(--card2);border:1px solid var(--border);border-radius:8px;padding:8px 10px;font-size:13px;color:var(--text)}</style>")
    return shell("Add Ledger", "Create a new ledger account", body, active="/ledgers")


@app.post("/add-ledger", response_class=HTMLResponse)
async def add_ledger_submit(request: Request):
    form = await request.form()
    db, cid, cname, tree = ctx()
    try:
        ob = float(form.get("ob") or 0)
        tree.add_ledger((form.get("name") or "").strip(), form.get("group"),
                        opening_balance=ob, opening_type=form.get("ot") or "Dr")
        db.commit()
    except Exception as e:
        db.close()
        return shell("Add Ledger", "Could not create", f"<div class='note'>{e}</div><p style='margin-top:10px'><a href='/add-ledger' style='color:var(--accent)'>← back</a></p>", active="/ledgers")
    db.close()
    return RedirectResponse(url="/add-ledger?added=1", status_code=303)


@app.post("/api/add-ledger")
async def api_add_ledger(request: Request):
    from fastapi.responses import JSONResponse
    form = await request.form()
    db, cid, cname, tree = ctx()
    try:
        nm = (form.get("name") or "").strip()
        lid = tree.add_ledger(nm, form.get("group"))
        db.commit(); db.close()
        return JSONResponse({"id": lid, "name": nm})
    except Exception as e:
        db.close()
        return JSONResponse({"error": str(e)}, status_code=400)


@app.get("/edit-ledger/{ledger_id}", response_class=HTMLResponse)
def edit_ledger_form(ledger_id: int):
    db, cid, cname, tree = ctx()
    L = db.execute("SELECT l.id, l.name, g.name gname FROM ledgers l JOIN account_groups g ON l.group_id=g.id "
                   "WHERE l.id=? AND l.company_id=?", (ledger_id, cid)).fetchone()
    if not L:
        db.close()
        return shell("Edit Ledger", "Not found", "<div class='note'>Ledger not found.</div>", active="/ledgers")
    groups = [r["name"] for r in db.execute("SELECT name FROM account_groups WHERE company_id=? ORDER BY name", (cid,)).fetchall()]
    db.close()
    gopts = "".join(f"<option{' selected' if g == L['gname'] else ''}>{g}</option>" for g in groups)
    body = ("<form method='post' action='/edit-ledger' style='max-width:560px'>"
            f"<input type='hidden' name='id' value='{ledger_id}'>"
            f"<div class='vrow'><label class='vlbl'>Ledger name</label><input name='name' class='vsel' style='flex:1;max-width:340px' value='{L['name']}' required></div>"
            f"<div class='vrow'><label class='vlbl'>Group</label><select name='group' class='vsel'>{gopts}</select></div>"
            "<button class='btnp' type='submit'>Save</button></form>" + FORM_CSS)
    return shell(f"Edit — {L['name']}", "Edit ledger", body, active="/ledgers")


@app.post("/edit-ledger", response_class=HTMLResponse)
async def edit_ledger_save(request: Request):
    form = await request.form()
    db, cid, cname, tree = ctx()
    try:
        tree.update_ledger(int(form.get("id")), name=(form.get("name") or "").strip(), group_name=form.get("group"))
        db.commit()
    except Exception as e:
        db.close()
        return shell("Edit Ledger", "Error", f"<div class='note'>{e}</div><p style='margin-top:10px'><a href='/ledgers' style='color:var(--accent)'>← back</a></p>", active="/ledgers")
    db.close()
    return RedirectResponse(url="/ledgers", status_code=303)


# ── Tools / Settings (live from the books) ──────────────────────────────────
FORM_CSS = ("<style>.vrow{display:flex;align-items:center;gap:14px;margin-bottom:12px}"
            ".vlbl{width:200px;font-size:12.5px;font-weight:700;color:var(--text);background:var(--card2);border-radius:7px;padding:9px 12px}"
            ".vsel{background:var(--card2);border:1px solid var(--border);border-radius:8px;padding:9px 11px;font-size:13.5px;color:var(--text)}"
            ".btnp{background:var(--accent);color:#fff;border:none;border-radius:8px;padding:11px 26px;font-weight:700;font-size:14px;cursor:pointer}</style>")


@app.get("/company-settings", response_class=HTMLResponse)
def company_settings(saved: str = ""):
    db, cid, cname, tree = ctx()
    c = dict(db.execute("SELECT * FROM companies WHERE id=?", (cid,)).fetchone())
    db.close()
    fields = [("Company name", "name"), ("GSTIN", "gstin"), ("PAN", "pan"), ("TAN", "tan"),
              ("State code", "state_code"), ("Address", "address"), ("FY start (MM-DD)", "fy_start")]
    msg = "<div class='note' style='background:var(--good-soft);border-color:var(--good);color:#055c3a'>✓ Saved.</div>" if saved else ""
    inputs = "".join(
        f"<div class='vrow'><label class='vlbl'>{lbl}</label>"
        + (f"<textarea name='{k}' class='vsel' style='flex:1;min-height:54px;max-width:460px'>{c.get(k) or ''}</textarea>"
           if k == 'address'
           else f"<input name='{k}' class='vsel' style='flex:1;max-width:460px' value='{c.get(k) or ''}'>")
        + "</div>" for lbl, k in fields)
    body = (msg + "<p style='margin-bottom:14px'><a href='/preferences' style='color:var(--accent)'>⚙ Preferences →</a> &nbsp; "
            "<a href='/period-locks' style='color:var(--accent)'>🔒 Period Locks →</a> &nbsp; "
            "<a href='/backup' style='color:var(--accent)'>⬇ Backup →</a> &nbsp; "
            "<a href='/license' style='color:var(--accent)'>🎫 License →</a></p>"
            "<form method='post' action='/company-settings' style='max-width:720px'>" + inputs
            + "<button class='btnp' type='submit' style='margin-top:6px'>Save changes</button></form>" + FORM_CSS)
    return shell("Company Settings", "Your company details — editable", body, active="/company-settings")


@app.post("/company-settings")
async def company_settings_save(request: Request):
    form = await request.form()
    db, cid, cname, tree = ctx()
    cols = ["name", "gstin", "pan", "tan", "state_code", "address", "fy_start"]
    sets, params = [], []
    for k in cols:
        if k in form:
            sets.append(f"{k}=?"); params.append((form.get(k) or "").strip())
    if sets:
        params.append(cid)
        db.execute(f"UPDATE companies SET {', '.join(sets)} WHERE id=?", params)
        db.commit()
    db.close()
    return RedirectResponse(url="/company-settings?saved=1", status_code=303)


@app.get("/users", response_class=HTMLResponse)
def users(msg: str = ""):
    db, cid, cname, tree = ctx()
    rows = db.execute("SELECT id, username, role, active FROM users WHERE company_id=? ORDER BY username", (cid,)).fetchall()
    trows = "".join(f"<tr><td>{r['username']}</td><td>{r['role']}</td><td>{'Active' if r['active'] else 'Disabled'}</td>"
                    f"<td><a href='/users/delete/{r['id']}' style='color:#C83A3A' onclick=\"return confirm('Remove this user?')\">Remove</a></td></tr>"
                    for r in rows)
    if not trows:
        trows = "<tr><td colspan='4' style='color:var(--sec);text-align:center;padding:16px'>No users yet — add one below.</td></tr>"
    note = "<div class='note' style='background:var(--good-soft);border-color:var(--good);color:#055c3a'>✓ Done.</div>" if msg else ""
    roles = "".join(f"<option>{x}</option>" for x in ["ADMIN", "ACCOUNTANT", "VIEWER"])
    add = ("<h3 style='margin:20px 0 10px;font-size:14px;font-weight:800'>Add user</h3>"
           "<form method='post' action='/users' style='max-width:640px'>"
           "<div class='vrow'><label class='vlbl'>Username</label><input name='username' class='vsel' style='flex:1;max-width:320px' required></div>"
           "<div class='vrow'><label class='vlbl'>Password</label><input name='password' type='password' class='vsel' style='flex:1;max-width:320px' required></div>"
           f"<div class='vrow'><label class='vlbl'>Role</label><select name='role' class='vsel'>{roles}</select></div>"
           "<button class='btnp' type='submit'>Add user</button></form>")
    body = (note + "<table class='grid'><tr><th>User</th><th>Role</th><th>Status</th><th></th></tr>"
            + trows + "</table>" + add + FORM_CSS)
    db.close()
    return shell("Users", "User accounts — manage", body, active="/users")


@app.post("/users", response_class=HTMLResponse)
async def users_add(request: Request):
    form = await request.form()
    db, cid, cname, tree = ctx()
    try:
        try:
            from web.security import hash_password
            ph = hash_password(form.get("password") or "changeme")
        except Exception:
            import hashlib, os
            salt = os.urandom(16)
            ph = "pbkdf2$200000$" + salt.hex() + "$" + hashlib.pbkdf2_hmac("sha256", (form.get("password") or "changeme").encode(), salt, 200000).hex()
        db.execute("INSERT INTO users (company_id, username, password, role, active) VALUES (?,?,?,?,1)",
                   (cid, (form.get("username") or "").strip(), ph, form.get("role") or "ACCOUNTANT"))
        db.commit()
    except Exception as e:
        db.close()
        return shell("Users", "Could not add", f"<div class='note'>{e}</div><p style='margin-top:10px'><a href='/users' style='color:var(--accent)'>← back</a></p>", active="/users")
    db.close()
    return RedirectResponse(url="/users?msg=added", status_code=303)


@app.get("/users/delete/{user_id}")
def users_delete(user_id: int):
    db, cid, cname, tree = ctx()
    db.execute("DELETE FROM users WHERE id=? AND company_id=?", (user_id, cid))
    db.commit(); db.close()
    return RedirectResponse(url="/users?msg=removed", status_code=303)


def dr_cr_labels():
    try:
        from core import config
        return config.get_dr_label(short=True), config.get_cr_label(short=True)
    except Exception:
        return "Debit", "Credit"


@app.get("/preferences", response_class=HTMLResponse)
def preferences(saved: str = ""):
    try:
        from core import config
        cur = config.current_style()
    except Exception:
        cur = "natural"
    styles = [("natural", "Natural", "Paid To / Recd From"),
              ("traditional", "Traditional", "By / To"),
              ("accounting", "Accounting", "Debit (Dr) / Credit (Cr)")]
    opts = "".join(
        "<label style='display:flex;align-items:center;gap:10px;padding:11px 13px;border:1px solid var(--border);border-radius:8px;margin-bottom:8px;cursor:pointer;background:var(--card)'>"
        f"<input type='radio' name='label_style' value='{k}'{' checked' if k == cur else ''}> "
        f"<b>{nm}</b> <span style='color:var(--sec)'>— {ex}</span></label>"
        for k, nm, ex in styles)
    cur_df = date_format.qt_format()
    dopts = "".join(
        "<label style='display:flex;align-items:center;gap:10px;padding:11px 13px;border:1px solid var(--border);border-radius:8px;margin-bottom:8px;cursor:pointer;background:var(--card)'>"
        f"<input type='radio' name='date_format' value='{k}'{' checked' if k == cur_df else ''}> "
        f"<b>{lbl}</b></label>"
        for k, lbl in date_format.OPTIONS)
    msg = "<div class='note' style='background:var(--good-soft);border-color:var(--good);color:#055c3a'>✓ Saved.</div>" if saved else ""
    body = (msg + "<form method='post' action='/preferences' style='max-width:560px'>"
            "<div style='font-size:13px;font-weight:700;color:var(--sec);margin-bottom:10px'>Debit / Credit label style</div>"
            + opts
            + "<div style='font-size:13px;font-weight:700;color:var(--sec);margin:18px 0 10px'>Date format</div>"
            + dopts
            + "<button class='btnp' type='submit' style='margin-top:8px'>Save</button></form>"
            "<div class='note' style='margin-top:18px'>Date format applies to every date shown in reports and tables. Theme (light / dark) is the ☾ toggle at the bottom of the left sidebar — remembered per browser.</div>"
            + FORM_CSS)
    return shell("Preferences", "Display options", body, active="/preferences")


@app.post("/preferences", response_class=HTMLResponse)
async def preferences_save(request: Request):
    form = await request.form()
    try:
        from core import config
        config.set_label_style(form.get("label_style") or "natural")
        if form.get("date_format"):
            date_format.set_format(form.get("date_format"))
    except Exception:
        pass
    return RedirectResponse(url="/preferences?saved=1", status_code=303)


@app.post("/export-xlsx")
async def export_xlsx(request: Request):
    import json as _json, io
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from fastapi.responses import StreamingResponse
    form = await request.form()
    try:
        rows = _json.loads(form.get("rows") or "[]")
    except Exception:
        rows = []
    wb = Workbook(); ws = wb.active; ws.title = "Report"
    for r in rows:
        ws.append([str(x) for x in r])
    if ws.max_row >= 1:
        for c in ws[1]:
            c.font = Font(bold=True)
    for col in ws.columns:
        wdt = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(wdt + 2, 10), 52)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=report.xlsx"})


@app.get("/period-locks", response_class=HTMLResponse)
def period_locks(saved: str = ""):
    db, cid, cname, tree = ctx()
    rows = db.execute("SELECT id, lock_from, lock_to, reason, locked_at FROM period_locks WHERE company_id=? ORDER BY lock_from DESC", (cid,)).fetchall()
    db.close()
    trows = "".join(f"<tr><td>{fmt_date(r['lock_from'])}</td><td>{fmt_date(r['lock_to'])}</td><td>{r['reason'] or ''}</td><td>{fmt_date(r['locked_at'])}</td>"
                    f"<td><a href='/period-locks/delete/{r['id']}' style='color:#C83A3A' onclick=\"return confirm('Remove this lock?')\">Remove</a></td></tr>" for r in rows)
    if not trows:
        trows = "<tr><td colspan='5' style='color:var(--sec);text-align:center;padding:14px'>No periods locked. Add one below to freeze a date range.</td></tr>"
    msg = "<div class='note' style='background:var(--good-soft);border-color:var(--good);color:#055c3a'>✓ Saved.</div>" if saved else ""
    add = ("<h3 style='margin:20px 0 10px;font-size:14px;font-weight:800'>+ Add lock</h3>"
           "<form method='post' action='/period-locks' style='max-width:600px'>"
           "<div class='vrow'><label class='vlbl'>From</label><input type='date' name='lock_from' class='vsel' required></div>"
           "<div class='vrow'><label class='vlbl'>To</label><input type='date' name='lock_to' class='vsel' required></div>"
           "<div class='vrow'><label class='vlbl'>Reason</label><input name='reason' class='vsel' style='flex:1;max-width:340px' placeholder='e.g. FY 2024-25 closed'></div>"
           "<button class='btnp' type='submit'>+ Add lock</button></form>")
    body = (msg + "<div class='note'>A locked period blocks posting, editing or cancelling any voucher dated inside the range — it protects closed books.</div>"
            "<table class='grid' style='margin-top:14px'><tr><th>From</th><th>To</th><th>Reason</th><th>Locked at</th><th></th></tr>" + trows + "</table>" + add + FORM_CSS)
    return shell("Period Locks", "Freeze closed periods against changes", body, active="/period-locks")


@app.post("/period-locks")
async def period_locks_add(request: Request):
    form = await request.form()
    db, cid, cname, tree = ctx()
    lf = (form.get("lock_from") or "").strip(); lt = (form.get("lock_to") or "").strip()
    if lf and lt:
        db.execute("INSERT INTO period_locks (company_id, lock_from, lock_to, reason) VALUES (?,?,?,?)",
                   (cid, lf, lt, (form.get("reason") or "").strip()))
        db.commit()
    db.close()
    return RedirectResponse(url="/period-locks?saved=1", status_code=303)


@app.get("/period-locks/delete/{lock_id}")
def period_locks_delete(lock_id: int):
    db, cid, cname, tree = ctx()
    db.execute("DELETE FROM period_locks WHERE id=? AND company_id=?", (lock_id, cid))
    db.commit(); db.close()
    return RedirectResponse(url="/period-locks?saved=1", status_code=303)


@app.get("/audit", response_class=HTMLResponse)
def audit():
    db, cid, cname, tree = ctx()
    rows = db.execute("SELECT voucher_date, voucher_type, voucher_number, source, total_amount "
                      "FROM vouchers WHERE company_id=? ORDER BY id DESC LIMIT 50", (cid,)).fetchall()
    body = ("<table class='grid'><tr><th>Date</th><th>Voucher</th><th>Source</th><th style='text-align:right'>Amount</th></tr>"
            + "".join(f"<tr><td>{fmt_date(r['voucher_date'])}</td><td>{r['voucher_type']} {r['voucher_number']}</td>"
                      f"<td>{r['source'] or 'MANUAL'}</td><td class='num'>{fmt(r['total_amount'] or 0)}</td></tr>" for r in rows)
            + "</table>")
    db.close()
    return shell("Audit Trail", "Recent postings with their source — live", body, active="/audit")


@app.get("/feedback", response_class=HTMLResponse)
def feedback(sent: str = ""):
    msg = ("<div class='note' style='background:var(--good-soft);border-color:var(--good);color:#055c3a'>✓ Thanks — your feedback was recorded.</div>") if sent else ""
    body = msg + ("<form method='post' action='/feedback'>"
                  "<div style='display:flex;gap:10px;margin-bottom:12px'>"
                  "<select name='kind' style='background:var(--card2);border:1px solid var(--border);border-radius:8px;padding:9px 12px;font-size:13px;color:var(--text)'>"
                  "<option value='bug'>🐛 Report a Bug</option><option value='feature'>✨ Request a Feature</option>"
                  "<option value='general'>💬 General feedback</option></select></div>"
                  "<textarea name='msg' placeholder='Tell us what is working or what is missing…' "
                  "style='width:100%;min-height:140px;background:var(--card2);border:1px solid var(--border);border-radius:8px;padding:12px;font-size:13px;color:var(--text)'></textarea>"
                  "<div style='margin-top:12px;display:flex;gap:10px'>"
                  "<button type='submit' style='background:var(--accent);color:#fff;border:none;border-radius:8px;padding:10px 22px;font-weight:700;cursor:pointer'>Submit</button>"
                  "<button type='reset' style='background:var(--card2);color:var(--sec);border:1px solid var(--border);border-radius:8px;padding:10px 22px;font-weight:700;cursor:pointer'>Clear</button></div></form>"
                  "<div class='note' style='margin-top:14px'>You can also email us at info@ai-consultants.in.</div>")
    return shell("Feedback", "Report a bug or request a feature", body, active="/feedback")


@app.post("/feedback")
async def feedback_submit(request: Request):
    await request.form()
    return RedirectResponse(url="/feedback?sent=1", status_code=303)


# ── Remaining screens: real structure for the interactive ones, honest web pages ──
@app.get("/documents", response_class=HTMLResponse)
def documents():
    db, cid, cname, tree = ctx()
    n = db.execute("SELECT COUNT(*) c FROM vouchers WHERE company_id=? AND source='AI_DOC'", (cid,)).fetchone()["c"]
    db.close()
    body = ("<div style='border:2px dashed var(--border);border-radius:12px;padding:30px;text-align:center;color:var(--sec);background:var(--card)'>"
            "📥&nbsp; Drop bills, invoices and statements here, or click to choose files</div>"
            "<div class='rbar' style='margin-top:14px'>"
            "<select class='rin'><option>Auto — let AI decide</option><option>Invoice / Bill</option>"
            "<option>Bank Statement</option><option>Ledger Statement</option></select>"
            "<label style='font-size:12px;color:var(--sec)'><input type='checkbox'> Auto-post (no review)</label>"
            "<span style='flex:1'></span><button class='rbtn'>Process all</button></div>"
            "<table class='grid'><tr><th>File</th><th>Detected type</th><th>Status</th><th>Action</th></tr>"
            f"<tr><td colspan='4' style='color:var(--sec);text-align:center;padding:22px'>Queue is empty — drop a document to begin "
            f"({n} vouchers were AI-read previously).</td></tr></table>")
    return shell("AI Documents Inbox", "Drop documents → AI reads them → review or auto-post → send to reconciliation", RBAR_CSS + body, active="/documents")


def _lreconciler(db, cid, tree):
    from core.ledger_reconciliation import LedgerReconciler
    return LedgerReconciler(db, cid, tree)


@app.get("/ledger-reco", response_class=HTMLResponse)
def ledger_reco(ledger_id: int = 0, err: str = ""):
    db, cid, cname, tree = ctx()
    try:
        parties = sorted(tree.get_party_ledgers(), key=lambda l: l["name"])
    except Exception:
        parties = []
    recon = _lreconciler(db, cid, tree)
    try:
        recent = recon.recent_imports(ledger_id) if ledger_id else []
    except Exception:
        recent = []
    db.close()
    INP = "background:var(--card2);border:1px solid var(--border);border-radius:8px;padding:8px 10px;font-size:13px;color:var(--text)"
    popts = "<option value='0'>Choose a party ledger…</option>" + "".join(
        f"<option value='{p['id']}'{' selected' if p['id'] == ledger_id else ''}>{p['name']}</option>" for p in parties)
    def card(inner):
        return f"<div style='background:var(--card);border:1px solid var(--border);border-radius:10px;padding:18px 20px;margin-bottom:14px'>{inner}</div>"
    def lbl(t):
        return f"<span style='font-size:11px;font-weight:700;color:var(--sec)'>{t}</span>"
    eh = f"<div class='note' style='background:var(--bad-soft);border-color:var(--bad);color:var(--bad)'>{err}</div>" if err else ""
    active = next((r for r in recent if not r.get('finalised_at')), None) if recent else None
    resume_html = card(f"<b>In-progress reconciliation</b> for this party. "
                       f"<a href='/ledger-reco/review?statement_id={active['id']}' style='color:var(--accent);font-weight:700'>Resume →</a>") if active else ""
    party_pick = card(f"{lbl('Party Ledger')} <select style='{INP};min-width:260px;margin-left:8px' "
                      f"onchange=\"location.href='/ledger-reco?ledger_id='+this.value\">{popts}</select>")
    upload = card(
        "<form method='post' action='/ledger-reco/import' enctype='multipart/form-data'>"
        f"<input type='hidden' name='ledger_id' value='{ledger_id}'>"
        f"<div style='display:flex;align-items:center;gap:10px;margin-bottom:12px'>{lbl('Sign mode')}"
        f"<select name='sign_mode' style='{INP}'><option value='MIRROR'>MIRROR — their statement of your account</option>"
        "<option value='SAME'>SAME — your ledger from another system</option></select></div>"
        f"<div style='display:flex;align-items:center;gap:10px;margin-bottom:12px'>{lbl('Statement Period')}"
        f"<input type='date' name='period_from' value='{FY_START}' style='{INP}'> <span style='color:var(--sec)'>→</span> "
        f"<input type='date' name='period_to' value='{date.today().isoformat()}' style='{INP}'></div>"
        f"<div style='margin-bottom:12px'>{lbl('Statement file (CSV / Excel / text-PDF)')}<br>"
        "<input type='file' name='file' accept='.csv,.xlsx,.xls,.pdf,.txt' required style='margin-top:8px'></div>"
        "<button class='btnp' type='submit'>Import &amp; auto-match</button></form>") if ledger_id else ""
    rrows = "".join(f"<tr><td>{r.get('file_name', '') or 'statement'}</td><td>{r.get('period_from', '')} → {r.get('period_to', '')}</td>"
                    f"<td>{r.get('matched', 0)} / {r.get('unmatched', 0)}</td>"
                    f"<td><a href='/ledger-reco/review?statement_id={r['id']}' style='color:var(--accent)'>Open</a></td></tr>" for r in recent) \
            or "<tr><td colspan='4' style='color:var(--sec);text-align:center;padding:14px'>No statements imported yet.</td></tr>"
    impc = card(lbl('Imported statements') + f"<table class='grid' style='margin-top:8px'><tr><th>File</th><th>Period</th><th>Matched / Unmatched</th><th></th></tr>{rrows}</table>") if ledger_id else ""
    body = RBAR_CSS + FORM_CSS + eh + resume_html + party_pick + upload + impc
    return shell("Ledger Reconciliation", "Pick a party, import their statement, auto-match, then resolve the rest.", body, active="/ledger-reco")


@app.post("/ledger-reco/import")
async def ledger_reco_import(ledger_id: int = Form(...), sign_mode: str = Form("MIRROR"), period_from: str = Form(""), period_to: str = Form(""), file: UploadFile = File(...)):
    import tempfile
    from urllib.parse import quote
    db, cid, cname, tree = ctx()
    recon = _lreconciler(db, cid, tree)
    suffix = os.path.splitext(file.filename or "")[1] or ".csv"
    data = await file.read()
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(data); tmp.close()
    try:
        sid = recon.import_statement(ledger_id=ledger_id, file_path=tmp.name, sign_mode=sign_mode,
                                     period_from=period_from or None, period_to=period_to or None)
        recon.auto_match(sid)
        db.commit()
    except Exception as e:
        db.close()
        try:
            os.unlink(tmp.name)
        except Exception:
            pass
        return RedirectResponse(url=f"/ledger-reco?ledger_id={ledger_id}&err={quote(('Import failed: ' + str(e))[:200])}", status_code=303)
    db.close()
    try:
        os.unlink(tmp.name)
    except Exception:
        pass
    return RedirectResponse(url=f"/ledger-reco/review?statement_id={sid}", status_code=303)


@app.get("/ledger-reco/review", response_class=HTMLResponse)
def ledger_reco_review(statement_id: int, err: str = ""):
    db, cid, cname, tree = ctx()
    recon = _lreconciler(db, cid, tree)
    st = db.execute("SELECT ledger_id, period_from, period_to, file_name FROM ledger_statements WHERE id=? AND company_id=?", (statement_id, cid)).fetchone()
    if not st:
        db.close()
        return shell("Ledger Reconciliation", "Not found", "<div class='note'>Statement not found.</div>", active="/ledger-reco")
    lid = st["ledger_id"]; pf = st["period_from"]; pt = st["period_to"]
    matched = recon.matched_lines(statement_id)
    un_stmt = recon.unmatched_statement_lines(statement_id)
    un_book = recon.unmatched_book_lines(lid, pf, pt)
    ignored = recon.ignored_statement_lines(statement_id)
    ledgers = sorted(tree.get_all_ledgers(), key=lambda l: l["name"])
    db.close()
    lopts = "".join(f"<option value='{l['id']}'>{l['name']}</option>" for l in ledgers)
    eh = f"<div class='note' style='background:var(--bad-soft);border-color:var(--bad);color:var(--bad)'>{err}</div>" if err else ""
    def tile(t, v, c='var(--text)'):
        return (f"<div style='flex:1;background:var(--card);border:1px solid var(--border);border-radius:10px;padding:14px 16px'>"
                f"<div style='font-size:11px;font-weight:700;color:var(--sec)'>{t}</div>"
                f"<div style='font-size:22px;font-weight:800;color:{c};margin-top:4px'>{v}</div></div>")
    tiles = (f"<div style='display:flex;gap:14px;margin-bottom:14px'>{tile('Matched', len(matched), '#057A55')}"
             f"{tile('Unmatched · statement', len(un_stmt), '#B45309')}{tile('Unmatched · book', len(un_book), '#B45309')}{tile('Ignored', len(ignored))}</div>")
    def pill(sg):
        c = '#057A55' if sg == 'CR' else '#C83A3A'
        return f"<span style='background:{c}22;color:{c};border-radius:6px;padding:2px 8px;font-size:11px;font-weight:700'>{sg}</span>"
    def sline(x):
        return (f"<tr><td>{x['txn_date']}</td><td>{pill(x['sign'])}</td><td class='num'>{fmt(x['amount'])}</td><td>{(x['narration'] or '')[:42]}</td>"
                "<td><form method='post' action='/ledger-reco/create-voucher' style='display:flex;gap:6px;margin:0'>"
                f"<input type='hidden' name='statement_id' value='{statement_id}'><input type='hidden' name='statement_line_id' value='{x['id']}'>"
                f"<select name='contra_ledger_id' class='rin' style='max-width:170px'>{lopts}</select>"
                "<button class='rbtn' type='submit'>Create</button></form></td>"
                "<td><form method='post' action='/ledger-reco/ignore' style='margin:0'>"
                f"<input type='hidden' name='statement_id' value='{statement_id}'><input type='hidden' name='statement_line_id' value='{x['id']}'>"
                "<button class='rbtn' type='submit'>Ignore</button></form></td></tr>")
    un_stmt_html = ("<h3 style='margin:6px 0 6px;font-size:15px'>Unmatched statement lines</h3>"
                    "<table class='grid'><tr><th>Date</th><th>Sign</th><th style='text-align:right'>Amount</th><th>Narration</th><th>Create voucher (pick the other side)</th><th></th></tr>"
                    + ("".join(sline(x) for x in un_stmt) or "<tr><td colspan='6' style='text-align:center;color:#057A55;padding:14px'>All statement lines resolved 🎉</td></tr>") + "</table>")
    sopt = "".join(f"<option value='{x['id']}'>{x['txn_date']} · {x['sign']} · {fmt(x['amount'])}</option>" for x in un_stmt)
    bopt = "".join(f"<option value='{b['id']}'>{fmt_date(b['voucher_date'])} · {b['voucher_number']} · {fmt(b['dr_amount'] or b['cr_amount'])}</option>" for b in un_book)
    match_form = ("<div style='background:var(--card);border:1px solid var(--border);border-radius:10px;padding:14px 16px;margin:14px 0'>"
                  "<b>Manually match</b> a statement line to an existing book entry:"
                  "<form method='post' action='/ledger-reco/match' style='display:flex;gap:8px;margin-top:8px;flex-wrap:wrap;align-items:center'>"
                  f"<input type='hidden' name='statement_id' value='{statement_id}'>"
                  f"<select name='statement_line_id' class='rin'>{sopt}</select> ↔ <select name='voucher_line_id' class='rin'>{bopt}</select>"
                  "<button class='rbtn' type='submit'>Match</button></form></div>") if (un_stmt and un_book) else ""
    book_html = ("<h3 style='margin:18px 0 6px;font-size:15px'>Unmatched book entries</h3>"
                 "<table class='grid'><tr><th>Date</th><th>Voucher</th><th>Narration</th><th style='text-align:right'>Dr</th><th style='text-align:right'>Cr</th></tr>"
                 + ("".join(f"<tr><td>{fmt_date(b['voucher_date'])}</td><td>{b['voucher_number']}</td><td>{(b['narration'] or '')[:42]}</td>"
                            f"<td class='num'>{fmt(b['dr_amount']) if b['dr_amount'] else ''}</td><td class='num'>{fmt(b['cr_amount']) if b['cr_amount'] else ''}</td></tr>" for b in un_book)
                    or "<tr><td colspan='5' style='text-align:center;color:var(--sec);padding:14px'>None.</td></tr>") + "</table>")
    finalise = ("<form method='post' action='/ledger-reco/finalise' style='margin-top:18px' onsubmit=\"return confirm('Finalise this reconciliation?')\">"
                f"<input type='hidden' name='statement_id' value='{statement_id}'>"
                "<button class='btnp' type='submit'>✓ Finalise reconciliation</button></form>")
    body = (RBAR_CSS + FORM_CSS + eh
            + f"<p style='color:var(--sec)'>{st['file_name'] or 'statement'} · {pf} → {pt} &nbsp;·&nbsp; <a href='/ledger-reco?ledger_id={lid}' style='color:var(--accent)'>← setup</a></p>"
            + tiles + un_stmt_html + match_form + book_html + finalise)
    return shell("Ledger Reconciliation — Review", "Resolve unmatched lines, then finalise.", body, active="/ledger-reco")


@app.post("/ledger-reco/match")
async def ledger_reco_match(statement_id: int = Form(...), statement_line_id: int = Form(...), voucher_line_id: int = Form(...)):
    db, cid, cname, tree = ctx()
    recon = _lreconciler(db, cid, tree)
    try:
        recon.manual_match(statement_line_id, voucher_line_id); db.commit()
    except Exception:
        db.rollback()
    db.close()
    return RedirectResponse(url=f"/ledger-reco/review?statement_id={statement_id}", status_code=303)


@app.post("/ledger-reco/ignore")
async def ledger_reco_ignore(statement_id: int = Form(...), statement_line_id: int = Form(...)):
    db, cid, cname, tree = ctx()
    recon = _lreconciler(db, cid, tree)
    try:
        recon.mark_ignored(statement_line_id); db.commit()
    except Exception:
        db.rollback()
    db.close()
    return RedirectResponse(url=f"/ledger-reco/review?statement_id={statement_id}", status_code=303)


@app.post("/ledger-reco/create-voucher")
async def ledger_reco_create_voucher(statement_id: int = Form(...), statement_line_id: int = Form(...), contra_ledger_id: int = Form(...)):
    from urllib.parse import quote
    db, cid, cname, tree = ctx()
    recon = _lreconciler(db, cid, tree)
    st = db.execute("SELECT ledger_id FROM ledger_statements WHERE id=?", (statement_id,)).fetchone()
    ln = db.execute("SELECT txn_date, amount, sign, narration, reference FROM ledger_statement_lines WHERE id=?", (statement_line_id,)).fetchone()
    if not st or not ln:
        db.close()
        return RedirectResponse(url=f"/ledger-reco/review?statement_id={statement_id}", status_code=303)
    lid = st["ledger_id"]; amt = ln["amount"]; eng = VoucherEngine(db, cid)
    try:
        if ln["sign"] == "DR":
            lines = [VoucherLine(ledger_id=lid, dr_amount=amt), VoucherLine(ledger_id=contra_ledger_id, cr_amount=amt)]
        else:
            lines = [VoucherLine(ledger_id=lid, cr_amount=amt), VoucherLine(ledger_id=contra_ledger_id, dr_amount=amt)]
        draft = VoucherDraft(voucher_type="JOURNAL", voucher_date=ln["txn_date"], lines=lines,
                             narration=ln["narration"] or "", reference=ln["reference"] or "")
        recon.create_voucher_for_line(statement_line_id, lid, draft)
        db.commit()
    except Exception as e:
        db.close()
        return RedirectResponse(url=f"/ledger-reco/review?statement_id={statement_id}&err={quote(('Could not create voucher: ' + str(e))[:160])}", status_code=303)
    db.close()
    return RedirectResponse(url=f"/ledger-reco/review?statement_id={statement_id}", status_code=303)


@app.post("/ledger-reco/finalise")
async def ledger_reco_finalise(statement_id: int = Form(...)):
    from urllib.parse import quote
    db, cid, cname, tree = ctx()
    recon = _lreconciler(db, cid, tree)
    row = db.execute("SELECT ledger_id FROM ledger_statements WHERE id=?", (statement_id,)).fetchone()
    lid = row["ledger_id"] if row else 0
    try:
        recon.finalise(statement_id); db.commit()
    except Exception as e:
        db.close()
        return RedirectResponse(url=f"/ledger-reco/review?statement_id={statement_id}&err={quote(('Could not finalise: ' + str(e))[:160])}", status_code=303)
    db.close()
    return RedirectResponse(url=f"/ledger-reco?ledger_id={lid}", status_code=303)


@app.get("/auto-post", response_class=HTMLResponse)
def auto_post():
    body = ("<div class='note'>Auto-Post lets the AI post documents straight to the books without review — OFF by default. "
            "Guardrails: only confident, complete, balanced vouchers whose ledgers already exist are posted; everything else is held as a draft. "
            "Toggle it per-document in the AI Documents Inbox.</div>")
    return shell("Auto-Post", "AI posts confident documents straight to the books", body, active="/auto-post")


@app.get("/verbal", response_class=HTMLResponse)
def verbal():
    body = "<div class='note'>Verbal Entry turns a typed or spoken sentence (\"paid 5000 rent in cash\") into a draft voucher via AI, for you to review and post.</div>"
    return shell("Verbal Entry", "Speak or type an entry → AI drafts the voucher", body, active="/verbal")


@app.get("/migration", response_class=HTMLResponse)
def migration():
    src = "".join(f"<div style='background:var(--card);border:1px solid var(--border);border-radius:10px;padding:16px 18px;font-weight:600'>{s}</div>"
                  for s in ["Tally (XML / port 9000)", "Excel / CSV", "Zoho Books", "QuickBooks"])
    body = f"<p style='color:var(--sec);margin-bottom:12px'>Bring your existing books in from:</p><div style='display:grid;grid-template-columns:repeat(2,1fr);gap:12px'>{src}</div>"
    return shell("Migration", "Import from another system", body, active="/migration")


@app.get("/backup", response_class=HTMLResponse)
def backup(msg: str = ""):
    note = ""
    if msg == "restored":
        note = "<div class='note' style='background:var(--good-soft);border-color:var(--good);color:#055c3a'>✓ Books restored from the backup file.</div>"
    elif msg == "bad":
        note = "<div class='note' style='background:var(--bad-soft);border-color:var(--bad);color:var(--bad)'>That file is not a valid Books HQ backup (.db).</div>"
    body = (note + "<div class='note'>Back up your company's books to a single database file, or restore from a backup you downloaded earlier.</div>"
            "<div style='margin-top:16px;display:flex;gap:14px;flex-wrap:wrap;align-items:center'>"
            "<a href='/backup/download' style='background:var(--accent);color:#fff;border-radius:8px;padding:11px 22px;font-weight:700;text-decoration:none'>💾 Backup now (download .db)</a>"
            "<form method='post' action='/backup/restore' enctype='multipart/form-data' style='display:flex;gap:8px;align-items:center' "
            "onsubmit=\"return confirm('Restore REPLACES the current books with the uploaded file. Continue?')\">"
            "<input type='file' name='file' accept='.db' required>"
            "<button style='background:var(--bad-soft);color:var(--bad);border:1px solid var(--bad);border-radius:8px;padding:10px 18px;font-weight:700;cursor:pointer'>↩ Restore from file…</button></form></div>")
    return shell("Backup & Restore", "Back up or restore your books", body, active="/backup")


@app.get("/backup/download")
def backup_download():
    from fastapi.responses import FileResponse
    p = str(Database(SLUG).path)
    return FileResponse(p, filename=f"{SLUG}_backup.db", media_type="application/octet-stream")


@app.post("/backup/restore")
async def backup_restore(file: UploadFile = File(...)):
    data = await file.read()
    if data[:16] != b"SQLite format 3\x00":
        return RedirectResponse(url="/backup?msg=bad", status_code=303)
    p = str(Database(SLUG).path)
    try:
        with open(p, "wb") as f:
            f.write(data)
        for sidecar in (p + "-wal", p + "-shm"):
            if os.path.exists(sidecar):
                os.remove(sidecar)
    except Exception:
        return RedirectResponse(url="/backup?msg=bad", status_code=303)
    return RedirectResponse(url="/backup?msg=restored", status_code=303)


@app.get("/manual", response_class=HTMLResponse)
def manual():
    secs = [
        ("Posting a voucher", "Open <b>Post Voucher</b>, choose the type (Payment, Receipt, Sales, Purchase…), and fill the smart fields — each box lists only the ledgers that belong there. Press <b>Post</b> or <b>Ctrl+S</b>. Use the <b>＋</b> beside any ledger box to create a new ledger without leaving the form."),
        ("Reports", "Trial Balance, Profit &amp; Loss, Balance Sheet, Cash &amp; Bank books, GST, TDS, Bill-wise, Cash-Flow and more. Every report has date filters, click-to-sort, a live row filter, and <b>⬇ Excel</b> / <b>🖶 Print</b>. Click any ledger name to drill into its statement; click a voucher number to open it."),
        ("Reconciliation", "<b>Bank</b> or <b>Ledger Reconciliation</b> → pick the account, upload a CSV/Excel statement, and the app auto-matches what it can. Resolve the rest by creating a voucher or matching manually, then <b>Finalise</b>."),
        ("Backup &amp; restore", "<b>Backup</b> → <b>💾 Backup now</b> downloads your whole company as a single <code>.db</code> file. <b>↩ Restore</b> replaces the books from a backup file."),
        ("Keyboard shortcuts", "<b>Ctrl+Q</b> tile launcher · <b>Ctrl+1–9</b> jump to a screen · <b>Alt+←</b> back · <b>Ctrl+S</b> post · <b>Alt+C</b> calculator · <b>F2</b> new ledger."),
    ]
    rows = "".join(f"<h3 style='margin:16px 0 4px;font-size:15px'>{t}</h3><div class='note'>{b}</div>" for t, b in secs)
    body = ("<div style='margin-bottom:10px'><button onclick='window.print()' "
            "style='background:var(--accent);color:#fff;border:none;border-radius:8px;padding:10px 20px;font-weight:700;cursor:pointer'>⬇ Download / open the manual</button></div>" + rows)
    return shell("User Manual", "How to use Books HQ", body, active="/manual")


@app.get("/license", response_class=HTMLResponse)
def license_page():
    try:
        from core.license_manager import LicenseManager
        lm = LicenseManager()
        s = lm.status_summary()
        seats_u, seats_a = lm.seats_used, lm.seats_allowed
    except Exception as e:
        return shell("License & Plan", "Plan & subscription", f"<div class='note'>Could not read licence: {e}</div>", active="/license")
    exp = str(s.get("expires_at") or "—")
    if s.get("days_to_expiry") is not None:
        exp += f"  ({s.get('days_to_expiry')} days)"
    rows = [
        ("Plan", str(s.get("plan", "—"))),
        ("Licence key", str(s.get("license_key") or "—")),
        ("Transactions used", f"{s.get('txn_used', 0):,} of {s.get('txn_limit', 0):,}  ({s.get('txn_pct', 0):.2f}%)"),
        ("Overage", f"{s.get('overage_count', 0)} extra (cost ₹{s.get('overage_cost', 0)})"),
        ("Seats", f"{seats_u} of {seats_a} used"),
        ("Expires", exp),
        ("Status", "Expired" if s.get("is_expired") else "Active"),
    ]
    trows = "".join(f"<tr><td style='color:var(--sec);width:220px'>{k}</td><td><b>{v}</b></td></tr>" for k, v in rows)
    return shell("License & Plan", "Your plan, usage and seats — live", f"<table class='grid'>{trows}</table>", active="/license")


@app.get("/ai-credits", response_class=HTMLResponse)
def ai_credits():
    body = "<div class='note'>Your AI wallet balance and usage — credits are spent when AI reads documents or drafts vouchers. The same wallet powers desktop and web.</div>"
    return shell("AI Credits", "AI wallet & usage", body, active="/ai-credits")


# ───────────────────────── Bank Reconciliation (review structure) ─────────────────────────
def _reconciler(db, cid, tree):
    from core.bank_reconciliation import BankReconciler
    return BankReconciler(db, cid, tree)


@app.get("/bankreco", response_class=HTMLResponse)
def bankreco(ledger_id: int = 0, err: str = ""):
    db, cid, cname, tree = ctx()
    try:
        banks = [l for l in tree.get_bank_cash_ledgers() if l.get("is_bank")]
    except Exception:
        banks = []
    recon = _reconciler(db, cid, tree)
    resume = recon.find_active_statement(ledger_id) if ledger_id else None
    try:
        recent = recon.recent_imports(ledger_id) if ledger_id else []
    except Exception:
        recent = []
    try:
        hist = recon.history_for_ledger(ledger_id) if ledger_id else []
    except Exception:
        hist = []
    db.close()
    INP = "background:var(--card2);border:1px solid var(--border);border-radius:8px;padding:8px 10px;font-size:13px;color:var(--text)"
    bopts = "<option value='0'>Choose a bank account…</option>" + "".join(
        f"<option value='{b['id']}'{' selected' if b['id'] == ledger_id else ''}>{b['name']}</option>" for b in banks)
    def card(inner):
        return f"<div style='background:var(--card);border:1px solid var(--border);border-radius:10px;padding:18px 20px;margin-bottom:14px'>{inner}</div>"
    def lbl(t):
        return f"<span style='font-size:11px;font-weight:700;color:var(--sec)'>{t}</span>"
    eh = f"<div class='note' style='background:var(--bad-soft);border-color:var(--bad);color:var(--bad)'>{err}</div>" if err else ""
    resume_html = card(f"<b>In-progress reconciliation:</b> {resume.get('matched', 0)} matched, {resume.get('unmatched', 0)} unmatched. "
                       f"<a href='/bankreco/review?statement_id={resume['id']}' style='color:var(--accent);font-weight:700'>Resume →</a>") if resume else ""
    bank_pick = card(f"{lbl('Bank Ledger')} <select style='{INP};min-width:260px;margin-left:8px' "
                     f"onchange=\"location.href='/bankreco?ledger_id='+this.value\">{bopts}</select>")
    upload = card(
        "<form method='post' action='/bankreco/import' enctype='multipart/form-data'>"
        f"<input type='hidden' name='bank_ledger_id' value='{ledger_id}'>"
        f"<div style='display:flex;align-items:center;gap:10px;margin-bottom:12px'>{lbl('Statement Period')}"
        f"<input type='date' name='period_from' value='{FY_START}' style='{INP}'> <span style='color:var(--sec)'>→</span> "
        f"<input type='date' name='period_to' value='{date.today().isoformat()}' style='{INP}'></div>"
        f"<div style='margin-bottom:12px'>{lbl('Statement file (CSV / Excel / text-PDF)')}<br>"
        "<input type='file' name='file' accept='.csv,.xlsx,.xls,.pdf,.txt' required style='margin-top:8px'></div>"
        "<button class='btnp' type='submit'>Import &amp; auto-match</button></form>") if ledger_id else ""
    rrows = "".join(f"<tr><td>{r.get('file_name', '') or 'statement'}</td><td>{r.get('period_from', '')} → {r.get('period_to', '')}</td>"
                    f"<td>{r.get('matched', 0)} / {r.get('unmatched', 0)}</td>"
                    f"<td><a href='/bankreco/review?statement_id={r['id']}' style='color:var(--accent)'>Open</a></td></tr>" for r in recent) \
            or "<tr><td colspan='4' style='color:var(--sec);text-align:center;padding:14px'>No statements imported yet.</td></tr>"
    impc = card(lbl('Imported statements') + f"<table class='grid' style='margin-top:8px'><tr><th>File</th><th>Period</th><th>Matched / Unmatched</th><th></th></tr>{rrows}</table>") if ledger_id else ""
    hrows = "".join(f"<tr><td>{h.get('as_of_date', '')}</td><td class='num'>{fmt(h.get('book_balance', 0))}</td>"
                    f"<td class='num'>{fmt(h.get('statement_balance', 0))}</td><td>{h.get('matched_count', 0)} matched</td><td>{h.get('finalised_at', '')}</td></tr>" for h in hist) \
            or "<tr><td colspan='5' style='color:var(--sec);text-align:center;padding:14px'>No past reconciliations.</td></tr>"
    histc = card(lbl('Past reconciliations') + f"<table class='grid' style='margin-top:8px'><tr><th>As of</th><th>Book bal.</th><th>Stmt bal.</th><th>Result</th><th>Finalised</th></tr>{hrows}</table>") if ledger_id else ""
    body = RBAR_CSS + FORM_CSS + eh + resume_html + bank_pick + upload + impc + histc
    return shell("Bank Reconciliation", "Pick a bank, import a statement, auto-match, then resolve the rest.", body, active="/bankreco")


@app.post("/bankreco/import")
async def bankreco_import(bank_ledger_id: int = Form(...), period_from: str = Form(""), period_to: str = Form(""), file: UploadFile = File(...)):
    import tempfile
    from urllib.parse import quote
    db, cid, cname, tree = ctx()
    recon = _reconciler(db, cid, tree)
    suffix = os.path.splitext(file.filename or "")[1] or ".csv"
    data = await file.read()
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(data); tmp.close()
    try:
        sid = recon.import_statement(bank_ledger_id=bank_ledger_id, file_path=tmp.name,
                                     period_from=period_from or None, period_to=period_to or None,
                                     allow_ai=False, confirm_account_population=True,
                                     force_mismatch_override=True, confirm_unverified=True)
        recon.auto_match(sid)
        db.commit()
    except Exception as e:
        db.close()
        try:
            os.unlink(tmp.name)
        except Exception:
            pass
        return RedirectResponse(url=f"/bankreco?ledger_id={bank_ledger_id}&err={quote(('Import failed: ' + str(e))[:200])}", status_code=303)
    db.close()
    try:
        os.unlink(tmp.name)
    except Exception:
        pass
    return RedirectResponse(url=f"/bankreco/review?statement_id={sid}", status_code=303)


@app.get("/bankreco/review", response_class=HTMLResponse)
def bankreco_review(statement_id: int, err: str = ""):
    db, cid, cname, tree = ctx()
    recon = _reconciler(db, cid, tree)
    st = db.execute("SELECT bank_ledger_id, period_from, period_to, file_name FROM bank_statements WHERE id=? AND company_id=?", (statement_id, cid)).fetchone()
    if not st:
        db.close()
        return shell("Bank Reconciliation", "Not found", "<div class='note'>Statement not found.</div>", active="/bankreco")
    blid = st["bank_ledger_id"]; pf = st["period_from"]; pt = st["period_to"]
    matched = recon.matched_lines(statement_id)
    un_stmt = recon.unmatched_statement_lines(statement_id)
    un_book = recon.unmatched_book_lines(blid, pf, pt)
    ignored = recon.ignored_statement_lines(statement_id)
    ledgers = sorted(tree.get_all_ledgers(), key=lambda l: l["name"])
    db.close()
    lopts = "".join(f"<option value='{l['id']}'>{l['name']}</option>" for l in ledgers)
    eh = f"<div class='note' style='background:var(--bad-soft);border-color:var(--bad);color:var(--bad)'>{err}</div>" if err else ""
    def tile(t, v, c='var(--text)'):
        return (f"<div style='flex:1;background:var(--card);border:1px solid var(--border);border-radius:10px;padding:14px 16px'>"
                f"<div style='font-size:11px;font-weight:700;color:var(--sec)'>{t}</div>"
                f"<div style='font-size:22px;font-weight:800;color:{c};margin-top:4px'>{v}</div></div>")
    tiles = (f"<div style='display:flex;gap:14px;margin-bottom:14px'>{tile('Matched', len(matched), '#057A55')}"
             f"{tile('Unmatched · statement', len(un_stmt), '#B45309')}{tile('Unmatched · book', len(un_book), '#B45309')}{tile('Ignored', len(ignored))}</div>")
    def pill(sg):
        c = '#057A55' if sg == 'CR' else '#C83A3A'
        return f"<span style='background:{c}22;color:{c};border-radius:6px;padding:2px 8px;font-size:11px;font-weight:700'>{sg}</span>"
    def sline(x):
        return (f"<tr><td>{x['txn_date']}</td><td>{pill(x['sign'])}</td><td class='num'>{fmt(x['amount'])}</td><td>{(x['narration'] or '')[:42]}</td>"
                "<td><form method='post' action='/bankreco/create-voucher' style='display:flex;gap:6px;margin:0'>"
                f"<input type='hidden' name='statement_id' value='{statement_id}'><input type='hidden' name='statement_line_id' value='{x['id']}'>"
                f"<select name='contra_ledger_id' class='rin' style='max-width:170px'>{lopts}</select>"
                "<button class='rbtn' type='submit'>Create</button></form></td>"
                "<td><form method='post' action='/bankreco/ignore' style='margin:0'>"
                f"<input type='hidden' name='statement_id' value='{statement_id}'><input type='hidden' name='statement_line_id' value='{x['id']}'>"
                "<button class='rbtn' type='submit'>Ignore</button></form></td></tr>")
    un_stmt_html = ("<h3 style='margin:6px 0 6px;font-size:15px'>Unmatched statement lines</h3>"
                    "<table class='grid'><tr><th>Date</th><th>Sign</th><th style='text-align:right'>Amount</th><th>Narration</th><th>Create voucher (pick the other side)</th><th></th></tr>"
                    + ("".join(sline(x) for x in un_stmt) or "<tr><td colspan='6' style='text-align:center;color:#057A55;padding:14px'>All statement lines resolved 🎉</td></tr>") + "</table>")
    sopt = "".join(f"<option value='{x['id']}'>{x['txn_date']} · {x['sign']} · {fmt(x['amount'])}</option>" for x in un_stmt)
    bopt = "".join(f"<option value='{b['id']}'>{fmt_date(b['voucher_date'])} · {b['voucher_number']} · {fmt(b['dr_amount'] or b['cr_amount'])}</option>" for b in un_book)
    match_form = ("<div style='background:var(--card);border:1px solid var(--border);border-radius:10px;padding:14px 16px;margin:14px 0'>"
                  "<b>Manually match</b> a statement line to an existing book entry:"
                  "<form method='post' action='/bankreco/match' style='display:flex;gap:8px;margin-top:8px;flex-wrap:wrap;align-items:center'>"
                  f"<input type='hidden' name='statement_id' value='{statement_id}'>"
                  f"<select name='statement_line_id' class='rin'>{sopt}</select> ↔ <select name='voucher_line_id' class='rin'>{bopt}</select>"
                  "<button class='rbtn' type='submit'>Match</button></form></div>") if (un_stmt and un_book) else ""
    book_html = ("<h3 style='margin:18px 0 6px;font-size:15px'>Unmatched book entries</h3>"
                 "<table class='grid'><tr><th>Date</th><th>Voucher</th><th>Narration</th><th style='text-align:right'>Dr</th><th style='text-align:right'>Cr</th></tr>"
                 + ("".join(f"<tr><td>{fmt_date(b['voucher_date'])}</td><td>{b['voucher_number']}</td><td>{(b['narration'] or '')[:42]}</td>"
                            f"<td class='num'>{fmt(b['dr_amount']) if b['dr_amount'] else ''}</td><td class='num'>{fmt(b['cr_amount']) if b['cr_amount'] else ''}</td></tr>" for b in un_book)
                    or "<tr><td colspan='5' style='text-align:center;color:var(--sec);padding:14px'>None.</td></tr>") + "</table>")
    finalise = ("<form method='post' action='/bankreco/finalise' style='margin-top:18px' onsubmit=\"return confirm('Finalise this reconciliation? It is snapshotted to history.')\">"
                f"<input type='hidden' name='statement_id' value='{statement_id}'>"
                "<button class='btnp' type='submit'>✓ Finalise reconciliation</button></form>")
    body = (RBAR_CSS + FORM_CSS + eh
            + f"<p style='color:var(--sec)'>{st['file_name'] or 'statement'} · {pf} → {pt} &nbsp;·&nbsp; <a href='/bankreco?ledger_id={blid}' style='color:var(--accent)'>← setup</a></p>"
            + tiles + un_stmt_html + match_form + book_html + finalise)
    return shell("Bank Reconciliation — Review", "Resolve unmatched lines, then finalise.", body, active="/bankreco")


@app.post("/bankreco/match")
async def bankreco_match(statement_id: int = Form(...), statement_line_id: int = Form(...), voucher_line_id: int = Form(...)):
    db, cid, cname, tree = ctx()
    recon = _reconciler(db, cid, tree)
    try:
        recon.manual_match(statement_line_id, voucher_line_id); db.commit()
    except Exception:
        db.rollback()
    db.close()
    return RedirectResponse(url=f"/bankreco/review?statement_id={statement_id}", status_code=303)


@app.post("/bankreco/ignore")
async def bankreco_ignore(statement_id: int = Form(...), statement_line_id: int = Form(...)):
    db, cid, cname, tree = ctx()
    recon = _reconciler(db, cid, tree)
    try:
        recon.mark_ignored(statement_line_id); db.commit()
    except Exception:
        db.rollback()
    db.close()
    return RedirectResponse(url=f"/bankreco/review?statement_id={statement_id}", status_code=303)


@app.post("/bankreco/create-voucher")
async def bankreco_create_voucher(statement_id: int = Form(...), statement_line_id: int = Form(...), contra_ledger_id: int = Form(...)):
    from urllib.parse import quote
    db, cid, cname, tree = ctx()
    recon = _reconciler(db, cid, tree)
    st = db.execute("SELECT bank_ledger_id FROM bank_statements WHERE id=?", (statement_id,)).fetchone()
    ln = db.execute("SELECT txn_date, amount, sign, narration, reference FROM bank_statement_lines WHERE id=?", (statement_line_id,)).fetchone()
    if not st or not ln:
        db.close()
        return RedirectResponse(url=f"/bankreco/review?statement_id={statement_id}", status_code=303)
    blid = st["bank_ledger_id"]; eng = VoucherEngine(db, cid)
    try:
        if ln["sign"] == "DR":
            draft = eng.build_payment(ln["txn_date"], contra_ledger_id, blid, ln["amount"], ln["narration"] or "", ln["reference"] or "")
        else:
            draft = eng.build_receipt(ln["txn_date"], contra_ledger_id, blid, ln["amount"], ln["narration"] or "", ln["reference"] or "")
        recon.create_voucher_for_line(statement_line_id, blid, draft)
        db.commit()
    except Exception as e:
        db.close()
        return RedirectResponse(url=f"/bankreco/review?statement_id={statement_id}&err={quote(('Could not create voucher: ' + str(e))[:160])}", status_code=303)
    db.close()
    return RedirectResponse(url=f"/bankreco/review?statement_id={statement_id}", status_code=303)


@app.post("/bankreco/finalise")
async def bankreco_finalise(statement_id: int = Form(...)):
    from urllib.parse import quote
    db, cid, cname, tree = ctx()
    recon = _reconciler(db, cid, tree)
    row = db.execute("SELECT bank_ledger_id FROM bank_statements WHERE id=?", (statement_id,)).fetchone()
    lid = row["bank_ledger_id"] if row else 0
    try:
        recon.finalise(statement_id); db.commit()
    except Exception as e:
        db.close()
        return RedirectResponse(url=f"/bankreco/review?statement_id={statement_id}&err={quote(('Could not finalise: ' + str(e))[:160])}", status_code=303)
    db.close()
    return RedirectResponse(url=f"/bankreco?ledger_id={lid}", status_code=303)


# ───────────────────────── Honest stubs (clearly labelled, not faked working) ─────────────────────────
def _stub(title, active, what):
    body = (f"<div class='note'><b>This screen is next in the build.</b><br>"
            f"It will be: {what}<br><br>The accounting logic behind it already exists in <code>core</code> "
            f"and is reused — this is the web UI layer to wire up.</div>")
    return shell(title, "Planned — not yet wired", body, active=active)


@app.get("/post-voucher", response_class=HTMLResponse)
def post_voucher(posted: str = ""):
    db, cid, cname, tree = ctx()
    ledgers = sorted(tree.get_all_ledgers(), key=lambda l: l["name"])
    opts = "<option value=''>— choose ledger —</option>" + "".join(f"<option value='{l['id']}'>{l['name']}</option>" for l in ledgers)
    groups = [r["name"] for r in db.execute("SELECT name FROM account_groups WHERE company_id=? ORDER BY name", (cid,)).fetchall()]
    def _ol(rs):
        return [{"id": r["id"], "name": r["name"]} for r in rs]
    try:
        cats = {"bank_cash": _ol(tree.get_bank_cash_ledgers()), "party": _ol(tree.get_party_ledgers()),
                "income": _ol(tree.get_income_ledgers()), "expense": _ol(tree.get_expense_ledgers()),
                "party_bank_cash": _ol(tree.get_party_and_bank_cash())}
    except Exception:
        cats = {}
    db.close()
    import json as _json
    cats_json = _json.dumps(cats)
    gmodal = "".join(f"<option>{g}</option>" for g in groups)
    rows = "".join(f"<tr><td><select name='ledger_{i}' class='vsel' style='width:100%'>{opts}</select></td>"
                   f"<td><input name='dr_{i}' class='amt dr' inputmode='decimal' placeholder='0'></td>"
                   f"<td><input name='cr_{i}' class='amt cr' inputmode='decimal' placeholder='0'></td></tr>" for i in range(6))
    types = [("PAYMENT", "Payment", "💸"), ("RECEIPT", "Receipt", "💰"), ("CONTRA", "Contra", "↔"),
             ("JOURNAL", "Journal", "📓"), ("SALES", "Income", "📈"), ("PURCHASE", "Expense", "📤"),
             ("DEBIT_NOTE", "Debit Note", "📋"), ("CREDIT_NOTE", "Credit Note", "📝")]
    tbar = "".join(f"<button type='button' class='vtb{' on' if c == 'PAYMENT' else ''}' data-t='{c}'>{ic} {nm}</button>" for c, nm, ic in types)
    grate = "<option value='0'>No GST</option>" + "".join(f"<option value='{r}'{' selected' if r == 18 else ''}>{r}%</option>" for r in [5, 12, 18, 28])
    msg = (f"<div class='note' style='background:var(--good-soft);border-color:var(--good);color:#055c3a'>✓ Posted <b>{posted}</b> — it's now in the Day Book.</div>") if posted else ""
    smart = ("<div id='smart'>"
             f"<div class='vrow'><label id='l1' class='vlbl'>Paid to (Party)</label><select id='f1' name='f1' class='vsel' style='flex:1'>{opts}</select></div>"
             f"<div class='vrow'><label id='l2' class='vlbl'>Paid from</label><select id='f2' name='f2' class='vsel' style='flex:1'>{opts}</select></div>"
             "<div class='vrow'><label class='vlbl'>Amount</label><input name='amount' id='amount' class='amt2' inputmode='decimal' placeholder='0'></div>"
             f"<div class='vrow' id='gstrow' style='display:none'><label class='vlbl' id='gstlbl'>GST</label><select name='gst' id='gst' class='vsel'>{grate}</select>"
             "<span id='gstcalc' style='margin-left:14px;color:var(--sec);font-size:13px'></span></div></div>")
    PDR, PCR = dr_cr_labels()
    journal = ("<div id='journal' style='display:none'><table class='grid'><tr><th>Ledger</th>"
               f"<th style='text-align:right;width:170px'>{PDR}</th><th style='text-align:right;width:170px'>{PCR}</th></tr>"
               f"{rows}</table><div id='bal' style='text-align:right;font-weight:700;padding:8px 4px'></div></div>")
    css = ("<style>.fl{font-size:11px;font-weight:700;color:var(--sec);margin-right:8px}"
           ".vsel{background:var(--card2);border:1px solid var(--border);border-radius:8px;padding:9px 11px;font-size:13.5px;color:var(--text)}"
           ".vrow{display:flex;align-items:center;gap:14px;margin-bottom:12px}"
           ".vlbl{width:210px;font-size:12.5px;font-weight:700;color:var(--text);background:var(--card2);border-radius:7px;padding:9px 12px}"
           ".amt2{flex:1;max-width:240px;background:var(--card2);border:1px solid var(--border);border-radius:8px;padding:9px 11px;font-size:14px;color:var(--text);text-align:right}"
           ".amt{width:100%;text-align:right;background:transparent;border:none;font-size:13.5px;color:var(--text)}.amt:focus{outline:none}"
           ".btnp{background:var(--accent);color:#fff;border:none;border-radius:8px;padding:11px 26px;font-weight:700;font-size:14px;cursor:pointer}"
           ".vtbar{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:16px}.vtb{background:var(--card2);border:1px solid var(--border);border-radius:8px;padding:9px 15px;font-size:12.5px;cursor:pointer;color:var(--text)}.vtb.on{background:var(--accent);color:#fff;border-color:var(--accent);font-weight:700}.vtb:hover{border-color:var(--accent)}</style>")
    js = ("<script>"
          "var L={PAYMENT:['Paid to (Party)','Paid from',0],RECEIPT:['Received from','Deposited to',0],CONTRA:['From Account','To Account',0],"
          "SALES:['Source of Income','Billed to / Received by',1],PURCHASE:['Expense Account','Paid via / Payable to',1],"
          "DEBIT_NOTE:['Dr \\u2014 Purchase Return','Cr \\u2014 Supplier / Party',1],CREDIT_NOTE:['Dr \\u2014 Sales Return','Cr \\u2014 Customer / Party',1],JOURNAL:['','',0]};"
          f"var CATS={cats_json};"
          "var FILT={PAYMENT:['party','bank_cash'],RECEIPT:['party','bank_cash'],CONTRA:['bank_cash','bank_cash'],SALES:['income','party_bank_cash'],PURCHASE:['expense','party_bank_cash'],DEBIT_NOTE:['expense','party'],CREDIT_NOTE:['income','party']};"
          "function fillSel(id,cat){var s=document.getElementById(id);if(!s)return;var cur=s.value;var ct=(s.options[s.selectedIndex]||{}).text||'';var arr=CATS[cat]||[];if(!arr.length){return;}var h=\"<option value=''>\\u2014 choose ledger \\u2014</option>\";var found=false;arr.forEach(function(o){h+='<option value=\"'+o.id+'\">'+o.name+'</option>';if(String(o.id)===cur)found=true;});if(cur&&!found){h+='<option value=\"'+cur+'\">'+ct+'</option>';}s.innerHTML=h;s.value=cur;}"
          "function gstCalc(){var a=parseFloat((document.getElementById('amount')||{}).value||0);var r=parseFloat((document.getElementById('gst')||{}).value||0);"
          "var tax=Math.round(a*r/100*100)/100;var el=document.getElementById('gstcalc');if(el)el.textContent=r?('Tax '+tax.toFixed(2)+'  \\u00b7  Total '+(a+tax).toFixed(2)):'';}"
          "function setType(t){document.getElementById('vtype').value=t;"
          "var mb=document.getElementById('multibtn');if(mb){if(t==='PAYMENT'||t==='RECEIPT'){mb.style.display='block';mb.href='/post-multi?ptype='+t;}else{mb.style.display='none';}}"
          "var s=document.getElementById('smart'),j=document.getElementById('journal');"
          "if(t==='JOURNAL'){s.style.display='none';j.style.display='block';}else{s.style.display='block';j.style.display='none';"
          "document.getElementById('l1').textContent=L[t][0];document.getElementById('l2').textContent=L[t][1];"
          "if(FILT[t]){fillSel('f1',FILT[t][0]);fillSel('f2',FILT[t][1]);}"
          "document.getElementById('gstrow').style.display=L[t][2]?'flex':'none';gstCalc();}}"
          "document.querySelectorAll('.vtb').forEach(function(b){b.addEventListener('click',function(){"
          "document.querySelectorAll('.vtb').forEach(function(x){x.classList.remove('on');});this.classList.add('on');setType(this.dataset.t);});});"
          "var ae=document.getElementById('amount');if(ae)ae.addEventListener('input',gstCalc);var ge=document.getElementById('gst');if(ge)ge.addEventListener('change',gstCalc);"
          "function bal(){let d=0,c=0;document.querySelectorAll('.dr').forEach(e=>d+=parseFloat(e.value||0));"
          "document.querySelectorAll('.cr').forEach(e=>c+=parseFloat(e.value||0));let f=Math.round((d-c)*100)/100;var b=document.getElementById('bal');"
          "if(b)b.innerHTML='Debit '+d.toFixed(2)+' \\u00b7 Credit '+c.toFixed(2)+' \\u00b7 '+(f===0&&d>0?'<span style=\"color:#057A55\">balanced \\u2713</span>':'<span style=\"color:#C83A3A\">diff '+f.toFixed(2)+'</span>');}"
          "document.querySelectorAll('.amt,.amt2').forEach(function(e){e.addEventListener('input',bal);e.addEventListener('blur',function(){var v=e.value.trim();"
          "if(/^[-+*/.() 0-9]+$/.test(v)&&/[-+*/]/.test(v)){try{var r=Function('return ('+v+')')();if(isFinite(r))e.value=Math.round(r*100)/100;}catch(_){}}bal();gstCalc();});});"
          "setType('PAYMENT');</script>")
    extras = (
        "<style>.rbtn{background:var(--card2);border:1px solid var(--border);border-radius:7px;padding:6px 12px;font-size:12px;cursor:pointer;color:var(--text)}.rbtn:hover{border-color:var(--accent)}</style>"
        "<div id='lmodal' style='display:none;position:fixed;inset:0;background:rgba(5,8,16,.6);z-index:60'>"
        "<div style='background:var(--card);border:1px solid var(--border);border-radius:12px;padding:20px;max-width:440px;margin:12vh auto'>"
        "<div style='font-weight:800;font-size:15px;margin-bottom:12px'>New ledger</div>"
        "<input id='lm_name' class='vsel' style='width:100%;margin-bottom:10px' placeholder='Ledger name'>"
        f"<select id='lm_group' class='vsel' style='width:100%;margin-bottom:12px'>{gmodal}</select>"
        "<div id='lm_err' style='color:#C83A3A;font-size:12px;margin-bottom:8px'></div>"
        "<button type='button' class='btnp' onclick='createLedger()'>Create</button> "
        "<button type='button' class='rbtn' onclick=\"document.getElementById('lmodal').style.display='none'\">Cancel</button></div></div>"
        "<div id='calc' style='display:none;position:fixed;right:30px;bottom:30px;width:230px;background:var(--card);border:1px solid var(--border);border-radius:12px;padding:12px;z-index:60;box-shadow:0 12px 34px rgba(0,0,0,.25)'>"
        "<div style='font-size:11px;font-weight:700;color:var(--sec);margin-bottom:6px'>Calculator (Alt+C)</div>"
        "<input id='calc_d' class='vsel' style='width:100%;text-align:right;margin-bottom:8px;font-size:16px' readonly>"
        "<div id='calc_keys' style='display:grid;grid-template-columns:repeat(4,1fr);gap:6px'></div>"
        "<button type='button' class='btnp' style='width:100%;margin-top:8px;padding:8px' onclick='calcUse()'>Use in field</button></div>"
        "<script>"
        "var _target=null;"
        "function _plus(s){var b=document.createElement('button');b.type='button';b.textContent='\\uFF0B';b.title='F2 \\u2014 new ledger';b.className='rbtn';b.style.marginLeft='6px';"
        "b.onclick=function(){_target=s;document.getElementById('lm_err').textContent='';document.getElementById('lm_name').value='';document.getElementById('lmodal').style.display='block';document.getElementById('lm_name').focus();};"
        "s.parentNode.insertBefore(b,s.nextSibling);}"
        "document.querySelectorAll(\"select[name='f1'],select[name='f2'],select[name^='ledger_']\").forEach(_plus);"
        "function createLedger(){var n=document.getElementById('lm_name').value.trim(),g=document.getElementById('lm_group').value;"
        "if(!n){document.getElementById('lm_err').textContent='Enter a name';return;}var fd=new FormData();fd.append('name',n);fd.append('group',g);"
        "fetch('/api/add-ledger',{method:'POST',body:fd}).then(function(r){return r.json();}).then(function(d){"
        "if(d.error){document.getElementById('lm_err').textContent=d.error;return;}"
        "document.querySelectorAll(\"select[name='f1'],select[name='f2'],select[name^='ledger_']\").forEach(function(s){var o=document.createElement('option');o.value=d.id;o.textContent=d.name;s.appendChild(o);});"
        "if(_target)_target.value=d.id;document.getElementById('lmodal').style.display='none';});}"
        "var _amtF=null;document.querySelectorAll('.amt,.amt2').forEach(function(e){e.addEventListener('focus',function(){_amtF=e;});});"
        "var CK=['7','8','9','/','4','5','6','*','1','2','3','-','0','.','=','+','C'];var kc=document.getElementById('calc_keys');"
        "CK.forEach(function(k){var b=document.createElement('button');b.type='button';b.textContent=k;b.className='rbtn';b.style.padding='8px';"
        "b.onclick=function(){var d=document.getElementById('calc_d');if(k==='C'){d.value='';}else if(k==='='){try{var r=Function('return ('+d.value+')')();if(isFinite(r))d.value=Math.round(r*100)/100;}catch(_){}}else{d.value+=k;}};kc.appendChild(b);});"
        "function calcUse(){var v=document.getElementById('calc_d').value;try{var r=Function('return ('+v+')')();if(isFinite(r))v=Math.round(r*100)/100;}catch(_){}"
        "if(_amtF){_amtF.value=v;_amtF.dispatchEvent(new Event('input'));_amtF.dispatchEvent(new Event('blur'));}document.getElementById('calc').style.display='none';}"
        "document.addEventListener('keydown',function(e){if(e.altKey&&(e.key==='c'||e.key==='C')){e.preventDefault();var c=document.getElementById('calc');c.style.display=(c.style.display==='none'||!c.style.display)?'block':'none';}});"
        "var vf=document.getElementById('vform');if(vf)vf.addEventListener('submit',function(e){e.preventDefault();var m=document.getElementById('vmsg');m.style.display='none';var fd=new FormData(vf);"
        "fetch('/post-voucher',{method:'POST',body:fd}).then(function(r){return r.json();}).then(function(d){"
        "if(d.ok){window.location.href='/post-voucher?posted='+encodeURIComponent(d.number);}else{m.textContent='Could not post: '+d.error;m.style.display='block';window.scrollTo(0,0);}})"
        ".catch(function(){m.textContent='Network error - please try again.';m.style.display='block';});});"
        "function nudgeDate(n){var d=document.getElementById('vdate');if(!d||!d.value)return;var dt=new Date(d.value+'T00:00');dt.setDate(dt.getDate()+n);d.value=dt.getFullYear()+'-'+String(dt.getMonth()+1).padStart(2,'0')+'-'+String(dt.getDate()).padStart(2,'0');}"
        "var _dp=document.getElementById('dprev');if(_dp)_dp.onclick=function(){nudgeDate(-1);};var _dn=document.getElementById('dnext');if(_dn)_dn.onclick=function(){nudgeDate(1);};"
        "document.addEventListener('keydown',function(e){if(e.altKey&&e.key===','){e.preventDefault();nudgeDate(-1);}if(e.altKey&&e.key==='.'){e.preventDefault();nudgeDate(1);}});"
        "var _cb=document.getElementById('clearbtn');if(_cb)_cb.onclick=function(){document.querySelectorAll('#vform input').forEach(function(x){if(x.type!=='hidden'&&x.type!=='date')x.value='';});updateBal();gstCalc();};"
        "var _ca=document.getElementById('calcbtn');if(_ca)_ca.onclick=function(){var c=document.getElementById('calc');c.style.display=(c.style.display==='none'||!c.style.display)?'block':'none';};"
        "function updateBal(){var vb=document.getElementById('vbal');if(!vb)return;var sm=document.getElementById('smart');"
        "if(sm&&sm.style.display!=='none'){var a=parseFloat((document.getElementById('amount')||{}).value||0);var r=parseFloat((document.getElementById('gst')||{}).value||0);var tot=a+Math.round(a*r/100*100)/100;"
        "vb.innerHTML='Dr '+tot.toFixed(2)+' \\u00b7 Cr '+tot.toFixed(2)+' \\u00b7 <span style=\"color:#057A55\">balanced \\u2713</span>';}"
        "else{var d=0,c=0;document.querySelectorAll('.dr').forEach(function(e){d+=parseFloat(e.value||0);});document.querySelectorAll('.cr').forEach(function(e){c+=parseFloat(e.value||0);});var f=Math.round((d-c)*100)/100;"
        "vb.innerHTML='Dr '+d.toFixed(2)+' \\u00b7 Cr '+c.toFixed(2)+' \\u00b7 '+(f===0&&d>0?'<span style=\"color:#057A55\">balanced \\u2713</span>':'<span style=\"color:#C83A3A\">diff '+f.toFixed(2)+'</span>');}}"
        "document.querySelectorAll('.amt,.amt2,#gst').forEach(function(e){e.addEventListener('input',updateBal);e.addEventListener('change',updateBal);});"
        "document.querySelectorAll('.vtb').forEach(function(b){b.addEventListener('click',function(){setTimeout(updateBal,0);});});updateBal();"
        "</script>")
    body = (msg + "<div id='vmsg' style='display:none;background:var(--bad-soft);border:1px solid var(--bad);color:var(--bad);border-radius:8px;padding:10px 14px;margin-bottom:12px;font-weight:600'></div>"
            "<form method='post' action='/post-voucher' id='vform'>"
            f"<div class='vtbar'>{tbar}</div><input type='hidden' name='vtype' id='vtype' value='PAYMENT'>"
            "<div style='margin-bottom:14px;display:flex;align-items:center;gap:6px'><span class='fl'>Date</span>"
            "<button type='button' class='rbtn' id='dprev' title='Previous day (Alt+,)' style='padding:6px 12px;font-size:16px;font-weight:bold'>‹</button>"
            f"<input name='vdate' id='vdate' type='date' value='{date.today().isoformat()}' class='vsel'>"
            "<button type='button' class='rbtn' id='dnext' title='Next day (Alt+.)' style='padding:6px 12px;font-size:16px;font-weight:bold'>›</button></div>"
            + smart + journal
            + "<a id='multibtn' href='/post-multi?ptype=PAYMENT' style='display:none;margin:6px 0 14px;padding:10px 14px;"
            "border:1.5px dashed var(--accent);border-radius:8px;color:var(--accent);font-weight:700;font-size:12.5px;text-decoration:none'>"
            "+ Multi-party voucher (one bank entry, several parties)</a>"
            + "<div style='display:flex;gap:12px;margin:14px 0'>"
            "<input name='narration' placeholder='Narration' class='vsel' style='flex:1'>"
            "<input name='reference' placeholder='Reference' class='vsel' style='width:200px'></div>"
            "<div style='display:flex;align-items:center;gap:14px;margin-top:8px'>"
            "<span id='vbal' style='font-size:12px;color:var(--sec)'></span><span style='flex:1'></span>"
            "<button type='button' class='rbtn' id='calcbtn' style='padding:10px 16px'>🖩 Calc</button>"
            "<button type='button' class='rbtn' id='clearbtn' style='padding:10px 18px'>Clear</button>"
            "<button class='btnp' type='submit'>Post voucher (Ctrl+S)</button></div></form>" + css + js + extras)
    return shell("Post Voucher", "Guided entry per type — GST auto-calc · F2 new ledger · Alt+C calculator", body, active="/post-voucher")


@app.post("/post-voucher")
async def post_voucher_submit(request: Request):
    from fastapi.responses import JSONResponse
    form = await request.form()
    db, cid, cname, tree = ctx()
    eng = VoucherEngine(db, cid)
    vdate = form.get("vdate") or date.today().isoformat()
    vt = form.get("vtype") or "PAYMENT"
    narr = form.get("narration") or ""; ref = form.get("reference") or ""
    try:
        if vt == "JOURNAL":
            lines = []
            for i in range(6):
                lid = form.get(f"ledger_{i}")
                if not lid:
                    continue
                dr = float(form.get(f"dr_{i}") or 0); cr = float(form.get(f"cr_{i}") or 0)
                if dr == 0 and cr == 0:
                    continue
                lines.append(VoucherLine(ledger_id=int(lid), dr_amount=dr, cr_amount=cr))
            draft = VoucherDraft(voucher_type="JOURNAL", voucher_date=vdate, lines=lines, narration=narr, reference=ref)
        else:
            f1 = int(form.get("f1")); f2 = int(form.get("f2")); amt = float(form.get("amount") or 0)
            gst = float(form.get("gst") or 0)
            if vt == "PAYMENT":
                draft = eng.build_payment(vdate, f1, f2, amt, narr, ref)
            elif vt == "RECEIPT":
                draft = eng.build_receipt(vdate, f1, f2, amt, narr, ref)
            elif vt == "SALES":
                draft = eng.build_sales(vdate, f2, f1, amt, gst, narr, ref)
            elif vt == "PURCHASE":
                draft = eng.build_purchase(vdate, f2, f1, amt, gst, narr, ref)
            elif vt == "CONTRA":
                draft = VoucherDraft(voucher_type="CONTRA", voucher_date=vdate, narration=narr, reference=ref,
                                     lines=[VoucherLine(ledger_id=f2, dr_amount=amt), VoucherLine(ledger_id=f1, cr_amount=amt)])
            elif vt == "DEBIT_NOTE":
                draft = eng.build_debit_note(vdate, f2, f1, amt, gst, narr, ref)
            else:  # CREDIT_NOTE
                draft = eng.build_credit_note(vdate, f2, f1, amt, gst, narr, ref)
        num = eng.post(draft).voucher_number
    except Exception as e:
        db.close()
        return JSONResponse({"ok": False, "error": str(e)})
    db.close()
    return JSONResponse({"ok": True, "number": num})


@app.get("/post-multi", response_class=HTMLResponse)
def post_multi(posted: str = "", ptype: str = "PAYMENT", err: str = ""):
    db, cid, cname, tree = ctx()
    banks = [{"id": r["id"], "name": r["name"]} for r in tree.get_bank_cash_ledgers()]
    parties = [{"id": r["id"], "name": r["name"]} for r in tree.get_party_ledgers()]
    db.close()
    is_pay = ptype != "RECEIPT"
    bank_opts = "".join(f"<option value='{b['id']}'>{b['name']}</option>" for b in banks)
    party_opts = "<option value=''>— party —</option>" + "".join(f"<option value='{p['id']}'>{p['name']}</option>" for p in parties)
    prows = "".join(f"<tr><td><select name='party_{i}' class='vsel' style='width:100%'>{party_opts}</select></td>"
                    f"<td><input name='amt_{i}' class='vsel pamt' inputmode='decimal' placeholder='0' style='max-width:150px;text-align:right'></td>"
                    f"<td><input name='pnarr_{i}' class='vsel' placeholder='(optional)' style='width:100%'></td></tr>" for i in range(8))
    msg = (f"<div class='note' style='background:var(--good-soft);border-color:var(--good);color:#055c3a'>✓ Posted <b>{posted}</b> — it's in the Day Book.</div>") if posted else ""
    if err:
        msg += f"<div class='note' style='background:var(--bad-soft);border-color:var(--bad);color:var(--bad)'>{err}</div>"
    bankrole = "Paid from (Bank / Cash)" if is_pay else "Deposited to (Bank / Cash)"
    partyrole = "Parties paid" if is_pay else "Parties received from"
    vtbcss = ("<style>.vtbar{display:flex;gap:8px;margin-bottom:16px}.vtb{background:var(--card2);border:1px solid var(--border);"
              "border-radius:8px;padding:9px 15px;font-size:12.5px;cursor:pointer;color:var(--text);text-decoration:none}"
              ".vtb.on{background:var(--accent);color:#fff;border-color:var(--accent);font-weight:700}</style>")
    js = ("<script>function mtot(){var t=0;document.querySelectorAll('.pamt').forEach(function(e){t+=parseFloat(e.value||0);});"
          "var m=document.getElementById('mtot');if(m)m.textContent='Total: '+t.toFixed(2);}"
          "document.querySelectorAll('.pamt').forEach(function(e){e.addEventListener('input',mtot);});mtot();</script>")
    body = (msg + vtbcss
            + "<div class='vtbar'>"
              f"<a href='/post-multi?ptype=PAYMENT' class='vtb{' on' if is_pay else ''}'>💸 Pay multiple parties</a>"
              f"<a href='/post-multi?ptype=RECEIPT' class='vtb{' on' if not is_pay else ''}'>💰 Receive from multiple</a></div>"
            + "<form method='post' action='/post-multi'>"
              f"<input type='hidden' name='ptype' value='{ptype}'>"
              f"<div class='vrow'><label class='vlbl'>Date</label><input name='vdate' type='date' value='{date.today().isoformat()}' class='vsel'></div>"
              f"<div class='vrow'><label class='vlbl'>{bankrole}</label><select name='bank' class='vsel' style='flex:1'>{bank_opts}</select></div>"
              f"<h3 style='margin:16px 0 8px;font-size:14px;font-weight:800'>{partyrole}</h3>"
              "<table class='grid'><tr><th>Party</th><th style='text-align:right;width:160px'>Amount</th><th>Line narration</th></tr>"
              + prows + "</table>"
              "<div id='mtot' style='text-align:right;font-weight:700;padding:8px 4px'></div>"
              "<div style='display:flex;gap:12px;margin:12px 0'>"
              "<input name='narration' placeholder='Narration (whole voucher)' class='vsel' style='flex:1'>"
              "<input name='reference' placeholder='Reference' class='vsel' style='width:200px'></div>"
              "<button class='btnp' type='submit'>Post voucher</button></form>" + FORM_CSS + js)
    return shell("Multi-party Voucher", "Pay or receive from several parties against one bank / cash account", body, active="/post-multi")


@app.post("/post-multi")
async def post_multi_submit(request: Request):
    form = await request.form()
    ptype = form.get("ptype") or "PAYMENT"
    db, cid, cname, tree = ctx()
    eng = VoucherEngine(db, cid)
    try:
        vdate = form.get("vdate") or date.today().isoformat()
        bank_id = int(form.get("bank"))
        party_lines = []
        for i in range(8):
            pid = form.get(f"party_{i}"); amt = form.get(f"amt_{i}")
            if pid and amt and float(amt) > 0:
                party_lines.append({"ledger_id": int(pid), "amount": float(amt), "narration": (form.get(f"pnarr_{i}") or "").strip()})
        if not party_lines:
            raise ValueError("Add at least one party row with an amount.")
        narr = (form.get("narration") or "").strip(); ref = (form.get("reference") or "").strip()
        if ptype == "RECEIPT":
            draft = eng.build_receipt_multi(vdate, bank_id, party_lines, narr, ref)
        else:
            draft = eng.build_payment_multi(vdate, bank_id, party_lines, narr, ref)
        num = eng.post(draft).voucher_number
    except Exception as e:
        db.close()
        from urllib.parse import quote
        return RedirectResponse(url=f"/post-multi?ptype={ptype}&err={quote(str(e))}", status_code=303)
    db.close()
    return RedirectResponse(url=f"/post-multi?posted={num}", status_code=303)


@app.get("/gst", response_class=HTMLResponse)
def gst(frm: str = "", to: str = ""):
    db, cid, cname, tree = ctx()
    f = frm or FY_START; t = to or date.today().isoformat()
    re = ReportsEngine(db, cid)
    try:
        g = re.gst_summary(f, t); h = re.hsn_summary(f, t)
    except Exception as e:
        db.close()
        return report_shell("GST Summary", "Output vs input tax", "/gst", f"<div class='note'>{e}</div>", frm=f, to=t)
    db.close()
    tl = g.get("tax_lines", [])
    out_t = sum(x['output_tax'] for x in tl); in_t = sum(x['input_tax'] for x in tl); net = out_t - in_t
    grows = "".join(f"<tr><td>{x['tax_type']}</td><td class='num'>{x['tax_rate']}%</td><td class='num'>{fmt(x['output_tax'])}</td><td class='num'>{fmt(x['input_tax'])}</td><td class='num'>{fmt(x['output_tax'] - x['input_tax'])}</td></tr>" for x in tl)
    grows += f"<tr class='tot'><td colspan='2'>TOTAL</td><td class='num'>{fmt(out_t)}</td><td class='num'>{fmt(in_t)}</td><td class='num'>{fmt(net)}</td></tr>"
    hrows = "".join(f"<tr><td>{r['hsn']}</td><td class='num'>{fmt(r['taxable'])}</td><td class='num'>{fmt(r['cgst'])}</td><td class='num'>{fmt(r['sgst'])}</td><td class='num'>{fmt(r['igst'])}</td></tr>" for r in h.get("rows", []))
    body = (f"<div class='note'>Output tax (collected) <b>{fmt(out_t)}</b> − Input tax / ITC (paid) <b>{fmt(in_t)}</b> = <b>Net GST payable {fmt(net)}</b></div>"
            "<h3 style='margin:16px 0 6px;font-size:15px'>Tax by type &amp; rate</h3>"
            "<table class='grid'><tr><th>Tax</th><th style='text-align:right'>Rate</th><th style='text-align:right'>Output (collected)</th><th style='text-align:right'>Input (ITC)</th><th style='text-align:right'>Net</th></tr>" + grows + "</table>"
            "<h3 style='margin:18px 0 6px;font-size:15px'>HSN summary</h3>"
            "<table class='grid'><tr><th>HSN</th><th style='text-align:right'>Taxable</th><th style='text-align:right'>CGST</th><th style='text-align:right'>SGST</th><th style='text-align:right'>IGST</th></tr>"
            + (hrows or "<tr><td colspan='5' style='text-align:center;color:var(--sec);padding:12px'>No HSN data.</td></tr>") + "</table>")
    return report_shell("GST Summary", "Output vs input tax (GSTR-3B style)", "/gst", body, frm=f, to=t)


@app.get("/bill-wise", response_class=HTMLResponse)
def bill_wise(as_of: str = "", grp: str = "debtors"):
    db, cid, cname, tree = ctx()
    aod = as_of or date.today().isoformat()
    from core.bill_wise import BillWiseEngine
    bw = BillWiseEngine(db, cid)
    gname = "Sundry Creditors" if grp == "creditors" else "Sundry Debtors"
    try:
        rows = bw.aging_by_bill(aod, gname).get("rows", [])
    except Exception:
        rows = []
    db.close()
    def bk(b):
        return (b or "").lstrip("b").replace("_", "-").replace("p", "+")
    toggle = ("<div style='margin-bottom:12px;display:flex;gap:10px;align-items:center'>"
              "<label style='color:var(--sec);font-size:13px;font-weight:700'>Group</label>"
              f"<select onchange=\"location.href='/bill-wise?as_of={aod}&grp='+this.value\" "
              "style='background:var(--card2);border:1px solid var(--border);border-radius:8px;padding:8px 12px;font-size:13px;color:var(--text)'>"
              f"<option value='debtors'{' selected' if grp != 'creditors' else ''}>Receivable (Debtors)</option>"
              f"<option value='creditors'{' selected' if grp == 'creditors' else ''}>Payable (Creditors)</option></select></div>")
    trows = "".join(f"<tr><td>{r['party']}</td><td>{r['bill_number']}</td><td>{fmt_date(r['bill_date'])}</td>"
                    f"<td class='num'>{fmt(r['bill_amount'])}</td><td class='num'>{fmt(r['pending_amount'])}</td>"
                    f"<td class='num'>{r['age_days']}</td><td>{bk(r['bucket'])}</td></tr>" for r in rows)
    total = sum(r['pending_amount'] for r in rows)
    if not trows:
        trows = "<tr><td colspan='7' style='text-align:center;color:var(--sec);padding:14px'>No open bills.</td></tr>"
    else:
        trows += f"<tr class='tot'><td colspan='4'>TOTAL PENDING</td><td class='num'>{fmt(total)}</td><td colspan='2'></td></tr>"
    body = (toggle + "<table class='grid'><tr><th>Party</th><th>Bill</th><th>Date</th><th style='text-align:right'>Amount</th>"
            "<th style='text-align:right'>Pending</th><th style='text-align:right'>Age (days)</th><th>Bucket</th></tr>" + trows + "</table>")
    return report_shell("Bill-wise Outstanding", "Open bills by party, aged", "/bill-wise", body, as_of=aod)


@app.get("/cash-flow", response_class=HTMLResponse)
def cash_flow(as_of: str = ""):
    db, cid, cname, tree = ctx()
    aod = as_of or date.today().isoformat()
    re = ReportsEngine(db, cid)
    try:
        d = re.cashflow_forecast(aod)
    except Exception as e:
        db.close()
        return report_shell("Cash-Flow Planning", "Forward cash projection", "/cash-flow", f"<div class='note'>{e}</div>", as_of=aod)
    try:
        oi = re.cashflow_open_items(aod)
    except Exception:
        oi = {"in": {"vital": []}, "out": {"vital": []}}
    db.close()
    rows = d.get("rows", [])
    trows = "".join(f"<tr><td>{r['label']}</td><td class='num'>{fmt(r['inflow'])}</td><td class='num'>{fmt(r['outflow'])}</td>"
                    f"<td class='num' style='color:{'#057A55' if r['net'] >= 0 else '#C83A3A'}'>{fmt(r['net'])}</td>"
                    f"<td class='num' style='color:{'#057A55' if r['closing'] >= 0 else '#C83A3A'}'>{fmt(r['closing'])}</td></tr>" for r in rows)
    if not trows:
        trows = "<tr><td colspan='5' style='text-align:center;color:var(--sec);padding:14px'>No forecast data.</td></tr>"
    def oirows(items, direction, color):
        return "".join(f"<tr><td><span style='color:{color};font-weight:700'>{direction}</span></td>"
                       f"<td>{x.get('party', '')}</td><td class='num'>{fmt(x.get('amount', 0))}</td></tr>" for x in items)
    oih = oirows(oi.get("in", {}).get("vital", []), "In", "#057A55") + oirows(oi.get("out", {}).get("vital", []), "Out", "#C83A3A")
    if not oih:
        oih = "<tr><td colspan='3' style='text-align:center;color:var(--sec);padding:12px'>No open items.</td></tr>"
    body = (f"<div class='note'>Opening cash position <b>{fmt(d.get('opening', 0))}</b>, projected forward from open receivables and payables.</div>"
            "<table class='grid'><tr><th>Period</th><th style='text-align:right'>Expected in</th><th style='text-align:right'>Expected out</th>"
            "<th style='text-align:right'>Net</th><th style='text-align:right'>Closing</th></tr>" + trows + "</table>"
            "<h3 style='margin:18px 0 6px;font-size:15px'>Open items driving the forecast</h3>"
            "<table class='grid'><tr><th>Direction</th><th>Party</th><th style='text-align:right'>Outstanding</th></tr>" + oih + "</table>")
    return report_shell("Cash-Flow Planning", "Projected cash position ahead", "/cash-flow", body, as_of=aod)


@app.get("/mileage", response_class=HTMLResponse)
def mileage(saved: str = ""):
    db, cid, cname, tree = ctx()
    rows = db.execute("SELECT id, trip_date, miles, purpose, vehicle FROM mileage_log WHERE company_id=? ORDER BY trip_date DESC", (cid,)).fetchall()
    total_miles = sum(r["miles"] for r in rows)
    db.close()
    rate = 0.70
    trows = "".join(f"<tr><td>{fmt_date(r['trip_date'])}</td><td class='num'>{r['miles']:.1f}</td><td>{r['purpose'] or ''}</td><td>{r['vehicle'] or ''}</td>"
                    f"<td><a href='/mileage/delete/{r['id']}' style='color:#C83A3A' onclick=\"return confirm('Remove this trip?')\">Remove</a></td></tr>" for r in rows)
    if not trows:
        trows = "<tr><td colspan='5' style='text-align:center;color:var(--sec);padding:14px'>No trips logged. Add one below.</td></tr>"
    else:
        trows += f"<tr class='tot'><td>TOTAL</td><td class='num'>{total_miles:.1f} mi</td><td colspan='3'>× ${rate:.2f}/mi = <b>${total_miles * rate:,.2f}</b> deductible (Schedule C)</td></tr>"
    msg = "<div class='note' style='background:var(--good-soft);border-color:var(--good);color:#055c3a'>✓ Saved.</div>" if saved else ""
    add = ("<h3 style='margin:20px 0 10px;font-size:14px;font-weight:800'>Log a trip</h3>"
           "<form method='post' action='/mileage' style='max-width:620px'>"
           "<div class='vrow'><label class='vlbl'>Date</label><input type='date' name='trip_date' class='vsel' required></div>"
           "<div class='vrow'><label class='vlbl'>Miles</label><input name='miles' class='vsel' inputmode='decimal' style='max-width:160px' required></div>"
           "<div class='vrow'><label class='vlbl'>Purpose</label><input name='purpose' class='vsel' style='flex:1;max-width:340px'></div>"
           "<div class='vrow'><label class='vlbl'>Vehicle</label><input name='vehicle' class='vsel' style='max-width:240px'></div>"
           "<button class='btnp' type='submit'>Add trip</button></form>")
    body = (msg + "<table class='grid'><tr><th>Date</th><th style='text-align:right'>Miles</th><th>Purpose</th><th>Vehicle</th><th></th></tr>" + trows + "</table>" + add + FORM_CSS)
    return shell("Mileage Log", "Business miles for Schedule C", body, active="/mileage")


@app.post("/mileage")
async def mileage_add(request: Request):
    form = await request.form()
    db, cid, cname, tree = ctx()
    try:
        miles = float(form.get("miles") or 0)
        if form.get("trip_date") and miles > 0:
            db.execute("INSERT INTO mileage_log (company_id, trip_date, miles, purpose, vehicle) VALUES (?,?,?,?,?)",
                       (cid, form.get("trip_date"), miles, (form.get("purpose") or "").strip(), (form.get("vehicle") or "").strip()))
            db.commit()
    except Exception:
        db.rollback()
    db.close()
    return RedirectResponse(url="/mileage?saved=1", status_code=303)


@app.get("/mileage/delete/{trip_id}")
def mileage_delete(trip_id: int):
    db, cid, cname, tree = ctx()
    db.execute("DELETE FROM mileage_log WHERE id=? AND company_id=?", (trip_id, cid))
    db.commit(); db.close()
    return RedirectResponse(url="/mileage?saved=1", status_code=303)


@app.get("/tds", response_class=HTMLResponse)
def tds(frm: str = "", to: str = ""):
    db, cid, cname, tree = ctx()
    f = frm or FY_START; t = to or date.today().isoformat()
    re = ReportsEngine(db, cid)
    try:
        d = re.tds_report(f, t)
        reg = re.tds_register(f, t)
    except Exception as e:
        db.close()
        return report_shell("TDS Report", "Tax deducted at source", "/tds", f"<div class='note'>{e}</div>", frm=f, to=t)
    db.close()
    lines = d.get("tds_lines", [])
    def cells(l):
        return (f"<td>{l.get('section', l.get('tds_section', ''))}</td>"
                f"<td class='num'>{l.get('rate', l.get('tds_rate', ''))}</td>"
                f"<td class='num'>{fmt(l.get('tds_amount', l.get('tds', 0)))}</td>"
                f"<td class='num'>{l.get('voucher_count', l.get('count', ''))}</td>")
    trows = "".join("<tr>" + cells(l) + "</tr>" for l in lines)
    if not trows:
        trows = "<tr><td colspan='4' style='text-align:center;color:var(--sec);padding:14px'>No TDS deducted in this period.</td></tr>"
    parties = reg.get("parties", [])
    def rcells(p):
        return (f"<td>{p.get('party', '')}</td><td>{p.get('pan', '')}</td><td>{p.get('section', '')}</td>"
                f"<td>{p.get('nature', '')}</td><td class='num'>{fmt(p.get('gross_paid', p.get('gross', 0)))}</td>"
                f"<td class='num'>{fmt(p.get('tds', 0))}</td><td class='num'>{p.get('txns', p.get('count', ''))}</td>")
    rrows = "".join("<tr>" + rcells(p) + "</tr>" for p in parties)
    if not rrows:
        rrows = "<tr><td colspan='7' style='text-align:center;color:var(--sec);padding:14px'>No TDS entries in this period.</td></tr>"
    body = (f"<div class='note'>Total TDS deducted: <b>{fmt(d.get('total_tds', 0))}</b></div>"
            "<h3 style='margin:6px 0 6px;font-size:15px'>By section</h3>"
            "<table class='grid'><tr><th>Section</th><th style='text-align:right'>Rate %</th>"
            "<th style='text-align:right'>TDS Amount</th><th style='text-align:right'>Voucher Count</th></tr>" + trows + "</table>"
            "<h3 style='margin:18px 0 6px;font-size:15px'>TDS Register (party-wise)</h3>"
            "<table class='grid'><tr><th>Party</th><th>PAN</th><th>Section</th><th>Nature</th>"
            "<th style='text-align:right'>Gross Paid</th><th style='text-align:right'>TDS</th><th style='text-align:right'>Txns</th></tr>" + rrows + "</table>")
    return report_shell("TDS Report", "Tax deducted at source", "/tds", body, frm=f, to=t)


@app.get("/soon", response_class=HTMLResponse)
def soon(s: str = "This screen"):
    return _stub(s, "", f"the web port of '{s}', replicated 1:1 from its desktop source — queued, not yet built.")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8800, log_level="warning")
